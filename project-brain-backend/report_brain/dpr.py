"""
report_brain.dpr — DPR (xlsx) extractor with FORMAT FINGERPRINTING.

Each DPR vendor uses a different sheet layout. We fingerprint by
(sheet-name signature + header signature), then apply a column map. First time a
new fingerprint appears, a one-shot mapping is proposed (best-effort header
detection) and can be TAUGHT/corrected on the staging screen; thereafter it is
automatic.

Emits:
  - progress atoms  -> staged to S-curve (activity, uom, cumulative, as-on)
  - status atoms    -> which areas completed / under progress (activity-wise)
  - manpower atoms  -> agency/category counts for monthly averages

Staging: every progress row is returned as a StagingRow (raw name + proposed
plan-activity mapping + confidence) so nothing writes to the S-curve unconfirmed.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field

import openpyxl

from report_brain.atoms import Atom, AliasRegistry, discipline_of


@dataclass
class StagingRow:
    raw_activity: str
    discipline: str
    uom: str
    scope: float | None
    cumulative: float | None
    as_on: str
    proposed_plan_activity: str
    confidence: float
    source_ref: str


@dataclass
class DprResult:
    fingerprint: str
    project_hint: str
    progress: list[StagingRow] = field(default_factory=list)
    status_atoms: list[Atom] = field(default_factory=list)
    manpower_atoms: list[Atom] = field(default_factory=list)


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def fingerprint(wb) -> str:
    sig = "|".join(sorted(s[:14] for s in wb.sheetnames))
    return hashlib.sha1(sig.encode()).hexdigest()[:10]


def _find_header_row(ws, needles=("execution particular", "activity", "activity/area")):
    for r in range(1, min(12, ws.max_row) + 1):
        cells = [str(c.value).lower() if c.value else "" for c in ws[r]]
        joined = " ".join(cells)
        if any(n in joined for n in needles):
            return r, cells
    return None, []


def _col(cells: list[str], *needles) -> int | None:
    for i, c in enumerate(cells):
        if any(n in c for n in needles):
            return i
    return None


def _project_hint(ws, reg: AliasRegistry) -> str:
    blob = " ".join(str(ws.cell(r, c).value or "")
                    for r in range(1, 5) for c in range(1, 6))
    return reg.project_of(blob) or "?"


def extract_dpr(path: str, reg: AliasRegistry, as_on: str = "") -> DprResult:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    fp = fingerprint(wb)
    res = DprResult(fingerprint=fp, project_hint="?")

    # --- progress + area status: prefer a 'Daily Progress Report'/'Daily Report' sheet
    prog_sheet = next((s for s in wb.sheetnames
                       if re.search(r"daily (progress )?report|daily report", s, re.I)), wb.sheetnames[0])
    ws = wb[prog_sheet]
    res.project_hint = _project_hint(ws, reg)
    hr, cells = _find_header_row(ws)
    if hr:
        c_act = _col(cells, "execution particular", "activity/area", "activity")
        c_scope = _col(cells, "scope")
        c_uom = _col(cells, "uom", "unit")
        c_cum = _col(cells, "cum till", "cumulative till", "actual till", "cum.")
        c_wtype = _col(cells, "work type")
        cur_disc = ""
        maxc = (max(x for x in (c_act, c_scope, c_uom, c_cum, c_wtype) if x is not None) + 1)
        for r_idx, rowcells in enumerate(ws.iter_rows(min_row=hr + 1, max_col=maxc, values_only=True), start=hr + 1):
            row = list(rowcells) + [None] * (maxc - len(rowcells))
            wtype = str(row[c_wtype]).strip() if (c_wtype is not None and row[c_wtype]) else ""
            if wtype:
                cur_disc = discipline_of(wtype) or wtype.title()
            name = str(row[c_act]).strip() if (c_act is not None and row[c_act]) else ""
            if not name or name.lower() in ("none", "total"):
                continue
            uom = str(row[c_uom]).strip() if (c_uom is not None and row[c_uom]) else ""
            scope = _num(row[c_scope]) if c_scope is not None else None
            cum = _num(row[c_cum]) if c_cum is not None else None
            r = r_idx
            disc = cur_disc or discipline_of(name)
            mapped = reg.activity_map.get(name.lower().strip(), "")
            res.progress.append(StagingRow(
                raw_activity=name, discipline=disc, uom=uom, scope=scope,
                cumulative=cum, as_on=as_on,
                proposed_plan_activity=mapped or name,
                confidence=1.0 if mapped else (0.55 if scope else 0.4),
                source_ref=f"dpr:{path.split('/')[-1]}:{prog_sheet}:r{r}",
            ))
            # area status atom (completed vs in-progress from scope vs cum)
            state = ""
            if scope and cum is not None:
                state = "completed" if cum >= scope - 1e-6 else ("in_progress" if cum > 0 else "planned")
            res.status_atoms.append(Atom(
                kind="status", date=as_on, project=res.project_hint,
                section_affinity="status", discipline=disc, area=name,
                text=f"{name}: {('completed' if state=='completed' else 'under progress' if state=='in_progress' else 'yet to start')}"
                     + (f" ({cum:g}/{scope:g} {uom})" if (scope and cum is not None) else ""),
                quantities=[{"value": cum, "of": scope, "unit": uom}] if (scope and cum is not None) else [],
                verb_state=state, source_type="dpr",
                source_ref=f"dpr:{path.split('/')[-1]}:{prog_sheet}:r{r}",
            ))

    # --- manpower: DLR sheet (category x skill) or Resource/Weekly manpower
    mp_sheet = next((s for s in wb.sheetnames if re.search(r"dlr|manpower|resource|weekly report", s, re.I)), None)
    if mp_sheet:
        ws2 = wb[mp_sheet]
        hr2, cells2 = _find_header_row(ws2, needles=("category", "agency", "asset description"))
        c_cat = _col(cells2, "category", "agency", "asset description") if cells2 else None
        c_tot = _col(cells2, "total") if cells2 else None
        if hr2 and c_cat is not None:
            maxc2 = ws2.max_column
            for r_idx2, rowcells in enumerate(ws2.iter_rows(min_row=hr2 + 1, max_col=maxc2, values_only=True), start=hr2 + 1):
                row2 = list(rowcells) + [None] * (maxc2 - len(rowcells))
                cat = row2[c_cat]
                if not cat or str(cat).strip().lower() in ("total", "none"):
                    continue
                total = _num(row2[c_tot]) if c_tot is not None else None
                if total is None:  # sum skill columns to the right of category
                    total = sum(v for i in range(c_cat + 1, maxc2)
                                if (v := _num(row2[i])) is not None) or None
                r = r_idx2
                if total:
                    res.manpower_atoms.append(Atom(
                        kind="manpower", date=as_on, project=res.project_hint,
                        section_affinity="manpower", discipline=discipline_of(str(cat)),
                        text=f"{str(cat).strip()}: {total:g}",
                        quantities=[{"value": total, "unit": "nos"}],
                        source_type="dpr",
                        source_ref=f"dpr:{path.split('/')[-1]}:{mp_sheet}:r{r}",
                        extra={"category": str(cat).strip()},
                    ))
    return res
