"""I1 — One-click ministry pack.

Composes the monthly review deliverable from the verified engines:
  Executive Summary  headline metrics + deterministic narrative bullets
                     (+ month-over-month "why changed" when a comparison
                     snapshot is given)
  CAPEX Matrix       the Matrix Engine grid (MoS layout, reused composer)
  EVM Portfolio      per-scheme SPI/CPI/EAC with traffic-light colouring
  Delay Watch        delayed schemes, worst first, days + applicable dates
  Data Quality       every configured check with violation counts
  Reconciliation / Details  (from the matrix composer)

Every figure is engine-calculated; the narrative is deterministic composition
over those figures (spec §18 discipline) — an LLM may later polish wording,
never numbers.
"""
from __future__ import annotations

from datetime import date
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.services import matrix_ai as MA
from app.services import matrix_engine as ME
from app.services import matrix_trust as MT
from app.services.matrix_export import add_report_sheets

FONT = "Arial"
THIN = Side(style="thin", color="9CA3AF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="DCE6F1")
HEALTH_FILL = {"green": PatternFill("solid", fgColor="C6EFCE"),
               "amber": PatternFill("solid", fgColor="FFEB9C"),
               "red": PatternFill("solid", fgColor="FFC7CE"),
               "unknown": PatternFill("solid", fgColor="EFEFEF")}


def _find_row(result: dict, row_id: str) -> Optional[dict]:
    return next((r for r in result["rows"] if r["id"] == row_id), None)


def build_summary(result: dict, evm: Optional[dict], delays: list[dict],
                  dq: dict, change_lines: list[str]) -> dict[str, Any]:
    """Headline metrics + narrative bullets, all engine-derived."""
    top = result["rows"][0] if result["rows"] else None
    cells = (top or {}).get("cells", {})
    be = next((cells[c["key"]] for c in result["columns"]
               if (c.get("measure") or {}).get("field") == "be_fy"), None)
    exp = next((cells[c["key"]] for c in result["columns"]
                if (c.get("measure") or {}).get("field") == "exp_fy"), None)
    util = round(exp / be * 100, 1) if (be and exp is not None) else None

    bullets: list[str] = []
    if top:
        bullets.append(f"{top['scheme_count']} scheme(s) in '{top['name']}'"
                       + (f"; FY expenditure ₹{exp:,.1f} Cr against BE ₹{be:,.1f} Cr "
                          f"— {util}% utilisation." if util is not None else "."))
    if evm and evm.get("schemes"):
        c = evm["counts"]
        worst = [s for s in evm["schemes"] if s["health"] in ("red", "amber")][:3]
        bullets.append(f"EVM health: {c.get('green', 0)} green / {c.get('amber', 0)} amber / "
                       f"{c.get('red', 0)} red / {c.get('unknown', 0)} unknown."
                       + (" Attention: " + "; ".join(
                           f"{s['scheme_name']} (SPI {s['spi']}, CPI {s['cpi']})"
                           for s in worst) + "." if worst else ""))
    if delays:
        w = delays[0]
        bullets.append(f"{len(delays)} scheme(s) running delayed; worst is "
                       f"{w['scheme_name']} at {int(w['delay_days'])} days beyond "
                       f"its applicable completion date.")
    else:
        bullets.append("No scheme is past its applicable completion date.")
    bullets.append(f"Data quality: {dq['error_violations']} error(s), "
                   f"{dq['warning_violations']} warning(s) across "
                   f"{len(dq['checks'])} checks on {dq['population']} records.")
    recon = result.get("reconciliation") or []
    passed = sum(1 for c in recon if c["passed"])
    bullets.append(f"Reconciliation: {passed}/{len(recon)} checks passed.")
    if change_lines:
        bullets.append("Since the compared position: " + " ".join(change_lines[:5]))
    return {"utilisation_pct": util, "be": be, "exp": exp, "bullets": bullets}


def build_pack(db, definition: dict, report_name: str, report_date: date,
               compare_snapshot_result: Optional[dict] = None,
               compare_rules: Optional[dict] = None) -> tuple[Workbook, dict]:
    """Returns (workbook, summary_json)."""
    result = ME.run_report(db, definition, report_date, include_ids=True)
    population = ME.fetch_population(db, report_date, definition.get("dataset"))
    dq = MT.run_dq(db, report_date, definition.get("dataset"))
    delays = sorted((r for r in population if (r.get("delay_days") or 0) > 0),
                    key=lambda r: -r["delay_days"])

    evm = None
    try:
        from app.services import evm_engine as EVM
        evm = EVM.portfolio_evm(db, ME.fy_start_year_of(report_date), today=report_date)
    except Exception:
        evm = None  # EVM optional: pack still composes without physical plans

    change_lines: list[str] = []
    if compare_snapshot_result is not None:
        cmp = MT.compare(compare_snapshot_result, result,
                         compare_rules or {}, ME.load_rules(db),
                         "previous position", "this position")
        change_lines = MA.narrate_comparison(cmp, result["columns"])

    summary = build_summary(result, evm, delays, dq, change_lines)

    # ---------------- workbook ----------------
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Executive Summary")
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"{report_name} — Monthly Review Pack"
    t.font = Font(name=FONT, size=14, bold=True)
    t.alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Position as on {report_date.isoformat()} · FY {result['fy']}"
    ws["A2"].font = Font(name=FONT, size=10, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    r = 4
    for b in summary["bullets"]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=f"•  {b}")
        c.font = Font(name=FONT, size=10.5)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(16, 14 * (1 + len(b) // 95))
        r += 1
    for col, w in zip("ABCDEF", (18, 18, 18, 18, 18, 18)):
        ws.column_dimensions[col].width = w

    add_report_sheets(wb, result, report_name, population, sheet_title="CAPEX Matrix")

    if evm and evm.get("schemes"):
        ws2 = wb.create_sheet("EVM Portfolio")
        heads = ["Scheme", "Health", "BAC", "PV", "EV", "AC", "SPI", "CPI", "EAC", "VAC", "% Compl."]
        for j, h in enumerate(heads, start=1):
            c = ws2.cell(row=1, column=j, value=h)
            c.font = Font(name=FONT, size=10, bold=True)
            c.fill = HEAD_FILL
            c.border = BORDER
        for i, s in enumerate(evm["schemes"], start=2):
            vals = [s["scheme_name"], s["health"].upper(), s.get("bac"), s.get("pv"),
                    s.get("ev"), s.get("ac"), s.get("spi"), s.get("cpi"),
                    s.get("eac"), s.get("vac"), s.get("pct_complete")]
            for j, v in enumerate(vals, start=1):
                c = ws2.cell(row=i, column=j, value=v)
                c.font = Font(name=FONT, size=10)
                c.border = BORDER
                if j == 2:
                    c.fill = HEALTH_FILL[s["health"]]
                if j >= 3:
                    c.number_format = "#,##0.00"
        ws2.column_dimensions["A"].width = 36
        ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Delay Watch")
    heads = ["Scheme", "Category", "Delay (days)", "Applicable Completion",
             "Original Completion", "Revised Completion", "Cost (Cr)"]
    for j, h in enumerate(heads, start=1):
        c = ws3.cell(row=1, column=j, value=h)
        c.font = Font(name=FONT, size=10, bold=True)
        c.fill = HEAD_FILL
        c.border = BORDER
    for i, d in enumerate(delays, start=2):
        vals = [d.get("scheme_name"), d.get("scheme_type"), int(d["delay_days"]),
                str(d.get("applicable_completion") or ""), str(d.get("planned_completion") or ""),
                str(d.get("revised_completion") or ""), d.get("applicable_cost")]
        for j, v in enumerate(vals, start=1):
            c = ws3.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
    ws3.column_dimensions["A"].width = 34
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet("Data Quality")
    for j, h in enumerate(["Check", "Severity", "Violations", "Schemes"], start=1):
        c = ws4.cell(row=1, column=j, value=h)
        c.font = Font(name=FONT, size=10, bold=True)
        c.fill = HEAD_FILL
        c.border = BORDER
    for i, chk in enumerate(dq["checks"], start=2):
        listing = ", ".join(f"{v['scheme_id']} {v['scheme_name']}" for v in chk["violations"][:8])
        for j, v in enumerate([chk["name"], chk["severity"],
                               chk["violation_count"], listing], start=1):
            c = ws4.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=10,
                          color="C00000" if (chk["severity"] == "error"
                                             and (chk["violation_count"] or 0) > 0) else None)
            c.border = BORDER
    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["D"].width = 60

    # review-flow sheet order: summary, matrix, EVM, delays, DQ, recon, details
    desired = ["Executive Summary", "CAPEX Matrix", "EVM Portfolio", "Delay Watch",
               "Data Quality", "Reconciliation", "Details"]
    order = [n for n in desired if n in wb.sheetnames] +             [n for n in wb.sheetnames if n not in desired]
    wb._sheets = [wb[n] for n in order]

    return wb, {"report_date": report_date.isoformat(), "fy": result["fy"],
                "summary": summary, "delay_count": len(delays),
                "evm_counts": (evm or {}).get("counts"),
                "dq": {"errors": dq["error_violations"],
                       "warnings": dq["warning_violations"]},
                "narrative": summary["bullets"]}
