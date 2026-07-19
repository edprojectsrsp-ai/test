"""Matrix Engine M4 — AI assistance (spec §18).

Hard rule: AI never calculates. Everything here produces DRAFTS or
EXPLANATIONS; report values always come from the deterministic engine, and
every draft passes the same validation + preview gate a hand-written rule
would (compile, field check, matching-count preview) before a user can save it.

Four capabilities:
  · draft_rule_from_text : natural language → rule condition JSON.
      A deterministic phrase parser handles the PSU vocabulary (ongoing,
      corporate, plant, delayed by more/less than N year/days, started
      before/during the selected FY, cost above/below N, completed, on time).
      If AI_SERVICE_URL is configured, the LLM is tried first and its output
      is forced through the same schema validation; the parser is the
      always-available fallback, so the feature works with no LLM at all.
  · explain_condition / explain_cell : rule trees rendered as plain English +
      the exact contributing schemes (drilldown made narrative).
  · narrate_comparison : M2 compare output → review-meeting sentences
      ("Cost rose ₹50.0 Cr; scheme 104 moved from X to Y; rule Z changed v1→v2").
  · report_skeleton_from_xlsx : an uploaded management workbook → draft
      report definition (hierarchy from indentation, rules/measures matched
      by name against the libraries) for the user to review and refine.
"""
from __future__ import annotations

import json
import os
import re
from typing import Any, Optional


# ─────────────────────────────────────────── NL → rule (deterministic parser)

_NUM = r"(\d+(?:\.\d+)?)"


def _phrase_rules(text: str) -> list[dict]:
    """Ordered pattern table. Each hit contributes one condition."""
    t = " " + text.lower().strip() + " "
    for word, digit in (("one", "1"), ("two", "2"), ("three", "3"), ("four", "4"),
                        ("five", "5"), ("six", "6"), ("half a", "0.5")):
        t = re.sub(rf"\b{word}\b", digit, t)
    conds: list[dict] = []

    def has(*words):
        return all(re.search(rf"\b{w}\b", t) for w in words)

    # lifecycle
    if has("ongoing"):
        conds.append({"field": "current_status", "op": "=", "value": "ongoing"})
    if has("completed") and not has("not", "completed"):
        conds.append({"field": "current_status", "op": "=", "value": "completed"})
    # category
    if has("corporate"):
        conds.append({"field": "scheme_type", "op": "=", "value": "corporate"})
    if re.search(r"\bplant\b", t):
        conds.append({"field": "scheme_type", "op": "=", "value": "plant"})
    # cost thresholds: "below/under/less than 30 cr", "above/over/more than 100 cr"
    m = re.search(rf"(?:below|under|less than)\s+{_NUM}\s*(?:cr|crore)", t)
    if m:
        conds.append({"field": "applicable_cost", "op": "<", "value": float(m.group(1))})
    m = re.search(rf"(?:above|over|exceeding|more than)\s+{_NUM}\s*(?:cr|crore)", t)
    if m:
        conds.append({"field": "applicable_cost", "op": ">", "value": float(m.group(1))})
    # implementation period
    if re.search(r"(?:started|implemented|being implemented).{0,30}"
                 r"(?:before|prior to|from (?:the )?(?:previous|last))", t):
        conds.append({"field": "effective_start", "op": "<", "value": {"token": "fy_start"}})
    elif re.search(r"started.{0,30}during (?:the )?(?:current|selected|this) (?:financial year|fy)", t):
        conds.append({"field": "effective_start", "op": ">=", "value": {"token": "fy_start"}})
        conds.append({"field": "effective_start", "op": "<=", "value": {"token": "fy_end"}})
    # delay buckets — years then days
    m = re.search(rf"delayed by (?:more than|over|at least)\s+{_NUM}\s*year", t)
    if m:
        conds.append({"field": "delay_days", "op": ">=", "value": float(m.group(1)) * 365})
    m = re.search(rf"delayed by (?:less than|under|below)\s+{_NUM}\s*year", t)
    if m:
        conds.append({"field": "delay_days", "op": ">", "value": 0})
        conds.append({"field": "delay_days", "op": "<", "value": float(m.group(1)) * 365})
    m = re.search(rf"delayed by (?:more than|over)\s+{_NUM}\s*day", t)
    if m:
        conds.append({"field": "delay_days", "op": ">", "value": float(m.group(1))})
    if re.search(r"\bon[- ]?time\b", t):
        conds.append({"field": "delay_days", "op": "=", "value": 0})
    return conds


def draft_rule_from_text(prompt: str, semantic_fields: dict) -> dict[str, Any]:
    """Returns {condition, source, notes}. Raises ValueError if nothing parsed."""
    conds = _phrase_rules(prompt)
    # de-duplicate identical conditions from overlapping phrasings
    seen, unique = set(), []
    for c in conds:
        key = json.dumps(c, sort_keys=True)
        if key not in seen:
            seen.add(key)
            unique.append(c)
    if not unique:
        raise ValueError(
            "Could not derive any condition from that phrasing. Recognised "
            "vocabulary: ongoing/completed, corporate/plant, cost above/below "
            "N Cr, started before / during the selected FY, on time, delayed "
            "by more/less than N years (or N days).")
    return {"condition": {"op": "AND", "conditions": unique},
            "source": "deterministic_parser",
            "notes": f"{len(unique)} condition(s) derived — review, preview and "
                     "save to version it. The engine, not the AI, computes results."}


def maybe_llm_draft(prompt: str, semantic_fields: dict) -> Optional[dict]:
    """Optional LLM transport (AI service). Output must parse as a condition
    tree or it is discarded in favour of the deterministic parser."""
    url = os.environ.get("AI_SERVICE_URL")
    if not url:
        return None
    try:
        import httpx
        fields_doc = ", ".join(f"{k}({v['type']})" for k, v in semantic_fields.items())
        r = httpx.post(f"{url.rstrip('/')}/v1/complete", timeout=20, json={
            "system": ("Convert the user's report-population description into a "
                       "JSON condition tree {\"op\":\"AND\",\"conditions\":[...]} "
                       "where each condition is {field, op, value} using ONLY these "
                       f"fields: {fields_doc}. Ops: = != > >= < <= between in "
                       "is_null not_null. Period tokens as value: "
                       "{\"token\":\"fy_start\"} etc. Respond with JSON only."),
            "prompt": prompt})
        txt = r.json().get("text", "")
        cond = json.loads(re.sub(r"^```json|```$", "", txt.strip(), flags=re.M))
        if isinstance(cond, dict) and "conditions" in cond:
            return {"condition": cond, "source": "ai_service",
                    "notes": "LLM draft — validated deterministically before save."}
    except Exception:
        return None
    return None


# ─────────────────────────────────────────── condition → English

_OP_EN = {"=": "is", "!=": "is not", ">": "is greater than", ">=": "is at least",
          "<": "is less than", "<=": "is at most", "between": "is between",
          "in": "is one of", "not_in": "is not one of", "contains": "contains",
          "is_null": "is blank", "not_null": "is not blank"}
_TOK_EN = {"fy_start": "the start of the selected FY",
           "fy_end": "the end of the selected FY",
           "prev_fy_start": "the start of the previous FY",
           "report_date": "the reporting date",
           "one_year_before_report": "one year before the reporting date"}


def _val_en(v: Any) -> str:
    if isinstance(v, dict) and "token" in v:
        return _TOK_EN.get(v["token"], v["token"])
    if isinstance(v, list):
        return " and ".join(_val_en(x) for x in v)
    return str(v)


def explain_condition(cond: Optional[dict], rules: dict, semantic: dict,
                      depth: int = 0) -> str:
    if not cond:
        return "all records"
    if "rule" in cond:
        r = rules.get(cond["rule"])
        inner = explain_condition(r["condition"], rules, semantic, depth + 1) if r else "?"
        return f"[{(r or {}).get('name', cond['rule'])}: {inner}]"
    if "conditions" in cond:
        joiner = {"AND": " and ", "OR": " or ", "NOT": " and not "}[
            (cond.get("op") or "AND").upper()]
        parts = [explain_condition(c, rules, semantic, depth + 1)
                 for c in cond["conditions"]]
        body = joiner.join(parts) if parts else "all records"
        return f"({body})" if depth and len(parts) > 1 else body
    label = semantic.get(cond.get("field"), {}).get("label", cond.get("field"))
    op = _OP_EN.get(cond.get("op"), cond.get("op"))
    if cond.get("op") in ("is_null", "not_null"):
        return f"{label} {op}"
    return f"{label} {op} {_val_en(cond.get('value'))}"


def explain_cell(drill: dict, rules: dict, semantic: dict) -> str:
    """Drilldown payload → plain-English paragraph (deterministic; §5.9 + §18)."""
    conds = explain_condition(drill["effective_conditions"], rules, semantic)
    names = [s["scheme_name"] for s in drill["schemes"][:8]]
    more = drill["qualifying_count"] - len(names)
    listing = ", ".join(str(n) for n in names) + (f" and {more} more" if more > 0 else "")
    return (f"'{drill['column']}' for '{drill['row']}' is {drill['value']}. "
            f"A scheme qualifies when {conds}. "
            f"{drill['qualifying_count']} scheme(s) qualify: {listing}.")


# ─────────────────────────────────────────── comparison → narrative

def narrate_comparison(cmp: dict, columns: list[dict]) -> list[str]:
    unit_of = {c["key"]: c.get("unit", "") for c in columns}
    name_of = {c["key"]: c["name"] for c in columns}
    lines: list[str] = []
    for row in cmp.get("changed_rows", []):
        moves = []
        if row.get("entered"):
            moves.append(f"scheme(s) {', '.join(map(str, row['entered']))} entered")
        if row.get("left"):
            moves.append(f"scheme(s) {', '.join(map(str, row['left']))} left")
        deltas = []
        for k, d in row["cells"].items():
            if d.get("delta") not in (None, 0):
                sign = "+" if d["delta"] > 0 else ""
                deltas.append(f"{name_of.get(k, k)} {sign}{d['delta']} {unit_of.get(k, '')}".strip())
        if deltas or moves:
            why = f" because {' and '.join(moves)}" if moves else ""
            what = "; ".join(deltas) if deltas else "membership changed"
            lines.append(f"'{row['name']}': {what}{why}.")
    for rc in cmp.get("rule_version_changes", []):
        vals = [f"{k} {v}" for k, v in rc.items() if k != "rule_key"]
        lines.append(f"Rule '{rc['rule_key']}' changed between the two sides "
                     f"({', '.join(map(str, vals))}) — classification logic differs.")
    if not lines:
        lines.append("No cell values, memberships or rule versions differ.")
    return lines


# ─────────────────────────────────────────── Excel → report skeleton

def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _match_key(name: str, library: dict[str, str]) -> Optional[str]:
    """library: key -> display name. Exact-norm first, then containment."""
    n = _norm(name)
    if not n:
        return None
    for key, disp in library.items():
        if _norm(disp) == n or _norm(key) == n:
            return key
    for key, disp in library.items():
        d = _norm(disp)
        if d and (d in n or n in d):
            return key
    # last pass: the KEY itself inside the name (min length guards noise)
    for key in library:
        k = _norm(key)
        if len(k) >= 5 and k in n:
            return key
    return None


def report_skeleton_from_xlsx(path: str, rule_names: dict[str, str],
                              measure_names: dict[str, str]) -> dict[str, Any]:
    """Parse a management workbook: header row = first row where ≥2 trailing
    cells are non-empty after a text first cell; hierarchy from cell
    indentation (or leading spaces); rules/measures matched by name."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)))

    header_i = None
    for i, r in enumerate(rows):
        vals = [c.value for c in r]
        if vals and isinstance(vals[0], str) and sum(1 for v in vals[1:] if v not in (None, "")) >= 2:
            nxt = [c.value for c in rows[i + 1]] if i + 1 < len(rows) else []
            if any(isinstance(v, (int, float)) for v in nxt[1:]):
                header_i = i
                break
    if header_i is None:
        raise ValueError("Could not locate a header row (title + ≥2 columns "
                         "followed by numeric data)")

    columns, unmatched_cols = [], []
    for j, cell in enumerate(rows[header_i][1:], start=1):
        if cell.value in (None, ""):
            continue
        name = str(cell.value).replace("\n", " ").strip()
        mk = _match_key(name, measure_names)
        col = {"key": f"c{j}", "name": name}
        if mk:
            col["measure_key"] = mk
        else:
            col["measure"] = {"field": "scheme_id", "agg": "count_distinct"}
            unmatched_cols.append(name)
        columns.append(col)

    tree: list[dict] = []
    stack: list[tuple[int, dict]] = []          # (depth, node)
    unmatched_rows = []
    idx = 0
    for r in rows[header_i + 1:]:
        raw = r[0].value
        if raw in (None, ""):
            continue
        name = str(raw).strip()
        if not name or _norm(name).startswith("systemcalculated"):
            continue
        indent = int(r[0].alignment.indent or 0)
        if indent == 0 and raw != name:          # fallback: leading spaces
            indent = (len(str(raw)) - len(str(raw).lstrip())) // 2
        idx += 1
        node: dict[str, Any] = {"id": f"r{idx}", "name": name, "children": []}
        rk = _match_key(name, rule_names)
        if rk:
            node["rule"] = rk
        else:
            unmatched_rows.append(name)
        while stack and stack[-1][0] >= indent:
            stack.pop()
        if stack:
            stack[-1][1]["children"].append(node)
        else:
            tree.append(node)
        stack.append((indent, node))

    return {"definition": {"columns": columns, "rows": tree},
            "unmatched_rows": unmatched_rows,
            "unmatched_columns": unmatched_cols,
            "notes": "Draft only — unmatched rows have no rule and count "
                     "columns default to scheme count. Review, assign, preview."}
