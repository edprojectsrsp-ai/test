"""Matrix report → Excel in the MoS management-report layout (spec §14).

Renders a run result (live or frozen snapshot) into a workbook that visually
matches the approved ministry format: merged title band, bold bordered header,
hierarchy expressed through indentation and weight, Indian number formats,
reconciliation sheet, and a drill-ready "Details" sheet listing the population.

The workbook contains system-calculated VALUES (a rendering of an approved
calculation, like the submitted PMC file), not formulas — the deterministic
engine is the calculator of record and every figure remains drillable in the
app. Provenance is stamped in the footer.
"""
from __future__ import annotations

from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
THIN = Side(style="thin", color="9CA3AF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="DCE6F1")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
NUM_FMT = "#,##0.00;(#,##0.00);\"-\""
INT_FMT = "#,##0;(#,##0);\"-\""


def build_workbook(result: dict[str, Any], report_name: str,
                   population: list[dict] | None = None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    cols = result["columns"]
    ncols = 1 + len(cols)

    # ── title band (merged, like the MoS sheet)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=report_name)
    t.font = Font(name=FONT, size=13, bold=True)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sub = ws.cell(row=2, column=1,
                  value=f"Position as on {result['report_date']}   ·   FY {result['fy']}")
    sub.font = Font(name=FONT, size=10, italic=True)
    sub.alignment = Alignment(horizontal="center")

    # ── header row
    hr = 4
    ws.cell(row=hr, column=1, value="Category")
    for j, c in enumerate(cols, start=2):
        label = c["name"] + (f"\n({c['unit']})" if c.get("unit") else "")
        ws.cell(row=hr, column=j, value=label)
    for j in range(1, ncols + 1):
        cell = ws.cell(row=hr, column=j)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.fill = HEAD_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hr].height = 30

    # ── body: indentation + weight express the hierarchy
    r = hr
    for row in result["rows"]:
        r += 1
        depth = row.get("depth", 0)
        name_cell = ws.cell(row=r, column=1, value=row["name"])
        name_cell.font = Font(name=FONT, size=10, bold=depth <= 1,
                              italic=row.get("type") == "formula")
        name_cell.alignment = Alignment(indent=depth, vertical="center")
        name_cell.border = BORDER
        if depth == 0:
            name_cell.fill = TOTAL_FILL
        for j, c in enumerate(cols, start=2):
            v = row["cells"].get(c["key"])
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = Font(name=FONT, size=10, bold=depth == 0)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
            is_count = "count" in str((c.get("measure") or {}).get("agg", ""))
            dec = c.get("decimals")
            if dec == 0 or is_count:
                cell.number_format = INT_FMT
            elif dec is not None:
                z = "0" * dec
                cell.number_format = f"#,##0.{z};(#,##0.{z});\"-\""
            else:
                cell.number_format = NUM_FMT
            if depth == 0:
                cell.fill = TOTAL_FILL

    # provenance footer (assumption/source documentation)
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    note = ws.cell(row=r, column=1, value=(
        "System-calculated by Project Brain Matrix Engine — every figure is "
        "drillable to contributing schemes in the application. Values rendered "
        f"as calculated for the {result['report_date']} position; rule versions "
        "recorded in the snapshot."))
    note.font = Font(name=FONT, size=8, italic=True, color="6B7280")

    # widths
    ws.column_dimensions["A"].width = 44
    for j in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.freeze_panes = f"B{hr + 1}"

    # ── reconciliation sheet
    ws2 = wb.create_sheet("Reconciliation")
    heads = ["Check", "Type", "Result", "Detail", "Left/Parent", "Right/Children"]
    for j, h in enumerate(heads, start=1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font = Font(name=FONT, size=10, bold=True)
        c.fill = HEAD_FILL
        c.border = BORDER
    for i, chk in enumerate(result.get("reconciliation") or [], start=2):
        vals = [chk["parent"], chk["type"], "PASS" if chk["passed"] else "FAIL",
                chk["detail"], chk.get("parent_count"), chk.get("children_union_count")]
        for j, v in enumerate(vals, start=1):
            c = ws2.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=10,
                          color=None if chk["passed"] else "C00000",
                          bold=(j == 3 and not chk["passed"]))
            c.border = BORDER
    for j, w in enumerate([34, 20, 8, 50, 14, 14], start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ── details sheet (supporting population — spec §14 drill workbook)
    if population:
        ws3 = wb.create_sheet("Details")
        keys = list(population[0].keys())
        for j, k in enumerate(keys, start=1):
            c = ws3.cell(row=1, column=j, value=k)
            c.font = Font(name=FONT, size=9, bold=True)
            c.fill = HEAD_FILL
            c.border = BORDER
        for i, rec in enumerate(population, start=2):
            for j, k in enumerate(keys, start=1):
                v = rec[k]
                if isinstance(v, date):
                    v = v.isoformat()
                ws3.cell(row=i, column=j, value=v).font = Font(name=FONT, size=9)
        ws3.freeze_panes = "A2"

    return wb
