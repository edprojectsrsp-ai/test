"""Contractor DPR file ingestion — parse the site's heterogeneous daily-report
Excel formats into one normalized shape and auto-match rows to plan_activities.

Three real format families (fixtures: Downloads/DPR 1.07.2026.zip):

  A. "rsp_dpr"        — sheet 'Daily Progress Report' (172_RSP_DPR / "by product"):
     WorkType | Sr | Execution Particulars | Scope | UOM | ActTillLastFY |
     CumTillLastMonth P/A | FTM Plan | TillYesterday P/A | ForTheDay P/A |
     CumMonth P/A | CumTotal P/A · then MANPOWER and CONSTRUCTION EQUIPMENT blocks.
  B. "weekly"         — sheet 'Weekly Report' (C.GWZ7 4th Stove):
     Activity | Scope | UoM | CumUpToDate P/A | CumTillLastMonth P/A |
     MonthPlan | ActualTillDate | NextPlan | Remarks · section header rows.
  C. "site_progress"  — sheet 'Daily Report' (L&T Battery Proper):
     Sl | Activity/Area | Scope | DrawingAvl | CumTillLastMonth | PlanFTM |
     CumPrevDay | FTD P/A | NextDayPlan | CumThisMonth | CumAsOnDate | Remarks
     · lettered sections ("A. Piling Work"), 'Resource' sheet for equipment.

Matching: work-type/section + discipline keywords + fuzzy name similarity
against the scheme's locked+current plan activities, with top-3 candidates and
a confidence score — the UI lets the analyst override before commit.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from difflib import SequenceMatcher

import openpyxl
from sqlalchemy import text


# ─────────────────────────── helpers ────────────────────────────────────────

def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").strip().lower()).strip()


def _f(v):
    try:
        if v is None or str(v).strip() == "":
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _cell(row, i):
    return row[i] if i < len(row) else None


def _text(row, i):
    v = _cell(row, i)
    return str(v).strip() if v is not None else ""


def _is_section(label: str) -> bool:
    """Group/section rows: ALL-CAPS work types or lettered headings."""
    t = label.strip()
    if not t:
        return False
    return bool(re.match(r"^[A-Z]\.?\s", t)) or (t.upper() == t and len(t) > 3)


ROMAN = {"i", "ii", "iii", "iv", "v", "vi", "vii", "viii", "ix", "x", "misc"}


# ─────────────────────────── format detection ───────────────────────────────

def detect_format(wb) -> str:
    names = [n.strip().lower() for n in wb.sheetnames]
    if "daily report" in names and "resource" in names:
        return "site_progress"
    if "weekly report" in names:
        return "weekly"
    if "daily progress report" in names:
        return "rsp_dpr"
    raise ValueError(f"Unrecognized DPR workbook (sheets: {wb.sheetnames})")


def _find_header(rows, must=("scope", "uom")):
    for i, row in enumerate(rows[:12]):
        joined = _norm(" ".join(str(c) for c in row if c is not None))
        if all(k in joined for k in must):
            return i
    return None


# ─────────────────────────── format A: rsp_dpr ──────────────────────────────

def _parse_rsp_dpr(wb):
    ws = next(s for s in wb.worksheets if s.title.strip().lower() == "daily progress report")
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    project = ""
    for r in rows[:4]:
        t = " ".join(str(c) for c in r if c)
        m = re.search(r"project\s*name\s*:?\s*(.+)", t, re.I)
        if m:
            project = m.group(1).split("For ")[0].strip()
            break
    hi = _find_header(rows)
    activities, manpower, equipment = [], [], []
    mode = "activities"
    work_type = ""
    parent = None
    _base = (hi or 3) + 2
    for _ri, r in enumerate(rows[_base:]):
        _src_row = _base + _ri + 1  # 1-based Excel row
        c0, c1, c2 = _text(r, 0), _text(r, 1), _text(r, 2)
        label = c2 or c0
        up = (c0 + " " + c2).upper()
        if "MANPOWER" in up and not _f(_cell(r, 3)):
            mode = "manpower"
        if "CONSTRUCTION EQUIPMENT" in up or "EQUIPMENT" == c0.upper():
            mode = "equipment"
        if mode == "manpower":
            qty = _f(_cell(r, 6))
            name = (c2 or c1 or c0).strip()
            if name and qty is not None and "total" not in name.lower():
                trade = _text(r, 3)
                manpower.append({"category": name, "trade": trade, "ftd": qty})
            if "CONSTRUCTION" in up:
                mode = "equipment"
                continue
            continue
        if mode == "equipment":
            qty = _f(_cell(r, 6))
            name = (c2 or c0).strip()
            if name and qty is not None and "EQUIPMENT" not in name.upper():
                equipment.append({"name": name, "count": qty})
            continue
        # activities
        if c0 and not _f(_cell(r, 3)) and not c2:
            continue
        if c0 and c0.upper() == c0 and len(c0) > 3:
            work_type = c0.title()
        if not (c2 or c1):
            continue
        scope, uom = _f(_cell(r, 3)), _text(r, 4)
        is_detail = _norm(c1) in ROMAN or bool(re.match(r"^\d+$", c1))
        entry = {
            "workType": work_type,
            "activity": c2 if not is_detail else (parent or work_type),
            "area": c2 if is_detail else "",
            "kind": "detail" if is_detail else "group",
            "scope": scope, "uom": uom,
            "actualTillLastFy": _f(_cell(r, 5)),
            "cumPlanLastMonth": _f(_cell(r, 6)), "cumActualLastMonth": _f(_cell(r, 7)),
            "ftmPlan": _f(_cell(r, 8)),
            "dayPlan": _f(_cell(r, 11)), "dayActual": _f(_cell(r, 12)),
            "cumPlanToDate": _f(_cell(r, 15)), "cumActualToDate": _f(_cell(r, 16)),
            "_srcRow": _src_row,
            "_colMap": {"scope": 3, "uom": 4, "actualTillLastFy": 5,
                        "cumPlanLastMonth": 6, "cumActualLastMonth": 7, "ftmPlan": 8,
                        "dayPlan": 11, "dayActual": 12, "cumPlanToDate": 15,
                        "cumActualToDate": 16},
            "remarks": "",
        }
        if not is_detail and c2:
            parent = c2
        if entry["scope"] is not None or entry["dayActual"] is not None or entry["cumActualToDate"] is not None:
            activities.append(entry)
    return {"format": "rsp_dpr", "projectName": project,
            "activities": activities, "manpower": manpower, "equipment": equipment}


# ─────────────────────────── format B: weekly ───────────────────────────────

def _parse_weekly(wb):
    ws = next(s for s in wb.worksheets if s.title.strip().lower() == "weekly report")
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    project = ""
    report_date = None
    for r in rows[:6]:
        t = " ".join(str(c) for c in r if c)
        m = re.search(r"name of project\s*:?\s*(.+)", t, re.I)
        if m:
            project = m.group(1).strip()
        for c in r:
            if isinstance(c, datetime):
                report_date = report_date or c.date()
    hi = _find_header(rows, must=("scope", "uom")) or _find_header(rows, must=("scope", "uo"))
    activities, manpower = [], []
    work_type = ""
    for r in rows[(hi or 5) + 3:]:
        c0 = _text(r, 0)
        if not c0:
            continue
        scope, uom = _f(_cell(r, 1)), _text(r, 2)
        if scope is None and not uom:
            work_type = re.sub(r"\(.*?\)", "", c0).strip().title()
            continue
        # manpower rows embedded in the weekly activity grid
        if re.search(r"manpower|officers|staff deployed|workmen", f"{work_type} {c0}", re.I):
            qty = _f(_cell(r, 8)) or _f(_cell(r, 4))
            if qty is not None:
                manpower.append({"category": c0, "trade": "", "ftd": qty})
            continue
        activities.append({
            "workType": work_type, "activity": c0, "area": "", "kind": "group",
            "scope": scope, "uom": uom,
            "actualTillLastFy": None,
            "cumPlanLastMonth": _f(_cell(r, 5)), "cumActualLastMonth": _f(_cell(r, 6)),
            "ftmPlan": _f(_cell(r, 7)),
            "dayPlan": None, "dayActual": _f(_cell(r, 8)),
            "cumPlanToDate": _f(_cell(r, 3)), "cumActualToDate": _f(_cell(r, 4)),
            "remarks": _text(r, 10),
        })
    return {"format": "weekly", "projectName": project,
            "reportDate": report_date.isoformat() if report_date else None,
            "activities": activities, "manpower": manpower, "equipment": []}


# ─────────────────────────── format C: site_progress ────────────────────────

def _parse_site_progress(wb):
    ws = next(s for s in wb.worksheets if s.title.strip().lower() == "daily report")
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    project = ""
    for r in rows[:5]:
        t = [str(c) for c in r if c]
        for j, c in enumerate(t):
            if "name of the project" in c.lower() and j + 1 < len(t):
                project = t[j + 1]
    hi = _find_header(rows, must=("scope", "activity"))
    activities = []
    work_type = ""
    for r in rows[(hi or 4) + 2:]:
        c0, c1 = _text(r, 0), _text(r, 1)
        if not c0 and c1 and _f(_cell(r, 2)) is None:
            work_type = re.sub(r"^[A-Z]\.?\s*", "", re.sub(r"\(.*?\)", "", c1)).strip().title()
            continue
        if not c1:
            continue
        scope = _f(_cell(r, 2))
        if scope is None and _f(_cell(r, 11)) is None:
            continue
        activities.append({
            "workType": work_type, "activity": work_type or c1, "area": c1, "kind": "detail",
            "scope": scope, "uom": "",
            "actualTillLastFy": None,
            "cumPlanLastMonth": None, "cumActualLastMonth": _f(_cell(r, 4)),
            "ftmPlan": _f(_cell(r, 5)),
            "dayPlan": _f(_cell(r, 7)), "dayActual": _f(_cell(r, 8)),
            "cumPlanToDate": None, "cumActualToDate": _f(_cell(r, 11)),
            "remarks": _text(r, 12),
        })
    # equipment from Resource sheet
    equipment = []
    res = next((s for s in wb.worksheets if s.title.strip().lower() == "resource"), None)
    if res is not None:
        for r in res.iter_rows(values_only=True):
            r = list(r)
            name, qty = _text(r, 1), _f(_cell(r, 3))
            if name and qty is not None and "description" not in name.lower():
                equipment.append({"name": name, "count": qty})
    return {"format": "site_progress", "projectName": project,
            "activities": activities, "manpower": [], "equipment": equipment}


def parse_dpr_file(content: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    fmt = detect_format(wb)
    if fmt == "rsp_dpr":
        return _parse_rsp_dpr(wb)
    if fmt == "weekly":
        return _parse_weekly(wb)
    return _parse_site_progress(wb)


# ─────────────────────────── auto-matching ──────────────────────────────────

# discipline keywords → the discipline expected in the plan-activity name
_DISCIPLINE_HINTS = [
    (("excavation", "pcc", "rcc", "civil", "piling", "pile", "concrete", "foundation"), ("civil",)),
    (("structural", "steel"), ("steel", "structur")),
    (("piping", "pipe"), ("piping", "pipe")),
    (("equipment", "mechanical"), ("mechanical", "equipment")),
    (("electrical", "cable"), ("electric",)),
    (("refractory", "refractories"), ("refractor",)),
    (("engineering", "design", "drawing"), ("design", "engineering")),
    (("commissioning", "testing"), ("commission", "testing")),
]

_STAGE_HINTS = [
    (("erection", "erect", "install"), "erection"),
    (("supply", "delivery", "dispatch", "receipt"), "supply"),
    (("design", "engineering", "drawing"), "design"),
]


def _hints(label: str):
    t = _norm(label)
    disc = set()
    for keys, targets in _DISCIPLINE_HINTS:
        if any(k in t for k in keys):
            disc.update(targets)
    stage = ""
    for keys, s in _STAGE_HINTS:
        if any(k in t for k in keys):
            stage = s
            break
    return disc, stage


def match_activities(db, scheme_id: int, parsed_rows, package_id=None):
    """Score each parsed group row against the scheme's current plan activities."""
    pkg_sql = "AND pkg.package_id = :pkg" if package_id else ""
    plan_acts = [dict(r) for r in db.execute(text(f"""
        SELECT pa.activity_id, pa.activity_name, COALESCE(pa.activity_category,'') AS category,
               COALESCE(pa.scope_qty,0) AS scope_qty, COALESCE(um.uom_code,'') AS uom,
               pkg.package_id, pkg.package_name
        FROM plan_activities pa
        JOIN progress_plans pp ON pp.plan_id = pa.plan_id
        JOIN packages pkg      ON pkg.package_id = pp.package_id
        LEFT JOIN uom_master um ON um.uom_id = pa.uom_id
        WHERE pkg.scheme_id = :sid AND pp.is_locked AND pp.is_current
          AND NOT pp.is_deleted AND NOT pa.is_deleted {pkg_sql}
        ORDER BY pkg.package_id, pa.sort_order
    """), {"sid": scheme_id, "pkg": package_id}).mappings().all()]

    def score(row, act):
        label = f'{row.get("workType", "")} {row.get("activity", "")}'
        target = f'{act["category"]} {act["activity_name"]}'
        s = SequenceMatcher(None, _norm(label), _norm(target)).ratio() * 40
        disc, stage = _hints(label)
        tnorm = _norm(target)
        if disc and any(d in tnorm for d in disc):
            s += 35
        elif disc:
            s -= 15
        if stage:
            if stage in tnorm:
                s += 20
            elif stage == "supply" and "erection" in tnorm or stage == "erection" and "supply" in tnorm:
                s -= 25
        # UOM agreement is a strong signal
        if row.get("uom") and act["uom"] and _norm(row["uom"])[:3] == _norm(act["uom"])[:3]:
            s += 10
        # scope proximity (same order of magnitude)
        ps, qs = row.get("scope"), float(act["scope_qty"] or 0)
        if ps and qs:
            ratio = min(ps, qs) / max(ps, qs)
            s += 8 * ratio
        return max(0.0, min(100.0, s))

    out = []
    for i, row in enumerate(parsed_rows):
        cands = sorted(
            ({"activity_id": a["activity_id"], "activity_name": a["activity_name"],
              "category": a["category"], "package": a["package_name"],
              "uom": a["uom"], "scope": float(a["scope_qty"] or 0),
              "confidence": round(score(row, a), 1)} for a in plan_acts),
            key=lambda c: c["confidence"], reverse=True)[:3]
        best = cands[0] if cands and cands[0]["confidence"] >= 45 else None
        out.append({"rowIndex": i, "candidates": cands,
                    "matchedActivityId": best["activity_id"] if best else None,
                    "confidence": best["confidence"] if best else 0.0})
    return out
