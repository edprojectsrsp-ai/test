from tkinter import *
from tkinter import ttk, messagebox, filedialog
import csv
from database import get_all_projects, add_project, add_child_project, get_db_connection, delete_project_everywhere, update_project_stage
from utils import generate_unique_id, get_project_status, normalize_buttons, keep_window_active, to_display_date, to_storage_date, parse_app_date, apply_page_watermark
from execution import ScrollableFrame   # Re-using existing ScrollableFrame
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

class RegistrationFrame(Frame):
    BUTTON_FONT = ("Arial", 9, "bold")
    BUTTON_WIDTH = 16
    BUTTON_HEIGHT = 1
    TABLE_COLUMNS = ("SR", "UID", "Name", "GrossCost", "REG_DATE", "Status", "Complete", "CompleteDate", "Commissioned", "CommissionedDate")
    TABLE_HEADINGS = {
        "SR": "Sr. No.",
        "UID": "Unique ID",
        "Name": "Project Name",
        "GrossCost": "Gross Cost",
        "REG_DATE": "Date of Registration",
        "Status": "Current Status",
        "Complete": "Complete",
        "CompleteDate": "Complete Date",
        "Commissioned": "Commissioned",
        "CommissionedDate": "Commissioned Date",
    }

    def __init__(self, parent, main_app=None):
        super().__init__(parent)
        self.main_app = main_app
        
        # ==================== SCROLLABLE PAGE WITH STEEL PLANT THEME ====================
        self.scroll_frame = ScrollableFrame(self, bg="#f0f4f8")   # Steel Plant Theme Background
        self.scroll_frame.pack(fill=BOTH, expand=True)
        main = self.scroll_frame.scrollable_frame
        # ===================================================================

        # ==================== CENTERED CONTENT (Right Side) ====================
        content_frame = Frame(main, bg="#f0f4f8")
        content_frame.pack(anchor="center", padx=25, pady=20, fill=BOTH, expand=True)
        self.content_frame = content_frame
        self.table_rows = {"corporate": [], "plant": []}
        self.table_search_vars = {}
        self.table_filter_vars = {}
        self.active_project_tree = None
        self._pending_tree_yview = {}
        self._pending_page_yview = None
        # ===================================================================

        self.configure_tree_style()

        Label(content_frame, text="Project Registration (Tabular Form)", 
              font=("Arial", 18, "bold"), bg="#f0f4f8").pack(pady=15)

        # Form Section - Centered
        f = Frame(content_frame, bg="#f0f4f8")
        f.pack(pady=10)
        
        Label(f, text="Project Type:", bg="#f0f4f8").grid(row=0, column=0, padx=8, sticky=W)
        self.type_var = StringVar()
        ttk.Combobox(f, textvariable=self.type_var, values=["Corporate AMR", "Plant Level AMR"], width=28).grid(row=0, column=1, padx=8)
        
        Label(f, text="Project Name:", bg="#f0f4f8").grid(row=1, column=0, padx=8, sticky=W)
        self.name_entry = Entry(f, width=32)
        self.name_entry.grid(row=1, column=1, padx=8)
        
        edit_state = NORMAL if not self.main_app or self.main_app.can_edit("registration") else DISABLED
        self.register_btn = Button(f, text="Register Project", command=self.register,
                                   bg="#003087", fg="white", font=("Arial", 11, "bold"),
                                   width=18, height=2, state=edit_state)
        self.register_btn.grid(row=2, column=1, pady=20)

        upload_frame = Frame(content_frame, bg="#f0f4f8")
        upload_frame.pack(pady=(5, 15))
        admin_template_state = NORMAL if (not self.main_app or self.main_app.is_admin()) else DISABLED

        self.template_btn = Button(
            upload_frame,
            text="Download Template",
            command=self.download_template,
            bg="#0f766e",
            fg="white",
            font=("Arial", 11, "bold"),
            state=admin_template_state,
        )
        self.template_btn.pack(side=LEFT, padx=10)
        self.upload_btn = Button(
            upload_frame,
            text="Upload",
            command=self.upload_sheet,
            bg="#7c3aed",
            fg="white",
            font=("Arial", 11, "bold"),
            state=admin_template_state,
        )
        self.upload_btn.pack(side=LEFT, padx=10)
        Label(
            upload_frame,
            text="Template includes project type, name, and stage-wise fields",
            bg="#f0f4f8",
            fg="#4b5563",
            font=("Arial", 10, "italic"),
        ).pack(side=LEFT, padx=12)

        # Corporate AMR Table - Centered & Expanded
        Label(content_frame, text="Corporate AMR Projects", 
              font=("Arial", 14, "bold"), fg="#003087", bg="#f0f4f8").pack(pady=(25, 8), anchor="w")
        self.create_table_search(content_frame, "corporate")
        
        tree_frame1 = Frame(content_frame, bg="#f0f4f8", height=240)
        tree_frame1.pack(fill=X, expand=True, padx=10, pady=5, anchor="w")
        tree_frame1.pack_propagate(False)
        tree_frame1.grid_rowconfigure(0, weight=1)
        tree_frame1.grid_columnconfigure(0, weight=1)
        self.tree_frame1 = tree_frame1
        
        self.tree_corporate = ttk.Treeview(
            tree_frame1,
            columns=self.TABLE_COLUMNS,
            show="headings",
            height=9,
            style="Registration.Treeview",
        )
        self.configure_table_headings(self.tree_corporate)
        self.tree_corporate.column("SR", width=80, anchor="center", stretch=False)
        self.tree_corporate.column("UID", width=230, anchor="center", stretch=False)
        self.tree_corporate.column("Name", width=360, anchor="w", stretch=False)
        self.tree_corporate.column("GrossCost", width=130, anchor="center", stretch=False)
        self.tree_corporate.column("REG_DATE", width=150, anchor="center", stretch=False)
        self.tree_corporate.column("Status", width=170, anchor="center", stretch=False)
        self.tree_corporate.column("Complete", width=110, anchor="center", stretch=False)
        self.tree_corporate.column("CompleteDate", width=130, anchor="center", stretch=False)
        self.tree_corporate.column("Commissioned", width=130, anchor="center", stretch=False)
        self.tree_corporate.column("CommissionedDate", width=150, anchor="center", stretch=False)
        self.tree_corporate.grid(row=0, column=0, sticky="nsew")

        scrollbar1 = ttk.Scrollbar(tree_frame1, orient="vertical", command=self.tree_corporate.yview)
        scrollbar1.grid(row=0, column=1, sticky="ns")
        hscrollbar1 = Scrollbar(tree_frame1, orient="horizontal", command=self.tree_corporate.xview, width=18)
        hscrollbar1.grid(row=1, column=0, sticky="ew", pady=(4, 0))
        self.tree_corporate.configure(yscrollcommand=scrollbar1.set, xscrollcommand=hscrollbar1.set)
        self.bind_project_tree(self.tree_corporate)
        self.tree_corporate.bind("<ButtonRelease-1>", self.on_tree_click)
        self.tree_corporate.bind("<Double-1>", self.on_double_click)

        # Plant Level AMR Table - Centered & Expanded
        Label(content_frame, text="Plant Level AMR Projects", 
              font=("Arial", 14, "bold"), fg="#003087", bg="#f0f4f8").pack(pady=(25, 8), anchor="w")
        self.create_table_search(content_frame, "plant")
        
        tree_frame2 = Frame(content_frame, bg="#f0f4f8", height=240)
        tree_frame2.pack(fill=X, expand=True, padx=10, pady=5, anchor="w")
        tree_frame2.pack_propagate(False)
        tree_frame2.grid_rowconfigure(0, weight=1)
        tree_frame2.grid_columnconfigure(0, weight=1)
        self.tree_frame2 = tree_frame2
        
        self.tree_plant = ttk.Treeview(
            tree_frame2,
            columns=self.TABLE_COLUMNS,
            show="headings",
            height=9,
            style="Registration.Treeview",
        )
        self.configure_table_headings(self.tree_plant)
        self.tree_plant.column("SR", width=80, anchor="center", stretch=False)
        self.tree_plant.column("UID", width=230, anchor="center", stretch=False)
        self.tree_plant.column("Name", width=360, anchor="w", stretch=False)
        self.tree_plant.column("GrossCost", width=130, anchor="center", stretch=False)
        self.tree_plant.column("REG_DATE", width=150, anchor="center", stretch=False)
        self.tree_plant.column("Status", width=170, anchor="center", stretch=False)
        self.tree_plant.column("Complete", width=110, anchor="center", stretch=False)
        self.tree_plant.column("CompleteDate", width=130, anchor="center", stretch=False)
        self.tree_plant.column("Commissioned", width=130, anchor="center", stretch=False)
        self.tree_plant.column("CommissionedDate", width=150, anchor="center", stretch=False)
        self.tree_plant.grid(row=0, column=0, sticky="nsew")

        scrollbar2 = ttk.Scrollbar(tree_frame2, orient="vertical", command=self.tree_plant.yview)
        scrollbar2.grid(row=0, column=1, sticky="ns")
        hscrollbar2 = Scrollbar(tree_frame2, orient="horizontal", command=self.tree_plant.xview, width=18)
        hscrollbar2.grid(row=1, column=0, sticky="ew", pady=(4, 0))
        self.tree_plant.configure(yscrollcommand=scrollbar2.set, xscrollcommand=hscrollbar2.set)
        self.bind_project_tree(self.tree_plant)
        self.tree_plant.bind("<ButtonRelease-1>", self.on_tree_click)
        self.tree_plant.bind("<Double-1>", self.on_double_click)

        # ==================== DELETE + REFRESH BUTTONS (Side by Side) ====================
        btn_frame = Frame(content_frame, bg="#f0f4f8")
        btn_frame.pack(pady=30)

        self.delete_btn = Button(btn_frame, text="Delete",
                                 command=self.delete_project,
                                 bg="#c8102e", fg="white", font=("Arial", 11, "bold"),
                                 width=18, height=2, state=edit_state)
        self.delete_btn.pack(side=LEFT, padx=10)
        self.child_btn = Button(
            btn_frame,
            text="Add Child",
            command=self.add_child_project,
            bg="#0f766e",
            fg="white",
            font=("Arial", 11, "bold"),
            width=18,
            height=2,
            state=edit_state,
        )
        self.child_btn.pack(side=LEFT, padx=10)
        Button(btn_frame, text="Refresh List", 
               command=self.load_list,
               bg="#0066cc", fg="white", font=("Arial", 11, "bold"), 
               width=18, height=2).pack(side=LEFT, padx=10)
        # ===================================================================

        self.load_list()
        apply_page_watermark(self)
        normalize_buttons(self)
        self.style_registration_buttons()
        self._lock_table_widths()
        self.content_frame.bind("<Configure>", self.on_layout_resize)

    # ==================== REST OF THE METHODS (unchanged) ====================
    def configure_tree_style(self):
        style = ttk.Style(self)
        style.configure("Registration.Treeview", rowheight=24)
        style.map(
            "Registration.Treeview",
            background=[("selected", "#0785d8")],
            foreground=[("selected", "white")],
        )

    def bind_project_tree(self, tree):
        tree.bind("<Button-1>", self.on_tree_button_press)
        tree.bind("<MouseWheel>", lambda event, widget=tree: self.on_tree_mousewheel(event, widget))

    def on_tree_button_press(self, event):
        tree = event.widget
        if tree not in (self.tree_corporate, self.tree_plant):
            return
        self.remember_scroll_positions(tree)
        self.active_project_tree = tree
        tree.focus_set()
        row_id = tree.identify_row(event.y)
        if row_id:
            self.clear_other_table_selection(tree)
            tree.selection_set(row_id)
            self.restore_scroll_positions(tree)

    def clear_other_table_selection(self, active_tree):
        other_tree = self.tree_plant if active_tree == self.tree_corporate else self.tree_corporate
        other_tree.selection_remove(other_tree.selection())

    def remember_scroll_positions(self, tree=None):
        trees = (tree,) if tree in (self.tree_corporate, self.tree_plant) else (self.tree_corporate, self.tree_plant)
        for current_tree in trees:
            try:
                self._pending_tree_yview[current_tree] = {
                    "y": current_tree.yview()[0],
                    "x": current_tree.xview()[0],
                }
            except Exception:
                pass
        try:
            self._pending_page_yview = self.scroll_frame.canvas.yview()[0]
        except Exception:
            self._pending_page_yview = None

    def restore_scroll_positions(self, tree=None):
        trees = (tree,) if tree in (self.tree_corporate, self.tree_plant) else (self.tree_corporate, self.tree_plant)

        def restore():
            for current_tree in trees:
                position = self._pending_tree_yview.get(current_tree)
                if position is not None:
                    try:
                        current_tree.yview_moveto(position["y"])
                        current_tree.xview_moveto(position["x"])
                    except Exception:
                        pass
            if self._pending_page_yview is not None:
                try:
                    self.scroll_frame.canvas.yview_moveto(self._pending_page_yview)
                except Exception:
                    pass

        self.after_idle(restore)
        self.after(20, restore)
        self.after(80, restore)

    def on_tree_mousewheel(self, event, tree):
        delta = int(-1 * (event.delta / 120))
        if delta:
            tree.yview_scroll(delta, "units")
        return "break"

    def iter_buttons(self, widget):
        for child in widget.winfo_children():
            if isinstance(child, Button):
                yield child
            yield from self.iter_buttons(child)

    def style_registration_buttons(self, widget=None):
        for button in self.iter_buttons(widget or self.content_frame):
            button.config(
                font=self.BUTTON_FONT,
                width=self.BUTTON_WIDTH,
                height=self.BUTTON_HEIGHT,
                anchor=CENTER,
                justify=CENTER,
            )

    def create_table_search(self, parent, table_key):
        search_frame = Frame(parent, bg="#f0f4f8")
        search_frame.pack(fill=X, padx=10, pady=(0, 4), anchor="w")

        Label(search_frame, text="Filter By:", bg="#f0f4f8", fg="#003087",
              font=("Arial", 10, "bold")).pack(side=LEFT)
        filter_var = StringVar(value="Project Name")
        self.table_filter_vars[table_key] = filter_var
        filter_combo = ttk.Combobox(
            search_frame,
            textvariable=filter_var,
            values=("Project Name", "Unique ID", "Date of Registration"),
            state="readonly",
            width=22,
        )
        filter_combo.pack(side=LEFT, padx=(8, 8))
        filter_var.trace_add("write", lambda *_args, key=table_key: self.apply_table_view(key))

        Label(search_frame, text="Value:", bg="#f0f4f8", fg="#003087",
              font=("Arial", 10, "bold")).pack(side=LEFT)
        search_var = StringVar()
        self.table_search_vars[table_key] = search_var
        search_entry = Entry(search_frame, textvariable=search_var, width=34)
        search_entry.pack(side=LEFT, padx=(8, 8))
        search_var.trace_add("write", lambda *_args, key=table_key: self.apply_table_view(key))

        Button(search_frame, text="Clear", command=lambda key=table_key: self.clear_table_search(key),
               bg="#666666", fg="white", font=("Arial", 9, "bold"),
               width=8, height=1).pack(side=LEFT)

    def configure_table_headings(self, tree):
        for column in self.TABLE_COLUMNS:
            tree.heading(column, text=self.TABLE_HEADINGS[column])

    def get_table(self, table_key):
        return self.tree_corporate if table_key == "corporate" else self.tree_plant

    def clear_table_search(self, table_key):
        search_var = self.table_search_vars.get(table_key)
        if search_var:
            search_var.set("")

    def make_table_row(self, values):
        searchable_text = " ".join(str(value or "") for value in values).lower()
        return {"values": values, "search": searchable_text}

    def format_gross_cost(self, project):
        value = project.get("stage2_cost")
        if value in (None, ""):
            value = project.get("stage1_cost")
        if value in (None, ""):
            return ""
        try:
            amount = float(value)
        except (TypeError, ValueError):
            return str(value)
        return str(int(amount)) if amount.is_integer() else f"{amount:.2f}"

    def apply_table_view(self, table_key):
        tree = self.get_table(table_key)
        for item in tree.get_children():
            tree.delete(item)

        rows = list(self.table_rows.get(table_key, []))
        search_var = self.table_search_vars.get(table_key)
        search_text = search_var.get().strip().lower() if search_var else ""
        if search_text:
            field_var = self.table_filter_vars.get(table_key)
            filter_field = field_var.get() if field_var else "Project Name"
            filter_columns = {
                "Project Name": "Name",
                "Unique ID": "UID",
                "Date of Registration": "REG_DATE",
            }
            filter_column = filter_columns.get(filter_field, "Name")
            column_index = self.TABLE_COLUMNS.index(filter_column)
            rows = [
                row for row in rows
                if search_text in str(row["values"][column_index] or "").lower()
            ]

        for row in rows:
            tree.insert("", END, values=row["values"])

        self._lock_table_widths()

    def register(self):
        if self.main_app and not self.main_app.can_edit("registration"):
            messagebox.showwarning("Edit Denied", "You have view access only for Project Registration.")
            return

        ptype = self.type_var.get()
        name = self.name_entry.get().strip()
        
        if not ptype or not name:
            messagebox.showerror("Error", "Please fill Project Type and Project Name")
            return

        # ========== NEW: Check for Duplicate Project Name ==========
        if self.project_name_exists(name):
            messagebox.showerror("Duplicate Name", 
                f"Project Name '{name}' already exists!\n\nPlease enter a different project name.")
            return
        # ===========================================================

        uid = generate_unique_id()
        add_project(uid, ptype, name)
        messagebox.showinfo("Success", f"Project {uid} registered successfully!")
        keep_window_active(self)
        self.load_list()
        self.name_entry.delete(0, END)   # Clear name field after successful registration

    def project_name_exists(self, name):
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT project_name FROM projects WHERE LOWER(project_name) = LOWER(%s)", (str(name or "").strip(),))
        existing = c.fetchone()
        conn.close()
        return bool(existing)

    def get_selected_project(self):
        tree_order = []
        if self.active_project_tree in (self.tree_corporate, self.tree_plant):
            tree_order.append(self.active_project_tree)
        tree_order.extend(tree for tree in (self.tree_corporate, self.tree_plant) if tree not in tree_order)

        selected = None
        tree = None
        for candidate_tree in tree_order:
            candidate_selection = candidate_tree.selection()
            if candidate_selection:
                selected = candidate_selection
                tree = candidate_tree
                break

        if not selected or not tree:
            return None

        item = tree.item(selected[0])
        values = item.get("values", [])
        if len(values) < 2:
            return None

        uid = values[1]
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM projects WHERE unique_id=%s", (uid,))
        project = c.fetchone()
        conn.close()
        return dict(project) if project else None

    def add_child_project(self):
        if self.main_app and not self.main_app.can_edit("registration"):
            messagebox.showwarning("Edit Denied", "You have view access only for Project Registration.")
            return

        parent_project = self.get_selected_project()
        if not parent_project:
            messagebox.showwarning("Select Project", "Please select a registered project first.")
            return

        popup = Toplevel(self)
        popup.title("Add Child Project")
        popup.geometry("560x350")
        popup.configure(bg="#f0f4f8")
        popup.transient(self.winfo_toplevel())
        popup.grab_set()

        Label(
            popup,
            text="Add Child Project",
            bg="#f0f4f8",
            fg="#003087",
            font=("Arial", 14, "bold"),
        ).pack(pady=(18, 14))

        form = Frame(popup, bg="#f0f4f8")
        form.pack(fill=X, padx=24)
        form.grid_columnconfigure(1, weight=1)

        Label(form, text="Parent Unique ID:", bg="#f0f4f8", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=6)
        parent_uid = Entry(form, width=42)
        parent_uid.grid(row=0, column=1, sticky="ew", pady=6)
        parent_uid.insert(0, str(parent_project.get("unique_id") or ""))
        parent_uid.config(state="readonly")

        Label(form, text="Parent Project:", bg="#f0f4f8", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky="w", pady=6)
        parent_name = Entry(form, width=42)
        parent_name.grid(row=1, column=1, sticky="ew", pady=6)
        parent_name.insert(0, str(parent_project.get("project_name") or ""))
        parent_name.config(state="readonly")

        Label(form, text="Project Type:", bg="#f0f4f8", font=("Arial", 10, "bold")).grid(row=2, column=0, sticky="w", pady=6)
        parent_type = Entry(form, width=42)
        parent_type.grid(row=2, column=1, sticky="ew", pady=6)
        parent_type.insert(0, str(parent_project.get("project_type") or ""))
        parent_type.config(state="readonly")

        Label(form, text="Child Project Name:", bg="#f0f4f8", font=("Arial", 10, "bold")).grid(row=3, column=0, sticky="w", pady=6)
        child_name_entry = Entry(form, width=42)
        child_name_entry.grid(row=3, column=1, sticky="ew", pady=6)

        Label(form, text="Stage-2 Gross Cost:", bg="#f0f4f8", font=("Arial", 10, "bold")).grid(row=4, column=0, sticky="w", pady=6)
        stage2_cost_entry = Entry(form, width=42)
        stage2_cost_entry.grid(row=4, column=1, sticky="ew", pady=6)

        def save_child():
            child_name = child_name_entry.get().strip()
            if not child_name:
                messagebox.showerror("Required", "Please enter Child Project Name.")
                keep_window_active(popup)
                return
            if self.project_name_exists(child_name):
                messagebox.showerror(
                    "Duplicate Name",
                    f"Project Name '{child_name}' already exists!\n\nPlease enter a different project name.",
                )
                keep_window_active(popup)
                return

            cost_text = stage2_cost_entry.get().strip().replace(",", "")
            if not cost_text:
                messagebox.showerror("Required", "Please enter Stage-2 Gross Cost for the child project.")
                keep_window_active(popup)
                return
            try:
                stage2_gross_cost = float(cost_text)
            except ValueError:
                messagebox.showerror("Invalid Cost", "Stage-2 Gross Cost must be a numeric value.")
                keep_window_active(popup)
                return
            if stage2_gross_cost < 0:
                messagebox.showerror("Invalid Cost", "Stage-2 Gross Cost cannot be negative.")
                keep_window_active(popup)
                return

            try:
                uid = generate_unique_id()
                add_child_project(parent_project["id"], uid, child_name, stage2_gross_cost)
            except Exception as e:
                messagebox.showerror("Add Child Failed", str(e))
                keep_window_active(popup)
                return

            popup.destroy()
            self.load_list()
            messagebox.showinfo("Success", f"Child project {uid} added successfully.")
            keep_window_active(self)

        button_frame = Frame(popup, bg="#f0f4f8")
        button_frame.pack(pady=22)
        Button(button_frame, text="Save", command=save_child,
               bg="#008000", fg="white", font=("Arial", 10, "bold"),
               width=14).pack(side=LEFT, padx=8)
        Button(button_frame, text="Cancel", command=popup.destroy,
               bg="#666666", fg="white", font=("Arial", 10, "bold"),
               width=14).pack(side=LEFT, padx=8)

        normalize_buttons(popup)
        self.style_registration_buttons(popup)
        child_name_entry.focus_set()

    def download_template(self):
        if self.main_app and not self.main_app.is_admin():
            messagebox.showwarning("Admin Only", "Only Admin can download the registration template.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Save Registration Template",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("CSV File", "*.csv")],
            initialfile="project_registration_template.xlsx",
        )
        if not file_path:
            return

        headers = [
            "Project Type",
            "Project Name",
            "DIC Recommendation Date",
            "COD Date",
            "Stage-1 Date",
            "Stage-1 Cost",
            "Expected Tender Opening Date",
            "Final Tender Opening Date",
            "Stage-2 Date",
            "Stage-2 Cost",
        ]
        sample_rows = [
            ["Corporate AMR", "Sample Corporate Project", "24-04-26", "", "", "", "", "", "", ""],
            ["Plant Level AMR", "Sample Plant Project", "24-04-26", "26-04-26", "28-04-26", "125.50", "", "", "", ""],
        ]

        try:
            if file_path.lower().endswith(".csv"):
                with open(file_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(sample_rows)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Registration Template"
                ws.append(headers)
                for row in sample_rows:
                    ws.append(row)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                widths = {
                    "A": 22,
                    "B": 38,
                    "C": 22,
                    "D": 16,
                    "E": 16,
                    "F": 16,
                    "G": 28,
                    "H": 24,
                    "I": 16,
                    "J": 16,
                }
                for col, width in widths.items():
                    ws.column_dimensions[col].width = width

                note = wb.create_sheet("Instructions")
                note["A1"] = "Use only these Project Type values:"
                note["A2"] = "Corporate AMR"
                note["A3"] = "Plant Level AMR"
                note["A5"] = "Date format:"
                note["A6"] = "DD-MM-YY"
                note["A8"] = "Required columns in first row:"
                for idx, header in enumerate(headers, start=9):
                    note[f"A{idx}"] = header
                note["C1"] = "Stage auto-shift logic:"
                note["C2"] = "If only DIC Recommendation Date is filled -> project stays at Stage-1"
                note["C3"] = "If Stage-1 Date, COD Date and Stage-1 Cost are also filled -> project moves to Tendering"
                note["C4"] = "If Expected TOD and Final TOD are also filled -> project moves to Stage-2"
                note["C5"] = "If Stage-2 Date and Stage-2 Cost are also filled -> project moves to Ongoing"
                note["C7"] = "Do not leave a stage partially filled."
                note["C8"] = "If any field of a stage is filled, all required fields of that stage must be filled."
                note.column_dimensions["A"].width = 42
                note.column_dimensions["C"].width = 75
                wb.save(file_path)

            messagebox.showinfo(
                "Template Saved",
                "Template saved successfully.\n\nIt now includes project type, project name, and stage-wise fields.",
            )
            keep_window_active(self)
        except Exception as e:
            messagebox.showerror("Template Error", f"Failed to create template:\n{e}")
            keep_window_active(self)

    def read_uploaded_rows(self, file_path):
        if file_path.lower().endswith(".csv"):
            with open(file_path, "r", newline="", encoding="utf-8-sig") as f:
                return list(csv.DictReader(f))

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        data_rows = []
        for row in rows[1:]:
            if not any(str(value or "").strip() for value in row):
                continue
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else ""
            data_rows.append(item)
        return data_rows

    def upload_sheet(self):
        if self.main_app and not self.main_app.is_admin():
            messagebox.showwarning("Admin Only", "Only Admin can upload registration sheets.")
            return

        file_path = filedialog.askopenfilename(
            title="Upload Registration Sheet",
            filetypes=[("Spreadsheet Files", "*.xlsx *.csv"), ("Excel Workbook", "*.xlsx"), ("CSV File", "*.csv")],
        )
        if not file_path:
            return

        try:
            raw_rows = self.read_uploaded_rows(file_path)
        except Exception as e:
            messagebox.showerror("Upload Error", f"Failed to read sheet:\n{e}")
            keep_window_active(self)
            return

        if not raw_rows:
            messagebox.showwarning("No Data", "No rows found in the uploaded sheet.")
            return

        normalized_rows = []
        for row in raw_rows:
            normalized = {str(k or "").strip().lower().replace("_", " "): str(v or "").strip() for k, v in row.items()}
            normalized_rows.append(normalized)

        created = 0
        skipped = []
        valid_types = {"corporate amr": "Corporate AMR", "plant level amr": "Plant Level AMR"}

        for index, row in enumerate(normalized_rows, start=2):
            project_type = row.get("project type", "")
            project_name = row.get("project name", "")
            dic_date_text = row.get("dic recommendation date", "")
            cod_date_text = row.get("cod date", "")
            stage1_date_text = row.get("stage-1 date", "") or row.get("stage 1 date", "")
            stage1_cost_text = row.get("stage-1 cost", "") or row.get("stage 1 cost", "")
            expected_tod_text = row.get("expected tender opening date", "")
            final_tod_text = row.get("final tender opening date", "")
            stage2_date_text = row.get("stage-2 date", "") or row.get("stage 2 date", "")
            stage2_cost_text = row.get("stage-2 cost", "") or row.get("stage 2 cost", "")

            if not project_type and not project_name:
                continue
            if not project_type or not project_name:
                skipped.append(f"Row {index}: Project Type and Project Name are required")
                continue

            project_type_key = project_type.lower()
            if project_type_key not in valid_types:
                skipped.append(f"Row {index}: Invalid Project Type '{project_type}'")
                continue

            if self.project_name_exists(project_name):
                skipped.append(f"Row {index}: Duplicate Project Name '{project_name}'")
                continue

            dic_date = to_storage_date(dic_date_text) if dic_date_text else None
            cod_date = to_storage_date(cod_date_text) if cod_date_text else None
            stage1_date = to_storage_date(stage1_date_text) if stage1_date_text else None
            expected_tod_date = to_storage_date(expected_tod_text) if expected_tod_text else None
            final_tod_date = to_storage_date(final_tod_text) if final_tod_text else None
            stage2_date = to_storage_date(stage2_date_text) if stage2_date_text else None

            invalid_dates = []
            if dic_date_text and not dic_date:
                invalid_dates.append("DIC Recommendation Date")
            if cod_date_text and not cod_date:
                invalid_dates.append("COD Date")
            if stage1_date_text and not stage1_date:
                invalid_dates.append("Stage-1 Date")
            if expected_tod_text and not expected_tod_date:
                invalid_dates.append("Expected Tender Opening Date")
            if final_tod_text and not final_tod_date:
                invalid_dates.append("Final Tender Opening Date")
            if stage2_date_text and not stage2_date:
                invalid_dates.append("Stage-2 Date")
            if invalid_dates:
                skipped.append(f"Row {index}: Invalid date format for {', '.join(invalid_dates)}. Use DD-MM-YY")
                continue

            try:
                stage1_cost = float(stage1_cost_text) if stage1_cost_text else None
                stage2_cost = float(stage2_cost_text) if stage2_cost_text else None
            except ValueError:
                skipped.append(f"Row {index}: Stage cost must be numeric")
                continue

            formulation_complete = bool(dic_date)
            stage1_any = any([cod_date_text, stage1_date_text, stage1_cost_text])
            # Relaxed: Cost is optional (can be 0 or blank)
            stage1_complete = bool(cod_date and stage1_date)
            tendering_any = any([expected_tod_text, final_tod_text])
            tendering_complete = bool(expected_tod_date and final_tod_date)
            stage2_any = any([stage2_date_text, stage2_cost_text])
            # Relaxed: Cost is optional (can be 0 or blank)
            stage2_complete = bool(stage2_date)

            if stage1_any and not formulation_complete:
                skipped.append(f"Row {index}: Fill DIC Recommendation Date before Stage-1 data")
                continue
            if stage1_any and not stage1_complete:
                skipped.append(f"Row {index}: COD Date and Stage-1 Date are required for Stage-1 (Cost can be 0 or blank)")
                continue
            if tendering_any and not stage1_complete:
                skipped.append(f"Row {index}: Complete Stage-1 fields before Tendering data")
                continue
            if tendering_any and not tendering_complete:
                skipped.append(f"Row {index}: Expected Tender Opening Date and Final Tender Opening Date are both required")
                continue
            if stage2_any and not tendering_complete:
                skipped.append(f"Row {index}: Complete Tendering dates before Stage-2 data")
                continue
            if stage2_any and not stage2_complete:
                skipped.append(f"Row {index}: Stage-2 Date is required (Cost can be 0 or blank)")
                continue

            uid = generate_unique_id()
            add_project(uid, valid_types[project_type_key], project_name)
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("SELECT id FROM projects WHERE unique_id=%s", (uid,))
            project_row = c.fetchone()
            conn.close()
            if not project_row:
                skipped.append(f"Row {index}: Project created but could not update stage data")
                continue

            update_data = {
                "dic_recommendation_date": dic_date,
                "cod_date": cod_date,
                "stage1_date": stage1_date,
                "stage1_cost": stage1_cost,
                "expected_tod_date": expected_tod_date,
                "final_tod_date": final_tod_date,
                "stage2_date": stage2_date,
                "stage2_cost": stage2_cost,
                "cod_cleared": "Y" if formulation_complete else "N",
                "stage1_cleared": "Y" if stage1_complete else "N",
                "stage2_cleared": "Y" if stage2_complete else "N",
            }
            clean_update_data = {k: v for k, v in update_data.items() if v is not None or k in ("cod_cleared", "stage1_cleared", "stage2_cleared")}
            update_project_stage(project_row["id"], **clean_update_data)
            created += 1

        self.load_list()

        if skipped:
            preview = "\n".join(skipped[:10])
            extra = f"\n...and {len(skipped) - 10} more" if len(skipped) > 10 else ""
            messagebox.showinfo(
                "Upload Completed",
                f"Projects created: {created}\nSkipped: {len(skipped)}\n\n{preview}{extra}",
            )
        else:
            messagebox.showinfo("Upload Completed", f"Projects created successfully: {created}")
        keep_window_active(self)

    def delete_project(self):
        if self.main_app and not self.main_app.can_edit("registration"):
            messagebox.showwarning("Edit Denied", "You do not have delete permission for Project Registration.")
            return

        selected = self.tree_corporate.selection() or self.tree_plant.selection()
        if not selected:
            messagebox.showwarning("Select", "Please select a project from any table to delete")
            return
        
        if self.tree_corporate.selection():
            item = self.tree_corporate.item(selected[0])
        else:
            item = self.tree_plant.item(selected[0])
        
        uid = item['values'][1]
        project_name = item['values'][2]
        
        if not messagebox.askyesno("Confirm Delete", 
                f"Are you sure you want to DELETE this project PERMANENTLY?\n\n"
                f"Unique ID: {uid}\n"
                f"Project Name: {project_name}\n\n"
                "⚠️ This will remove ALL data from EVERYWHERE:\n"
                "• Contract & Appendix-2\n"
                "• All S-Curve Plans & Activities\n"
                "• Daily Progress Reports\n"
                "• All Stage History\n\n"
                "This action CANNOT be undone!"):
            return
        
        try:
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("SELECT id FROM projects WHERE unique_id=%s", (uid,))
            row = c.fetchone()
            if not row:
                messagebox.showerror("Error", "Project not found!")
                return
            project_id = row['id']
            
            conn.close()
            delete_project_everywhere(project_id)
            
            messagebox.showinfo("Deleted", f"Project {uid} has been completely removed from the entire system!")
            keep_window_active(self)
            self.load_list()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete project:\n{str(e)}")
            keep_window_active(self)

    def load_list(self):
        self.remember_scroll_positions()
        self.table_rows = {"corporate": [], "plant": []}
        
        allowed_project_ids = self.main_app.get_allowed_project_ids() if self.main_app else None
        all_projects = [dict(project) for project in get_all_projects(allowed_project_ids)]

        def project_sort_key(project):
            return int(project.get("id") or 0)

        def build_values(project, sr_no, depth=0):
            status = get_project_status(dict(project))
            registration_date = to_display_date(project["registration_date"])
            mark_enabled = status == "Ongoing" or project.get("completion_marked") == "Y" or project.get("commissioned_marked") == "Y"
            complete_mark = "\u2611" if project.get("completion_marked") == "Y" else ("\u2610" if mark_enabled else "")
            complete_date = to_display_date(project.get("completion_date"))
            commissioned_mark = "\u2611" if project.get("commissioned_marked") == "Y" else ("\u2610" if mark_enabled else "")
            commissioned_date = to_display_date(project.get("commissioned_date"))
            prefix = ("    " * max(0, int(depth or 0))) + ("- " if depth else "")
            return (
                sr_no,
                project["unique_id"],
                f"{prefix}{project['project_name']}",
                self.format_gross_cost(project),
                registration_date,
                status,
                complete_mark,
                complete_date,
                commissioned_mark,
                commissioned_date,
            )

        def populate_project_type(project_type, table_key):
            project_rows = [project for project in all_projects if project.get("project_type") == project_type]
            ids_in_view = {int(project.get("id") or 0) for project in project_rows}
            children_by_parent = {}
            roots = []
            for project in project_rows:
                parent_id = project.get("parent_project_id")
                parent_id = int(parent_id) if parent_id else None
                if parent_id and parent_id in ids_in_view:
                    children_by_parent.setdefault(parent_id, []).append(project)
                else:
                    roots.append(project)

            roots.sort(key=project_sort_key, reverse=True)
            for child_rows in children_by_parent.values():
                child_rows.sort(key=project_sort_key, reverse=True)

            serial = 1
            visited = set()

            def append_row(project, depth=0):
                nonlocal serial
                project_id = int(project.get("id") or 0)
                if project_id in visited:
                    return
                visited.add(project_id)
                self.table_rows[table_key].append(self.make_table_row(build_values(project, serial, depth)))
                serial += 1
                for child in children_by_parent.get(project_id, []):
                    append_row(child, depth + 1)

            for root_project in roots:
                append_row(root_project, 0)

            for project in sorted(project_rows, key=project_sort_key, reverse=True):
                if int(project.get("id") or 0) not in visited:
                    append_row(project, 0)

        populate_project_type("Corporate AMR", "corporate")
        populate_project_type("Plant Level AMR", "plant")

        self.apply_table_view("corporate")
        self.apply_table_view("plant")
        self.restore_scroll_positions()
        return
        corporate_sr = 1
        plant_sr = 1

        for p in get_all_projects(allowed_project_ids):
            status = get_project_status(dict(p))
            registration_date = to_display_date(p["registration_date"])
            mark_enabled = status == "Ongoing" or p.get("completion_marked") == "Y" or p.get("commissioned_marked") == "Y"
            complete_mark = "☑" if p.get("completion_marked") == "Y" else ("☐" if mark_enabled else "")
            complete_date = to_display_date(p.get("completion_date"))
            commissioned_mark = "☑" if p.get("commissioned_marked") == "Y" else ("☐" if mark_enabled else "")
            commissioned_date = to_display_date(p.get("commissioned_date"))

            if p["project_type"] == "Corporate AMR":
                values = (
                    corporate_sr,
                    p["unique_id"],
                    p["project_name"],
                    self.format_gross_cost(p),
                    registration_date,
                    status,
                    complete_mark,
                    complete_date,
                    commissioned_mark,
                    commissioned_date,
                )
                self.table_rows["corporate"].append(self.make_table_row(values))
                corporate_sr += 1
            else:
                values = (
                    plant_sr,
                    p["unique_id"],
                    p["project_name"],
                    self.format_gross_cost(p),
                    registration_date,
                    status,
                    complete_mark,
                    complete_date,
                    commissioned_mark,
                    commissioned_date,
                )
                self.table_rows["plant"].append(self.make_table_row(values))
                plant_sr += 1

        self.apply_table_view("corporate")
        self.apply_table_view("plant")

    def _lock_table_widths(self):
        for tree in (self.tree_corporate, self.tree_plant):
            for col in tree["columns"]:
                current_width = int(tree.column(col, "width") or 100)
                tree.column(col, width=current_width, minwidth=current_width, stretch=False)

    def on_layout_resize(self, event=None):
        try:
            available_width = max(780, self.content_frame.winfo_width() - 20)
            for frame in (self.tree_frame1, self.tree_frame2):
                frame.configure(width=available_width)
        except Exception:
            pass

    def ask_mark_date(self, title, initial_date=""):
        popup = Toplevel(self)
        popup.title(title)
        popup.geometry("320x220")
        popup.configure(bg="#f0f4f8")
        popup.transient(self.winfo_toplevel())
        popup.grab_set()

        Label(popup, text=title, bg="#f0f4f8", fg="#003087", font=("Arial", 13, "bold")).pack(pady=(18, 10))
        Label(popup, text="Select Date", bg="#f0f4f8", font=("Arial", 10, "bold")).pack()

        initial_parsed = parse_app_date(initial_date)
        date_entry = DateEntry(
            popup,
            width=18,
            date_pattern="dd-mm-yy",
            background="darkblue",
            foreground="white",
        )
        if initial_parsed:
            date_entry.set_date(initial_parsed)
        date_entry.pack(pady=12)

        result = {"value": None}

        button_frame = Frame(popup, bg="#f0f4f8")
        button_frame.pack(pady=15)

        def save_date():
            result["value"] = date_entry.get_date().strftime("%d-%m-%y")
            popup.destroy()

        Button(button_frame, text="Save", command=save_date,
               bg="#008000", fg="white", font=("Arial", 10, "bold")).pack(side=LEFT, padx=8)
        Button(button_frame, text="Cancel", command=popup.destroy,
               bg="#666666", fg="white", font=("Arial", 10, "bold")).pack(side=LEFT, padx=8)
        normalize_buttons(popup)
        self.style_registration_buttons(popup)
        self.wait_window(popup)
        return result["value"]

    def update_project_mark(self, uid, mark_type, checked, date_value=None):
        flag_column = "completion_marked" if mark_type == "complete" else "commissioned_marked"
        date_column = "completion_date" if mark_type == "complete" else "commissioned_date"
        conn = get_db_connection()
        c = conn.cursor()
        c.execute(
            f"UPDATE projects SET {flag_column}=%s, {date_column}=%s WHERE unique_id=%s",
            ("Y" if checked else "N", to_storage_date(date_value) if checked else None, uid),
        )
        conn.commit()
        conn.close()

    def on_tree_click(self, event):
        tree = event.widget
        if tree not in (self.tree_corporate, self.tree_plant):
            return
        self.remember_scroll_positions(tree)
        self.active_project_tree = tree

        row_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)
        region = tree.identify("region", event.x, event.y)
        if region != "cell" or not row_id:
            return

        self.clear_other_table_selection(tree)
        tree.selection_set(row_id)
        self.restore_scroll_positions(tree)

        if self.main_app and not self.main_app.can_edit("registration"):
            return

        if column_id not in ("#7", "#9"):
            return

        item = tree.item(row_id)
        values = item.get("values", [])
        if len(values) < 10:
            return

        uid = values[1]
        status = values[5]
        if column_id == "#7":
            current_checked = values[6] == "☑"
            current_date = values[7]
            label = "Complete"
            mark_type = "complete"
        else:
            current_checked = values[8] == "☑"
            current_date = values[9]
            label = "Commissioned"
            mark_type = "commissioned"

        if not current_checked and status != "Ongoing":
            messagebox.showwarning("Not Allowed", f"{label} can be marked only when the project status is Ongoing.")
            return

        if current_checked:
            if not messagebox.askyesno("Clear Mark", f"Remove {label} mark for project {uid}?"):
                return
            self.update_project_mark(uid, mark_type, False)
        else:
            selected_date = self.ask_mark_date(f"{label} Date", current_date)
            if not selected_date:
                return
            self.update_project_mark(uid, mark_type, True, selected_date)

        self.load_list()
        if row_id in tree.get_children():
            tree.selection_set(row_id)

    def on_double_click(self, event):
        if event.widget.identify("region", event.x, event.y) == "cell" and event.widget.identify_column(event.x) in ("#7", "#9"):
            return
        if event.widget == self.tree_corporate:
            tree = self.tree_corporate
        else:
            tree = self.tree_plant
            
        selected = tree.selection()
        if not selected:
            return
        item = tree.item(selected[0])
        uid = item['values'][1]
        status = item['values'][5]
        
        from database import get_db_connection
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM projects WHERE unique_id=%s", (uid,))
        project_row = c.fetchone()
        conn.close()
        
        if project_row and self.main_app:
            project = dict(project_row)
            if not self.main_app.can_access_project(project["id"]):
                messagebox.showwarning("Access Denied", "You do not have access to this project.")
                return
            if status == "Ongoing":
                self.main_app.show_ongoing()
            else:
                self.main_app.show_project_details()
            messagebox.showinfo("Navigated", f"✅ Opened {status} section for project {uid}")
        else:
            messagebox.showinfo("Status", f"Current Status: {status}")
