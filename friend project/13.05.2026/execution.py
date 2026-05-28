from tkinter import *
from tkinter import ttk, messagebox, simpledialog, filedialog
from database import (
    get_db_connection,
    update_project_stage,
    project_has_plan,
    project_has_saved_planning,
    project_has_completed_planning,
    get_latest_planned_plan,
    get_activity_progress_rows,
    get_activities_for_plan,
    get_daily_progress_activity_matrix,
    get_daily_progress_display_rows,
    save_daily_progress_with_activities,
    get_appendix_activity_rows,
    get_daily_progress_by_date,
    get_daily_progress_manpower,
    get_daily_report_month_matrix,
)
from utils import add_months, normalize_buttons, keep_window_active, to_display_date, to_storage_date, parse_app_date, apply_page_watermark
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

class ScrollableFrame(Frame):
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        self.canvas = Canvas(self, bg=kwargs.get('bg', '#f0f4f8'), highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = Frame(self.canvas, bg=kwargs.get('bg', '#f0f4f8'))

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side=LEFT, fill=BOTH, expand=True)
        self.scrollbar.pack(side=RIGHT, fill=Y)

        def _resize_inner_frame(event):
            self.canvas.itemconfigure(self.canvas_window, width=event.width)

        self.canvas.bind("<Configure>", _resize_inner_frame)

        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)


class ContractWindow(Toplevel):
    def __init__(self, parent, project_id, uid, main_app=None):
        super().__init__(parent)
        self.project_id = project_id
        self.uid = uid
        self.main_app = main_app
        
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT project_name FROM projects WHERE id=%s", (project_id,))
        proj_row = c.fetchone()
        conn.close()
        self.project_name = proj_row["project_name"] if proj_row else "Unknown Project"

        self.title(f"Contract Details & Appendix-2 - {self.uid}")
        self.geometry("1550x980")
        self.configure(bg="#f0f4f8")

        self.scroll_frame = ScrollableFrame(self, bg="#f0f4f8")
        self.scroll_frame.pack(fill=BOTH, expand=True, padx=40, pady=20)
        main = self.scroll_frame.scrollable_frame

        content = Frame(main, bg="#f0f4f8")
        content.pack(anchor="center", fill=X, expand=True, padx=80)

        # Top Buttons
        btn_top = Frame(content, bg="#f0f4f8")
        btn_top.pack(fill=X, pady=5)
        admin_template_state = NORMAL if (self.main_app and hasattr(self.main_app, "is_admin") and self.main_app.is_admin()) else DISABLED
        Button(btn_top, text="📤 Upload", command=self.upload_contract_template,
               bg="#7c3aed", fg="white", font=("Arial", 10, "bold"),
               width=12, state=admin_template_state).pack(side=RIGHT, padx=5)
        Button(btn_top, text="📄 Template", command=self.download_contract_template,
               bg="#0f766e", fg="white", font=("Arial", 10, "bold"),
               width=12, state=admin_template_state).pack(side=RIGHT, padx=5)
        Button(btn_top, text="🔄 Refresh", command=self.refresh_data,
               bg="#28a745", fg="white", font=("Arial", 10, "bold"), width=12).pack(side=RIGHT, padx=5)
        Button(btn_top, text="🏠 Home", command=self.go_home,
               bg="#008000", fg="white", font=("Arial", 10, "bold"), width=12).pack(side=RIGHT, padx=5)
        Button(btn_top, text="← Back", command=self.destroy,
               bg="#555", fg="white", font=("Arial", 10, "bold"), width=12).pack(side=RIGHT, padx=5)

        Label(content, text="Contract Details & Appendix-2", 
              font=("Arial", 22, "bold"), bg="#f0f4f8", fg="#003087").pack(pady=(15, 2))
        Label(content, text=self.project_name, 
              font=("Arial", 16, "bold"), bg="#f0f4f8", fg="#0066cc").pack(pady=(0, 20))

        form = Frame(content, bg="#f0f4f8")
        form.pack(fill=X, pady=10)

        Label(form, text="Contractor Name:", bg="#f0f4f8", font=("Arial", 11)).pack(anchor=W)
        self.contractor_entry = Entry(form, width=70, font=("Arial", 11))
        self.contractor_entry.pack(fill=X, pady=(0, 12))

        Label(form, text="LOA Date (DD-MM-YY):", bg="#f0f4f8", font=("Arial", 11)).pack(anchor=W)
        loa_f = Frame(form, bg="#f0f4f8")
        loa_f.pack(fill=X, pady=(0, 12))
        self.loa_entry = Entry(loa_f, width=30, font=("Arial", 11))
        self.loa_entry.pack(side=LEFT)
        Button(loa_f, text="📅", width=4, command=lambda: self.pick_date(self.loa_entry)).pack(side=LEFT, padx=8)

        Label(form, text="Effective Date of Contract (DD-MM-YY):", bg="#f0f4f8", font=("Arial", 11)).pack(anchor=W)
        eff_f = Frame(form, bg="#f0f4f8")
        eff_f.pack(fill=X, pady=(0, 12))
        self.effective_entry = Entry(eff_f, width=30, font=("Arial", 11))
        self.effective_entry.pack(side=LEFT)
        Button(eff_f, text="📅", width=4, command=lambda: self.pick_date(self.effective_entry)).pack(side=LEFT, padx=8)

        Label(form, text="Project Schedule in Months:", bg="#f0f4f8", font=("Arial", 11)).pack(anchor=W)
        self.schedule_entry = Entry(form, width=15, font=("Arial", 11))
        self.schedule_entry.pack(anchor=W, pady=(0, 12))

        Label(form, text="Schedule Completion Date (DD-MM-YY):", bg="#f0f4f8", font=("Arial", 11)).pack(anchor=W)
        self.completion_label = Label(form, text="---", font=("Arial", 12, "bold"), fg="blue", bg="#f0f4f8")
        self.completion_label.pack(anchor=W, pady=(0, 12))

        Button(form, text="Calculate Completion Date", command=self.calculate_completion,
               bg="#0066cc", fg="white", font=("Arial", 10, "bold"), width=25).pack(anchor=W, pady=8)

        Label(content, text="Appendix-2 Time Schedule", 
              font=("Arial", 16, "bold"), bg="#f0f4f8", fg="#003087").pack(anchor=W, pady=(30, 8))

        cols = ("S.No.", "Parent", "Item of Work", "Commencement\n(Months)", "Completion\n(Months)", "Schedule Start", "Schedule Finish")
        tree_frame = Frame(content, bg="#f0f4f8")
        tree_frame.pack(fill=BOTH, expand=True, pady=5)

        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=14)
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=130 if col == "Item of Work" else 110)
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        y_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        y_scroll.pack(side=RIGHT, fill=Y)
        x_scroll = ttk.Scrollbar(content, orient="horizontal", command=self.tree.xview)
        x_scroll.pack(fill=X)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        btn_frame = Frame(content, bg="#f0f4f8")
        btn_frame.pack(pady=20)
        Button(btn_frame, text="➕ Add Appendix-2 Line Item", command=self.add_appendix_item,
               bg="#0066cc", fg="white", font=("Arial", 11, "bold"), width=28).pack(side=LEFT, padx=8)
        Button(btn_frame, text="🗑️", command=self.delete_selected,
               bg="#c84d10", fg="white", font=("Arial", 11, "bold"), width=22).pack(side=LEFT, padx=8)
        Button(btn_frame, text="Calculate All Schedule Dates", command=self.calculate_all_dates,
               bg="#555", fg="white", font=("Arial", 11, "bold"), width=28).pack(side=LEFT, padx=8)

        self.save_btn = Button(content, text="SAVE CONTRACT DETAILS & APPENDIX-2", 
                               command=self.save_contract, bg="#008000", fg="white", 
                               font=("Arial", 12, "bold"), height=2)
        self.save_btn.pack(pady=30, fill=X)

        self.load_data()
        apply_page_watermark(self)
        normalize_buttons(self)

    def is_admin_user(self):
        return bool(self.main_app and hasattr(self.main_app, "is_admin") and self.main_app.is_admin())

    def get_appendix_groups(self):
        return {
            "Design & Engineering": ["Basic Engineering", "Detailed Design Engineering"],
            "Civil Work": ["Civil Execution"],
            "Supply": [
                "Building Steel Structures & Sheeting",
                "Mechanical Plant & Equipment - Imported",
                "Mechanical Plant & Equipment - Indigenous",
                "Electrical Plant & Equipment - Imported",
                "Electrical Plant & Equipment - Indigenous",
                "Refractories - Imported",
                "Refractories - Indigenous",
            ],
            "Erection": [
                "Building Steel Structures & Sheeting",
                "Mechanical Plant & Equipment",
                "Electrical Plant & Equipment",
                "Refractories",
            ],
            "Testing & Commissioning": ["Preliminary Acceptance", "Commissioning"],
        }

    def download_contract_template(self):
        if not self.is_admin_user():
            messagebox.showwarning("Admin Only", "Only Admin can download the contract template.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Save Contract Template",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"{self.uid.replace('/', '_')}_contract_template.xlsx",
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Contract Details"
            ws.append(["Field", "Value"])
            contract_rows = [
                ["Project UID", self.uid],
                ["Project Name", self.project_name],
                ["Contractor Name", ""],
                ["LOA Date", ""],
                ["Effective Date of Contract", ""],
                ["Project Schedule in Months", ""],
                ["Schedule Completion Date", ""],
            ]
            for row in contract_rows:
                ws.append(row)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 35

            appendix = wb.create_sheet("Appendix-2")
            appendix_headers = [
                "S.No.",
                "Parent",
                "Item of Work",
                "Commencement (Months)",
                "Completion (Months)",
                "Schedule Start",
                "Schedule Finish",
            ]
            appendix.append(appendix_headers)
            for cell in appendix[1]:
                cell.font = Font(bold=True)
            widths = {"A": 10, "B": 26, "C": 38, "D": 24, "E": 22, "F": 18, "G": 18}
            for col, width in widths.items():
                appendix.column_dimensions[col].width = width

            groups = self.get_appendix_groups()
            list_sheet = wb.create_sheet("ValidationLists")
            list_sheet["A1"] = "Parent"
            list_sheet["B1"] = "Item of Work"
            parents = list(groups.keys())
            children = []
            for items in groups.values():
                for item in items:
                    if item not in children:
                        children.append(item)
            for row_idx, parent in enumerate(parents, start=2):
                list_sheet.cell(row=row_idx, column=1, value=parent)
            for row_idx, child in enumerate(children, start=2):
                list_sheet.cell(row=row_idx, column=2, value=child)
            list_sheet.sheet_state = "hidden"

            parent_validation = DataValidation(
                type="list",
                formula1=f"=ValidationLists!$A$2:$A${len(parents) + 1}",
                allow_blank=True,
            )
            child_validation = DataValidation(
                type="list",
                formula1=f"=ValidationLists!$B$2:$B${len(children) + 1}",
                allow_blank=True,
            )
            appendix.add_data_validation(parent_validation)
            appendix.add_data_validation(child_validation)
            parent_validation.add("B2:B500")
            child_validation.add("C2:C500")

            note = wb.create_sheet("Instructions")
            note["A1"] = "Date format"
            note["A2"] = "DD-MM-YY"
            note["A4"] = "Required Contract sheet fields"
            note["A5"] = "Contractor Name"
            note["A6"] = "LOA Date"
            note["A7"] = "Effective Date of Contract"
            note["A8"] = "Project Schedule in Months"
            note["A10"] = "Appendix-2 sheet"
            note["A11"] = "Fill one line item per row."
            note["A12"] = "Schedule Start and Schedule Finish are optional."
            note["A13"] = "If blank, they will be auto-calculated from Effective Date and month values."
            note["A14"] = "Parent and Item of Work columns have dropdowns."
            note.column_dimensions["A"].width = 60

            wb.save(file_path)
            messagebox.showinfo("Template Saved", "Contract template saved successfully.")
            keep_window_active(self)
        except Exception as e:
            messagebox.showerror("Template Error", f"Failed to create template:\n{e}")
            keep_window_active(self)

    def upload_contract_template(self):
        if not self.is_admin_user():
            messagebox.showwarning("Admin Only", "Only Admin can upload the contract template.")
            return

        file_path = filedialog.askopenfilename(
            title="Upload Contract Template",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            if "Contract Details" not in wb.sheetnames or "Appendix-2" not in wb.sheetnames:
                messagebox.showerror("Upload Error", "Workbook must contain 'Contract Details' and 'Appendix-2' sheets.")
                keep_window_active(self)
                return

            contract_map = {}
            contract_ws = wb["Contract Details"]
            for row in contract_ws.iter_rows(min_row=2, values_only=True):
                key = str(row[0] or "").strip()
                value = row[1] if len(row) > 1 else ""
                if key:
                    contract_map[key.lower()] = value

            contractor_name = str(contract_map.get("contractor name", "") or "").strip()
            loa_value = to_storage_date(contract_map.get("loa date")) if contract_map.get("loa date") else None
            eff_value = to_storage_date(contract_map.get("effective date of contract")) if contract_map.get("effective date of contract") else None
            schedule_months_raw = contract_map.get("project schedule in months")
            schedule_completion_value = to_storage_date(contract_map.get("schedule completion date")) if contract_map.get("schedule completion date") else None

            if not contractor_name:
                messagebox.showerror("Upload Error", "Contractor Name is required in the Contract Details sheet.")
                keep_window_active(self)
                return
            if contract_map.get("loa date") and not loa_value:
                messagebox.showerror("Upload Error", "Invalid LOA Date. Use DD-MM-YY or a valid Excel date cell.")
                keep_window_active(self)
                return
            if contract_map.get("effective date of contract") and not eff_value:
                messagebox.showerror("Upload Error", "Invalid Effective Date of Contract. Use DD-MM-YY or a valid Excel date cell.")
                keep_window_active(self)
                return
            if contract_map.get("schedule completion date") and not schedule_completion_value:
                messagebox.showerror("Upload Error", "Invalid Schedule Completion Date. Use DD-MM-YY or a valid Excel date cell.")
                keep_window_active(self)
                return

            if schedule_months_raw in ("", None):
                schedule_months = None
            else:
                try:
                    schedule_months = int(float(schedule_months_raw))
                except ValueError:
                    messagebox.showerror("Upload Error", "Project Schedule in Months must be numeric.")
                    keep_window_active(self)
                    return

            appendix_ws = wb["Appendix-2"]
            appendix_rows = []
            for excel_row_num, row in enumerate(appendix_ws.iter_rows(min_row=2, values_only=True), start=2):
                values = list(row[:7]) + [""] * max(0, 7 - len(row[:7]))
                if not any(str(v or "").strip() for v in values):
                    continue

                s_no = str(values[0] or "").strip()
                parent = str(values[1] or "").strip()
                item = str(values[2] or "").strip()
                comm_raw = str(values[3] or "").strip()
                comp_raw = str(values[4] or "").strip()
                start_value = to_storage_date(values[5]) if values[5] not in ("", None) else None
                finish_value = to_storage_date(values[6]) if values[6] not in ("", None) else None

                if not s_no or not parent or not item or not comm_raw or not comp_raw:
                    messagebox.showerror("Upload Error", f"Appendix-2 row {excel_row_num} is missing required fields.")
                    keep_window_active(self)
                    return
                if not comm_raw.isdigit() or not comp_raw.isdigit():
                    messagebox.showerror("Upload Error", f"Appendix-2 row {excel_row_num} month values must be numeric.")
                    keep_window_active(self)
                    return
                if values[5] not in ("", None) and not start_value:
                    messagebox.showerror("Upload Error", f"Appendix-2 row {excel_row_num} has an invalid Schedule Start date.")
                    keep_window_active(self)
                    return
                if values[6] not in ("", None) and not finish_value:
                    messagebox.showerror("Upload Error", f"Appendix-2 row {excel_row_num} has an invalid Schedule Finish date.")
                    keep_window_active(self)
                    return

                appendix_rows.append((s_no, parent, item, comm_raw, comp_raw, start_value, finish_value))

            self.contractor_entry.delete(0, END)
            self.contractor_entry.insert(0, contractor_name)
            self.loa_entry.delete(0, END)
            self.loa_entry.insert(0, to_display_date(loa_value))
            self.effective_entry.delete(0, END)
            self.effective_entry.insert(0, to_display_date(eff_value))
            self.schedule_entry.delete(0, END)
            self.schedule_entry.insert(0, str(schedule_months) if schedule_months is not None else "")

            if schedule_completion_value:
                self.completion_label.config(text=to_display_date(schedule_completion_value))
            elif eff_value and schedule_months is not None:
                self.completion_label.config(text=to_display_date(add_months(eff_value, schedule_months)))
            else:
                self.completion_label.config(text="---")

            for item_id in self.tree.get_children():
                self.tree.delete(item_id)

            effective_for_calc = eff_value or ""
            for s_no, parent, item, comm_raw, comp_raw, start_value, finish_value in appendix_rows:
                start_display = to_display_date(start_value) if start_value else to_display_date(add_months(effective_for_calc, int(comm_raw))) if effective_for_calc else ""
                finish_display = to_display_date(finish_value) if finish_value else to_display_date(add_months(effective_for_calc, int(comp_raw))) if effective_for_calc else ""
                self.tree.insert("", END, values=(s_no, parent, item, comm_raw, comp_raw, start_display, finish_display))

            self.save_contract()
        except Exception as e:
            messagebox.showerror("Upload Error", f"Failed to upload template:\n{e}")
            keep_window_active(self)

    def save_contract(self):
        try:
            data = {
                "contractor_name": self.contractor_entry.get().strip(),
                "loa_date": to_storage_date(self.loa_entry.get()),
                "effective_date": to_storage_date(self.effective_entry.get()),
                "schedule_months": int(self.schedule_entry.get().strip()) if self.schedule_entry.get().strip().isdigit() else None,
                "schedule_completion": to_storage_date(self.completion_label.cget("text"))
            }
            update_project_stage(self.project_id, **data)
            
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("DELETE FROM appendix2 WHERE project_id=%s", (self.project_id,))
            
            for child in self.tree.get_children():
                v = self.tree.item(child)["values"]
                c.execute("""INSERT INTO appendix2 
                    (project_id, s_no, category, item, commencement_months, completion_months, schedule_start, schedule_finish) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                    (self.project_id, v[0], v[1], v[2], v[3], v[4], to_storage_date(v[5]), to_storage_date(v[6])))
            
            conn.commit()
            conn.close()
            self.lock_everything()
            messagebox.showinfo("Success", "Contract saved successfully!")
            keep_window_active(self)
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save:\n{str(e)}")
            keep_window_active(self)

    def lock_everything(self):
        if self.main_app and hasattr(self.main_app, "is_admin") and self.main_app.is_admin():
            self.save_btn.config(state="normal")
            self.tree.bind("<Double-1>", lambda event: None)
            return
        self.contractor_entry.config(state="disabled", disabledbackground="#e8e8e8", disabledforeground="#333333")
        self.loa_entry.config(state="disabled", disabledbackground="#e8e8e8", disabledforeground="#333333")
        self.effective_entry.config(state="disabled", disabledbackground="#e8e8e8", disabledforeground="#333333")
        self.schedule_entry.config(state="disabled", disabledbackground="#e8e8e8", disabledforeground="#333333")
        self.save_btn.config(state="disabled")
        self.tree.tag_configure("locked", background="#e8e8e8", foreground="#333333")
        for child in self.tree.get_children():
            self.tree.item(child, tags=("locked",))
        self.tree.unbind("<Double-1>")

    def delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Select", "Please select a row to delete")
            return
        self.tree.delete(selected)
        messagebox.showinfo("Deleted", "Selected row removed!")
        keep_window_active(self)

    def go_home(self):
        self.destroy()
        if self.main_app:
            self.main_app.show_frame("registration")

    def pick_date(self, entry):
        cal_win = Toplevel(self)
        cal_win.title("Select Date")
        cal_win.geometry("320x300")
        cal = DateEntry(cal_win, width=25, date_pattern='dd-mm-yy', background='darkblue', foreground='white')
        cal.pack(pady=20)
        def set_d():
            entry.delete(0, END)
            entry.insert(0, cal.get_date().strftime("%d-%m-%y"))
            cal_win.destroy()
        Button(cal_win, text="Select This Date", command=set_d, bg="#003087", fg="white").pack(pady=10)
        normalize_buttons(cal_win)

    def calculate_completion(self):
        eff = self.effective_entry.get().strip()
        months = self.schedule_entry.get().strip()
        if eff and months.isdigit():
            self.completion_label.config(text=to_display_date(add_months(eff, int(months))))
        else:
            messagebox.showwarning("Input", "Enter Effective Date and Schedule Months")

    def calculate_all_dates(self):
        eff = self.effective_entry.get().strip()
        if not eff:
            messagebox.showwarning("Missing", "Enter Effective Date first")
            return
        for child in self.tree.get_children():
            values = list(self.tree.item(child)["values"])
            comm = int(values[3]) if str(values[3]).isdigit() else 0
            comp = int(values[4]) if str(values[4]).isdigit() else 0
            values[5] = add_months(eff, comm)
            values[6] = add_months(eff, comp)
            self.tree.item(child, values=values)
        messagebox.showinfo("Done", "All schedule dates calculated!")
        keep_window_active(self)

    def load_data(self):
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT contractor_name, loa_date, effective_date, schedule_months, schedule_completion FROM projects WHERE id=%s", (self.project_id,))
        row = c.fetchone()
        if row:
            self.contractor_entry.insert(0, row["contractor_name"] or "")
            self.loa_entry.insert(0, to_display_date(row["loa_date"]))
            self.effective_entry.insert(0, to_display_date(row["effective_date"]))
            self.schedule_entry.insert(0, str(row["schedule_months"]) if row["schedule_months"] else "")
            self.completion_label.config(text=to_display_date(row["schedule_completion"]) or "---")
        
        c.execute("""
            SELECT s_no, category, item, commencement_months, completion_months,
                   schedule_start, schedule_finish
            FROM appendix2
            WHERE project_id=%s
            ORDER BY s_no
        """, (self.project_id,))
        saved = c.fetchall()
        conn.close()
        
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        if saved:
            for r in saved:
                eff = self.effective_entry.get().strip()
                start = r["schedule_start"] or (add_months(eff, int(r['commencement_months'])) if eff and str(r['commencement_months']).isdigit() else "")
                finish = r["schedule_finish"] or (add_months(eff, int(r['completion_months'])) if eff and str(r['completion_months']).isdigit() else "")
                self.tree.insert("", END, values=(r['s_no'], r['category'] or "", r['item'], r['commencement_months'], r['completion_months'], to_display_date(start), to_display_date(finish)))
            self.lock_everything()
        else:
            self.save_btn.config(state="normal")

    def add_appendix_item(self):
        add_win = Toplevel(self)
        add_win.title("Add Appendix-2 Line Item")
        add_win.geometry("720x550")
        add_win.grab_set()
        
        Label(add_win, text="Select Parent → Child Item", font=("Arial", 12, "bold")).pack(pady=10)
        
        groups = self.get_appendix_groups()
        
        category_var = StringVar()
        category_combo = ttk.Combobox(add_win, textvariable=category_var, values=list(groups.keys()), width=45, height=8, state="readonly")
        category_combo.pack(pady=8, padx=40)
        
        item_var = StringVar()
        item_combo = ttk.Combobox(add_win, textvariable=item_var, width=60, height=12, state="readonly")
        item_combo.pack(pady=8, padx=40)

        item_list = Listbox(add_win, height=8, width=62, exportselection=False)
        item_list.pack(pady=(0, 8), padx=40)
        
        def update_items(*args):
            cat = category_var.get()
            items = groups.get(cat, [])
            item_combo["values"] = items
            item_combo.set("")
            item_list.delete(0, END)
            for child_item in items:
                item_list.insert(END, child_item)

        def choose_list_item(_event=None):
            selection = item_list.curselection()
            if selection:
                item_var.set(item_list.get(selection[0]))
        
        category_var.trace_add("write", update_items)
        category_combo.bind("<<ComboboxSelected>>", update_items)
        item_list.bind("<<ListboxSelect>>", choose_list_item)
        
        f = Frame(add_win)
        f.pack(pady=20)
        Label(f, text="Commencement Months:").grid(row=0, column=0, padx=10)
        comm_var = StringVar()
        Entry(f, textvariable=comm_var, width=12).grid(row=0, column=1)
        Label(f, text="Completion Months:").grid(row=0, column=2, padx=10)
        comp_var = StringVar()
        Entry(f, textvariable=comp_var, width=12).grid(row=0, column=3)
        
        def add_item():
            parent = category_var.get()
            item = item_var.get()
            comm = comm_var.get().strip()
            comp = comp_var.get().strip()
            if not parent or not item or not comm.isdigit() or not comp.isdigit():
                messagebox.showerror("Error", "Please select Parent, Child and fill months correctly")
                return
            s_no = len(self.tree.get_children()) + 1
            self.tree.insert("", END, values=(s_no, parent, item, comm, comp, "", ""))
            add_win.destroy()
        
        Button(add_win, text="➕ Add This Item", command=add_item, bg="#008000", fg="white", width=25).pack(pady=20)
        normalize_buttons(add_win)

    def refresh_data(self):
        self.contractor_entry.delete(0, END)
        self.loa_entry.delete(0, END)
        self.effective_entry.delete(0, END)
        self.schedule_entry.delete(0, END)
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.load_data()
        messagebox.showinfo("Refreshed", "Contract data refreshed successfully!")
        keep_window_active(self)


class ScurveWindow(Toplevel):
    def __init__(self, parent, project_id, uid, main_app=None):
        super().__init__(parent)
        self.project_id = project_id
        self.uid = uid
        self.main_app = main_app
        self.title(f"S Curve Planning - {uid}")
        self.geometry("1920x980")
        self.configure(bg="#f0f4f8")

        self.current_plan = None
        self.plan_list = []
        self.appendix_rows = []
        self.current_plan_saved = False

        main_container = Frame(self, bg="#f0f4f8")
        main_container.pack(fill=BOTH, expand=True, padx=15, pady=10)

        # LEFT: Plan Type
        left_frame = Frame(main_container, bg="#f0f4f8", width=240, relief="groove", bd=3)
        left_frame.pack(side=LEFT, fill=Y, padx=(0, 15))
        left_frame.pack_propagate(False)

        Label(left_frame, text="Plan Type", font=("Arial", 16, "bold"), bg="#f0f4f8", fg="#003087").pack(pady=15)

        self.plan_listbox = Listbox(left_frame, font=("Arial", 11), height=18, selectbackground="#0078d4")
        self.plan_listbox.pack(fill=BOTH, expand=True, padx=10, pady=5)
        self.plan_listbox.bind("<<ListboxSelect>>", self.on_plan_selected)

        # Add Plan + Delete Plan
        plan_btn_frame = Frame(left_frame, bg="#f0f4f8")
        plan_btn_frame.pack(pady=15, padx=15, fill=X)
        
        Button(plan_btn_frame, text="➕ Add Plan", command=self.add_new_plan,
               bg="#008000", fg="white", font=("Arial", 11, "bold"), height=2).pack(side=LEFT, padx=5, expand=True, fill=X)
        
        Button(plan_btn_frame, text="🗑️", command=self.delete_selected_plan,
               bg="#c8102e", fg="white", font=("Arial", 11, "bold"), height=2).pack(side=LEFT, padx=5, expand=True, fill=X)

        # RIGHT: Main Content
        right_frame = Frame(main_container, bg="#f0f4f8")
        right_frame.pack(side=LEFT, fill=BOTH, expand=True)
        self.right_frame = right_frame

        # Project dates
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT effective_date, schedule_completion FROM projects WHERE id=%s", (self.project_id,))
        proj = c.fetchone()
        self.effective_date = proj["effective_date"] if proj and proj["effective_date"] else "2026-04-01"
        self.schedule_completion = proj["schedule_completion"] if proj and proj["schedule_completion"] else "2028-12-31"
        conn.close()

        self.build_table(right_frame)
        self.load_plans()

        # ========== BUTTONS BELOW THE TABLE ==========
        bottom_btn_frame = Frame(right_frame, bg="#f0f4f8")
        bottom_btn_frame.pack(fill=X, pady=15)

        Button(bottom_btn_frame, text="➕ Add Activity", command=self.open_add_activity_popup,
               bg="#007580", fg="white", font=("Arial", 11, "bold"), height=2, width=15).pack(side=LEFT, padx=6)
        Button(bottom_btn_frame, text="💾 Save Activities", command=self.save_activities,
               bg="#0066cc", fg="white", font=("Arial", 11, "bold"), height=2, width=17).pack(side=LEFT, padx=6)
        Button(bottom_btn_frame, text="🗑️", command=self.delete_selected_row,
               bg="#c8102e", fg="white", font=("Arial", 11, "bold"), height=2, width=15).pack(side=LEFT, padx=6)
        Button(bottom_btn_frame, text="🔄 Refresh", command=self.refresh_current_plan,
               bg="#28a745", fg="white", font=("Arial", 11, "bold"), height=2, width=13).pack(side=LEFT, padx=6)
        self.add_plan_btn = plan_btn_frame.winfo_children()[0]
        self.delete_plan_btn = plan_btn_frame.winfo_children()[1]
        self.delete_plan_btn.config(text="🗑️", font=("Arial", 14, "bold"), width=8)
        self.add_plan_btn.pack_forget()
        self.delete_plan_btn.pack_forget()
        self.add_plan_btn.pack(side=LEFT, padx=(0, 6), fill=X, expand=True)
        self.delete_plan_btn.pack(side=LEFT, padx=(0, 0), fill=Y)
        self.add_activity_btn = bottom_btn_frame.winfo_children()[0]
        self.save_activities_btn = bottom_btn_frame.winfo_children()[1]
        self.delete_row_btn = bottom_btn_frame.winfo_children()[2]
        self.update_scurve_action_state()
        apply_page_watermark(self)
        normalize_buttons(self)
        self.add_plan_btn.config(width=13, height=2)
        self.delete_plan_btn.config(width=4, height=2)

    def load_plans(self):
        self.plan_listbox.delete(0, END)
        self.plan_list = []
        self.current_plan = None

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT plan_name FROM plans WHERE project_id=%s ORDER BY id", (self.project_id,))
        plans = c.fetchall()
        conn.close()

        for p in plans:
            plan_name = p["plan_name"]
            self.plan_list.append(plan_name)
            self.plan_listbox.insert(END, plan_name)

        if self.plan_list:
            self.plan_listbox.selection_set(0)
            self.current_plan = self.plan_list[0]
            self.load_activities_for_plan()
        else:
            self.clear_activity_table()
            self.current_plan_saved = False
            self.update_scurve_action_state()

    def add_new_plan(self):
        from utils import get_current_fy
        
        current_fy = get_current_fy()
        base_name = f"FY {current_fy}"
        
        # Count how many plans already exist with this FY
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT plan_name FROM plans WHERE project_id=%s AND plan_name LIKE %s", 
                  (self.project_id, f"{base_name}%"))
        existing = c.fetchall()
        conn.close()
        
        count = len(existing)
        
        if count == 0:
            new_plan = base_name                    # FY 2026-27
        else:
            new_plan = f"{base_name}-{count}"       # FY 2026-27-1, FY 2026-27-2, etc.
        
        plan_name = simpledialog.askstring("Add New Plan", "Enter Plan Name:", initialvalue=new_plan)
        if not plan_name or not plan_name.strip():
            return

        plan_name = plan_name.strip()

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT 1 FROM plans WHERE project_id=%s AND plan_name=%s", (self.project_id, plan_name))
        if c.fetchone():
            conn.close()
            messagebox.showwarning("Duplicate Plan", f"Plan '{plan_name}' already exists for this project.")
            return
        c.execute("INSERT INTO plans (project_id, plan_name) VALUES (%s, %s)", (self.project_id, plan_name))
        conn.commit()
        conn.close()

        self.load_plans()
        messagebox.showinfo("Success", f"Plan '{plan_name}' created successfully!")
        keep_window_active(self)

    def delete_selected_plan(self):
        if not self.current_plan:
            messagebox.showwarning("Select Plan", "Please select a plan to delete!")
            keep_window_active(self)
            return
        if self.current_plan_saved and not self.is_admin_user():
            messagebox.showwarning("Plan Locked", "This plan type is already saved. Only Admin can delete saved plan types.")
            keep_window_active(self)
            return
        
        if not messagebox.askyesno("Confirm Delete", f"Delete plan '{self.current_plan}' and all its activities?"):
            return

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("DELETE FROM daily_actuals WHERE activity_id IN (SELECT id FROM activities WHERE project_id=%s AND plan_name=%s)", (self.project_id, self.current_plan))
        c.execute("DELETE FROM monthly_plans WHERE project_id=%s AND plan_name=%s", (self.project_id, self.current_plan))
        c.execute("DELETE FROM activities WHERE project_id=%s AND plan_name=%s", (self.project_id, self.current_plan))
        c.execute("DELETE FROM plans WHERE project_id=%s AND plan_name=%s", (self.project_id, self.current_plan))
        conn.commit()
        conn.close()

        messagebox.showinfo("Deleted", f"Plan '{self.current_plan}' deleted!")
        self.load_plans()
        keep_window_active(self)

    def on_plan_selected(self, event):
        selection = self.plan_listbox.curselection()
        if selection:
            self.current_plan = self.plan_listbox.get(selection[0])
            self.load_activities_for_plan()

    def load_activities_for_plan(self):
        if not self.current_plan:
            return

        self.clear_activity_table()

        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""SELECT id, activity_type, uom, scope_qty, weight_percent, actuals_till_last_fy, start_date, finish_date,
                            COALESCE(NULLIF(expected_finish, ''), finish_date) AS expected_finish
                     FROM activities 
                     WHERE project_id=%s AND plan_name=%s 
                     ORDER BY id""", (self.project_id, self.current_plan))
        activities = c.fetchall()
        planning_locked = self.plan_has_saved_planning() and not self.is_admin_user()

        for act in activities:
            c.execute("""
                SELECT month, planned_qty
                FROM monthly_plans
                WHERE project_id=%s AND plan_name=%s AND activity_type=%s
            """, (self.project_id, self.current_plan, act["activity_type"]))
            monthly_values = {r["month"]: r["planned_qty"] for r in c.fetchall()}
            scope_qty = float(act["scope_qty"] or 0)
            actual_last_fy = float(act.get("actuals_till_last_fy") or 0)
            planned_total = sum(float(v or 0) for v in monthly_values.values())
            remaining_qty = scope_qty - actual_last_fy - planned_total
            values = [
                f"{self.parse_weight(act.get('weight_percent') or 10):.2f}%",
                act["activity_type"],
                act["uom"],
                str(act["scope_qty"]),
                f"{remaining_qty:.2f}",
                to_display_date(act["start_date"]),
                to_display_date(act["finish_date"]),
                f"{actual_last_fy:.2f}",
            ]
            for month in self.month_columns:
                allowed, _ = self.is_month_allowed_for_activity(month, values)
                if month in monthly_values:
                    values.append(f"{float(monthly_values[month] or 0):.2f}")
                elif allowed:
                    values.append(self.locked_marker() if planning_locked else self.editable_marker())
                else:
                    values.append("")
            values = self.recalculate_remaining_qty(values)
            self.tree.insert("", "end", iid=str(act["id"]), values=values, tags=(str(act["id"]),))
        conn.close()
        self.current_plan_saved = bool(activities)
        self.update_scurve_action_state()

    def is_admin_user(self):
        return bool(self.main_app and hasattr(self.main_app, "is_admin") and self.main_app.is_admin())

    def is_user_saved_plan_locked(self):
        return self.current_plan_saved and self.main_app and not self.is_admin_user()

    def editable_marker(self):
        return "🔲"

    def locked_marker(self):
        return "⬛"

    def planning_cell_qty(self, value):
        text = str(value or "").strip()
        if not text or text in (self.editable_marker(), self.locked_marker()):
            return 0.0
        return float(text)

    def recalculate_remaining_qty(self, values):
        values = list(values)
        try:
            scope = float(values[3] or 0)
        except Exception:
            scope = 0.0
        try:
            actual_last_fy = float(values[7] or 0)
        except Exception:
            actual_last_fy = 0.0
        planned_total = 0.0
        for idx in range(8, len(values)):
            planned_total += self.planning_cell_qty(values[idx])
        values[4] = f"{scope - actual_last_fy - planned_total:.2f}"
        return values

    def plan_has_saved_planning(self):
        if not self.current_plan:
            return False
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""
            SELECT 1 FROM monthly_plans
            WHERE project_id=%s AND plan_name=%s
            LIMIT 1
        """, (self.project_id, self.current_plan))
        row = c.fetchone()
        conn.close()
        return bool(row)

    def update_scurve_action_state(self):
        if hasattr(self, "add_activity_btn"):
            locked = self.is_user_saved_plan_locked()
            locked_bg = "#9e9e9e"
            locked_fg = "#eeeeee"
            normal_states = [
                (self.add_activity_btn, "#007580"),
                (self.delete_row_btn, "#c8102e"),
                (self.delete_plan_btn, "#c8102e"),
            ]
            self.add_activity_btn.config(state=DISABLED if locked else NORMAL)
            self.save_activities_btn.config(state=NORMAL)
            self.delete_row_btn.config(state=DISABLED if locked else NORMAL)
            can_delete_plan = bool(self.current_plan) and not locked
            self.delete_plan_btn.config(state=NORMAL if can_delete_plan else DISABLED)
            for btn, normal_bg in normal_states:
                btn.config(bg=locked_bg if locked else normal_bg, fg=locked_fg if locked else "white")
            self.save_activities_btn.config(
                text="💾 Save Planning" if self.current_plan_saved else "💾 Save Activities",
                bg="#0066cc",
                fg="white"
            )

    def is_original_plan(self):
        return str(self.current_plan or "").strip().lower() in ["original plan", "orignal plan", "original", "orignal"]

    def get_plan_fy_start(self):
        if self.is_original_plan():
            return None
        import re
        match = re.search(r"(20\d{2})", str(self.current_plan or ""))
        year = int(match.group(1)) if match else datetime.now().year
        return datetime(year, 4, 1).date()

    def month_label_to_date(self, month_label):
        return datetime.strptime(str(month_label), "%b-%y").date().replace(day=1)

    def is_month_allowed_for_activity(self, month_label, values):
        try:
            month_date = self.month_label_to_date(month_label)
            start_date = parse_app_date(values[5]).replace(day=1)
            finish_date = parse_app_date(values[6]).replace(day=1)
        except Exception:
            return False, "Activity schedule start/finish date is missing or invalid."

        if month_date < start_date or month_date > finish_date:
            return False, "Planning is allowed only between this activity's Schedule Start and Schedule Finish."

        fy_start = self.get_plan_fy_start()
        if fy_start and month_date < fy_start:
            last_fy_end = fy_start.replace(month=3, day=31)
            return False, f"Use 'Actuals Till Last FY' for quantities up to 31-Mar-{last_fy_end.year}."

        return True, ""

    def clear_activity_table(self):
        if hasattr(self, "tree"):
            for item in self.tree.get_children():
                self.tree.delete(item)

    def load_appendix_rows(self):
        self.appendix_rows = [dict(row) for row in get_appendix_activity_rows(self.project_id)]
        return self.appendix_rows

    def get_default_project_dates(self):
        return self.effective_date or "", self.schedule_completion or ""

    def format_activity_type(self, category, item):
        category = (category or "").strip()
        item = (item or "").strip()
        if category and item:
            return f"{category} -> {item}"
        return item or category

    def get_matching_appendix_row(self, category, item):
        category = (category or "").strip().lower()
        item = (item or "").strip().lower()
        for row in self.appendix_rows:
            if (row.get("category") or "").strip().lower() == category and (row.get("item") or "").strip().lower() == item:
                return row
        return None

    def parse_weight(self, value):
        try:
            return float(str(value).replace("%", "").strip() or 0)
        except Exception:
            return 0.0

    def get_total_weight(self, exclude_item=None):
        total = 0.0
        for child in self.tree.get_children():
            if exclude_item and child == exclude_item:
                continue
            values = self.tree.item(child)["values"]
            if values:
                total += self.parse_weight(values[0])
        return total

    def validate_total_weight(self, new_weight, exclude_item=None):
        total = self.get_total_weight(exclude_item) + self.parse_weight(new_weight)
        if total > 100:
            messagebox.showerror(
                "Weight Limit",
                f"Total Wtg. % cannot exceed 100%.\n\nCurrent total: {self.get_total_weight(exclude_item):.2f}%\nNew total: {total:.2f}%"
            )
            keep_window_active(self)
            return False
        return True

    def build_table(self, parent):
        self.month_columns = []
        start = parse_app_date(self.effective_date)
        end = parse_app_date(self.schedule_completion)
        if not start:
            start = datetime(2026, 4, 1).date()
        if not end:
            end = datetime(2028, 12, 31).date()
        current = start.replace(day=1)
        while current <= end:
            self.month_columns.append(current.strftime("%b-%y"))
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

        cols = ["Wtg. %", "Activity Type", "UOM", "Scope Qty", "Remaining Qty", "Schedule Start", "Schedule Finish", 
                "Actuals Till Last FY"] + self.month_columns

        table_frame = Frame(parent, bg="#f0f4f8")
        table_frame.pack(fill=BOTH, expand=True, padx=5, pady=5)

        scurve_style = ttk.Style(self)
        scurve_style.configure("Scurve.Treeview", font=("Arial", 11), rowheight=30)
        scurve_style.configure("Scurve.Treeview.Heading", font=("Arial", 11, "bold"))

        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=18, style="Scurve.Treeview")
        for col in cols:
            self.tree.heading(col, text=col)
            if col == "Activity Type":
                width = 260
            elif col in ["Wtg. %", "UOM", "Scope Qty", "Remaining Qty"]:
                width = 105
            elif col in ["Schedule Start", "Schedule Finish", "Actuals Till Last FY"]:
                width = 145
            else:
                width = 120
            self.tree.column(col, width=width, minwidth=width, anchor="center", stretch=False)

        hscroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        hscroll.pack(side=BOTTOM, fill=X)
        vscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        vscroll.pack(side=RIGHT, fill=Y)
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        self.tree.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)
        self.tree.bind("<Double-1>", self.edit_cell)

    def open_add_activity_popup(self):
        if not self.current_plan:
            messagebox.showwarning("Select Plan", "Please select or create a plan first!")
            return
        if self.is_user_saved_plan_locked():
            messagebox.showwarning("Plan Locked", "This plan is already saved. User cannot add more activities.")
            keep_window_active(self)
            return
        appendix_rows = self.load_appendix_rows()
        category_map = {}
        for row in appendix_rows:
            category = (row.get("category") or "").strip()
            item = (row.get("item") or "").strip()
            if not category or not item:
                continue
            category_map.setdefault(category, [])
            if item not in category_map[category]:
                category_map[category].append(item)

        popup = Toplevel(self)
        popup.title("Add Activity")
        popup.geometry("760x690")
        popup.grab_set()
        popup.configure(bg="#f0f4f8")

        Label(popup, text=f"Add Activity to {self.current_plan}",
              font=("Arial", 16, "bold"), bg="#f0f4f8", fg="#003087").pack(pady=15)

        f = Frame(popup, bg="#f0f4f8")
        f.pack(pady=10, padx=30, fill=X)
        f.grid_columnconfigure(1, weight=1)

        Label(f, text="Parent Activity:").grid(row=0, column=0, sticky=W, pady=6)
        category_var = StringVar()
        category_combo = ttk.Combobox(f, textvariable=category_var, values=list(category_map.keys()), width=42, state="normal")
        category_combo.grid(row=0, column=1, pady=6, sticky=W)

        Label(f, text="Child Activity / New Activity:").grid(row=1, column=0, sticky=W, pady=6)
        item_var = StringVar()
        item_combo = ttk.Combobox(f, textvariable=item_var, width=42, state="normal")
        item_combo.grid(row=1, column=1, pady=6, sticky=W)

        Label(f, text="Unit of Measurement:").grid(row=2, column=0, sticky=W, pady=6)
        uoms = ["Nos.", "MT", "Cum", "Rmtr", "Set", "IM"]
        uom_var = StringVar()
        ttk.Combobox(f, textvariable=uom_var, values=uoms, width=42, state="readonly").grid(row=2, column=1, pady=6, sticky=W)

        Label(f, text="Scope Qty:").grid(row=3, column=0, sticky=W, pady=6)
        scope_var = StringVar()
        Entry(f, textvariable=scope_var, width=15).grid(row=3, column=1, pady=6, sticky=W)

        Label(f, text="Wtg. %:").grid(row=4, column=0, sticky=W, pady=6)
        wtg_var = StringVar(value="10")
        Entry(f, textvariable=wtg_var, width=10).grid(row=4, column=1, pady=6, sticky=W)

        used_wtg = self.get_total_weight()
        Label(
            f,
            text=f"Wtg already added: {used_wtg:.2f}%    Remaining: {max(0, 100 - used_wtg):.2f}%",
            bg="#f0f4f8",
            fg="#c8102e" if used_wtg >= 100 else "#003087",
            font=("Arial", 10, "bold")
        ).grid(row=5, column=1, columnspan=2, sticky=W, pady=(0, 8))

        Label(f, text="Schedule Start (Auto DD-MM-YY):", fg="blue").grid(row=6, column=0, sticky=W, pady=6)
        start_var = StringVar()
        Entry(f, textvariable=start_var, width=25, state="normal").grid(row=6, column=1, pady=6, sticky=W)
        Button(f, text="📅", width=3, command=lambda: self.pick_date_popup(start_var, popup)).grid(row=6, column=2, padx=5)

        Label(f, text="Schedule Finish (Auto DD-MM-YY):", fg="blue").grid(row=7, column=0, sticky=W, pady=6)
        finish_var = StringVar()
        Entry(f, textvariable=finish_var, width=25, state="normal").grid(row=7, column=1, pady=6, sticky=W)
        Button(f, text="📅", width=3, command=lambda: self.pick_date_popup(finish_var, popup)).grid(row=7, column=2, padx=5)

        def auto_fill_dates():
            selected_row = self.get_matching_appendix_row(category_var.get(), item_var.get())
            if selected_row:
                start_value = selected_row.get("schedule_start")
                finish_value = selected_row.get("schedule_finish")

                if not start_value and self.effective_date and selected_row.get("commencement_months") is not None:
                    start_value = add_months(self.effective_date, int(selected_row["commencement_months"]))
                if not finish_value and self.effective_date and selected_row.get("completion_months") is not None:
                    finish_value = add_months(self.effective_date, int(selected_row["completion_months"]))

                start_var.set(to_display_date(start_value))
                finish_var.set(to_display_date(finish_value))
                return

            default_start, default_finish = self.get_default_project_dates()
            start_var.set(to_display_date(default_start))
            finish_var.set(to_display_date(default_finish))

        def update_items(*args):
            category = category_var.get().strip()
            item_combo["values"] = category_map.get(category, [])
            if category_map.get(category) and item_var.get().strip() not in category_map.get(category, []):
                item_var.set("")
            auto_fill_dates()

        category_var.trace_add("write", update_items)
        item_var.trace_add("write", lambda *args: auto_fill_dates())

        Button(f, text="📅 Auto Fill from Contract & Appendix-2",
               command=auto_fill_dates, bg="#003087", fg="white",
               font=("Arial", 10)).grid(row=8, column=0, columnspan=3, pady=10, sticky="ew")

        if not appendix_rows:
            Label(
                f,
                text="No Appendix-2 rows found. You can still type a new parent/activity manually.",
                fg="#a33",
                bg="#f0f4f8",
                font=("Arial", 9, "italic"),
            ).grid(row=9, column=0, columnspan=3, sticky=W, pady=(2, 0))

        popup.after(300, auto_fill_dates)

        def add_and_refresh():
            category = category_var.get().strip()
            item = item_var.get().strip()
            act_type = self.format_activity_type(category, item)
            uom = uom_var.get()
            scope = scope_var.get().strip()
            wtg = wtg_var.get().strip() or "10"
            start = start_var.get().strip()
            finish = finish_var.get().strip()

            if not all([act_type, uom, scope, start, finish]):
                messagebox.showerror("Error", "All fields are mandatory")
                return
            if not self.validate_total_weight(wtg):
                return

            values = [f"{self.parse_weight(wtg):.2f}%", act_type, uom, scope, scope, start, finish, "0.00"] + [""] * len(self.month_columns)
            draft_iid = f"draft_{len(self.tree.get_children()) + 1}"
            values = self.recalculate_remaining_qty(values)
            self.tree.insert("", "end", iid=draft_iid, values=values, tags=("draft",))

            popup.destroy()
            messagebox.showinfo("Draft Added", "Activity added as draft. Click 'Save Activities' to save it permanently.")
            keep_window_active(self)

        Button(popup, text="➕ Add Activity", command=add_and_refresh, bg="#008000", fg="white",
               font=("Arial", 12, "bold"), height=2, width=25).pack(pady=25)
        normalize_buttons(popup)

    def get_month_range_labels(self, start_date, finish_date):
        start = start_date.replace(day=1)
        finish = finish_date.replace(day=1)
        labels = []
        current = start
        while current <= finish:
            labels.append(current.strftime("%b-%y"))
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)
        return labels

    def load_saved_plan_activities(self):
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""
            SELECT id, activity_type, uom, scope_qty, weight_percent, start_date, finish_date,
                   COALESCE(NULLIF(expected_finish, ''), finish_date) AS expected_finish
            FROM activities
            WHERE project_id=%s AND plan_name=%s
            ORDER BY id
        """, (self.project_id, self.current_plan))
        activities = [dict(row) for row in c.fetchall()]

        c.execute("""
            SELECT activity_type, month, planned_qty
            FROM monthly_plans
            WHERE project_id=%s AND plan_name=%s
        """, (self.project_id, self.current_plan))
        monthly_values = {}
        for row in c.fetchall():
            monthly_values.setdefault(row["activity_type"], {})[row["month"]] = float(row["planned_qty"] or 0)
        fy_start = self.get_plan_fy_start()
        last_fy_end = fy_start.replace(month=3, day=31) if fy_start else None
        for act in activities:
            if last_fy_end:
                c.execute("""
                    SELECT COALESCE(SUM(actual_qty), 0) AS total
                    FROM daily_actuals
                    WHERE activity_id=%s AND actual_date <= %s
                """, (act["id"], last_fy_end.strftime("%Y-%m-%d")))
                row = c.fetchone()
                act["actuals_till_last_fy"] = float(row["total"] or 0) if row else 0
            else:
                act["actuals_till_last_fy"] = 0
        conn.close()
        return activities, monthly_values

    def open_planning_popup(self):
        if not self.current_plan:
            messagebox.showwarning("Select Plan", "Please select a plan first!")
            keep_window_active(self)
            return
        if not self.current_plan_saved:
            messagebox.showwarning("Save Activities First", "Save activities first, then open Planning.")
            keep_window_active(self)
            return

        activities, monthly_values = self.load_saved_plan_activities()
        if not activities:
            messagebox.showwarning("No Activities", "No saved activities found for this plan.")
            keep_window_active(self)
            return

        date_pairs = []
        for act in activities:
            try:
                start = parse_app_date(act["start_date"])
                finish = parse_app_date(act.get("expected_finish") or act["finish_date"])
                if not start or not finish:
                    raise ValueError
                date_pairs.append((start, finish))
            except Exception:
                pass
        if not date_pairs:
            messagebox.showwarning("Schedule Missing", "Saved activities do not have valid Schedule Start and Finish dates.")
            keep_window_active(self)
            return

        month_cols = self.get_month_range_labels(min(p[0] for p in date_pairs), max(p[1] for p in date_pairs))
        popup = Toplevel(self)
        popup.title(f"Planning - {self.current_plan}")
        popup.geometry("1500x760")
        popup.configure(bg="#f0f4f8")
        popup.grab_set()

        Label(
            popup,
            text=f"Planning for {self.current_plan}",
            bg="#f0f4f8",
            fg="#003087",
            font=("Arial", 18, "bold"),
        ).pack(pady=(12, 6))
        Label(
            popup,
            text="Double-click blue-grey cells to enter planning quantities. Grey cells are locked after user save.",
            bg="#f0f4f8",
            fg="#333333",
            font=("Arial", 10, "bold"),
        ).pack(pady=(0, 8))

        table_frame = Frame(popup, bg="#f0f4f8")
        table_frame.pack(fill=BOTH, expand=True, padx=15, pady=8)

        planning_style = ttk.Style(popup)
        planning_style.configure(
            "Planning.Treeview",
            font=("Arial", 12),
            rowheight=38,
            background="#fffaf2",
            fieldbackground="#fffaf2",
            borderwidth=1,
        )
        planning_style.configure(
            "Planning.Treeview.Heading",
            font=("Arial", 11, "bold"),
            foreground="#1f2933",
            background="#ffe9d6",
        )
        planning_style.map(
            "Planning.Treeview",
            background=[("selected", "#0b79d0")],
            foreground=[("selected", "white")],
        )

        fixed_cols = [
            "Activity ID", "Wtg. %", "Activity Type", "UOM", "Scope Qty",
            "Schedule Start", "Schedule Finish", "Expected Finish", "Editable Months", "Actual Upto Last FY",
            "Planned Total", "Balance",
        ]
        planning_cols = fixed_cols + month_cols
        editable_marker = "🔲"
        locked_marker = "⬛"
        user_planning_locked = bool(monthly_values) and not self.is_admin_user()

        def planning_cell_qty(value):
            text = str(value or "").strip()
            if not text or text in (editable_marker, locked_marker):
                return 0.0
            return float(text)

        planning_tree = ttk.Treeview(table_frame, columns=planning_cols, show="headings", height=18, style="Planning.Treeview")

        vscroll = ttk.Scrollbar(table_frame, orient=VERTICAL, command=planning_tree.yview)
        hscroll = ttk.Scrollbar(table_frame, orient=HORIZONTAL, command=planning_tree.xview)
        planning_tree.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)
        hscroll.pack(side=BOTTOM, fill=X)
        vscroll.pack(side=RIGHT, fill=Y)
        planning_tree.pack(side=LEFT, fill=BOTH, expand=True)

        widths = {
            "Activity ID": 0,
            "Wtg. %": 75,
            "Activity Type": 460,
            "UOM": 75,
            "Scope Qty": 95,
            "Schedule Start": 110,
            "Schedule Finish": 110,
            "Expected Finish": 115,
            "Editable Months": 170,
            "Actual Upto Last FY": 150,
            "Planned Total": 110,
            "Balance": 100,
        }
        for col in planning_cols:
            planning_tree.heading(col, text=col)
            planning_tree.column(col, width=widths.get(col, 115), minwidth=0 if col == "Activity ID" else 70, stretch=False)
        planning_tree.column("Activity ID", width=0, minwidth=0, stretch=False)
        planning_tree.tag_configure("plan_even", background="#fff7ec")
        planning_tree.tag_configure("plan_odd", background="#fff0f5")
        planning_tree.tag_configure("expected_extension", background="#ffedd5")

        activity_schedules = {}
        for row_index, act in enumerate(activities):
            iid = str(act["id"])
            activity_type = act["activity_type"]
            scope = float(act["scope_qty"] or 0)
            monthly = monthly_values.get(activity_type, {})
            try:
                start_date = parse_app_date(act["start_date"])
                finish_date = parse_app_date(act["finish_date"])
                expected_finish = parse_app_date(act.get("expected_finish") or act["finish_date"])
            except Exception:
                start_date = None
                finish_date = None
                expected_finish = None
            month_values = []
            planned_total = 0
            actual_last_fy = float(act.get("actuals_till_last_fy") or 0)
            editable_months = ""
            planning_finish = expected_finish or finish_date
            if start_date and planning_finish:
                editable_months = f"{start_date.strftime('%b-%y')} to {planning_finish.strftime('%b-%y')}"
            for month in month_cols:
                in_schedule = bool(
                    start_date
                    and planning_finish
                    and start_date.replace(day=1) <= self.month_label_to_date(month) <= planning_finish.replace(day=1)
                )
                value = float(monthly.get(month, 0) or 0) if in_schedule else 0
                planned_total += value
                marker = locked_marker if user_planning_locked else editable_marker
                month_values.append(f"{value:.2f}" if value else (marker if in_schedule else ""))
            balance = scope - actual_last_fy - planned_total
            fixed_values = [
                act["id"],
                f"{float(act.get('weight_percent') or 0):.2f}%",
                activity_type,
                act["uom"],
                f"{scope:.2f}",
                to_display_date(act["start_date"]),
                to_display_date(act["finish_date"]),
                to_display_date(act.get("expected_finish") or act["finish_date"]),
                editable_months,
                f"{actual_last_fy:.2f}",
                f"{planned_total:.2f}",
                f"{balance:.2f}",
            ] + month_values
            row_tags = [iid, "plan_even" if row_index % 2 == 0 else "plan_odd"]
            if start_date and finish_date and expected_finish and expected_finish.replace(day=1) > finish_date.replace(day=1):
                row_tags.append("expected_extension")
            planning_tree.insert("", END, iid=iid, values=fixed_values, tags=tuple(row_tags))
            try:
                activity_schedules[iid] = (
                    parse_app_date(act["start_date"]),
                    parse_app_date(act.get("expected_finish") or act["finish_date"]),
                )
            except Exception:
                activity_schedules[iid] = (None, None)

        def month_allowed_for_row(item_id, month_label):
            start, finish = activity_schedules.get(item_id, (None, None))
            if not start or not finish:
                return False
            month_date = self.month_label_to_date(month_label)
            return start.replace(day=1) <= month_date <= finish.replace(day=1)

        def edit_planning_cell(event):
            if planning_tree.identify("region", event.x, event.y) != "cell":
                return
            column = planning_tree.identify_column(event.x)
            item = planning_tree.identify_row(event.y)
            if not item:
                return
            col_idx = int(column[1:]) - 1
            col_name = planning_tree["columns"][col_idx]
            if col_name not in month_cols:
                return
            if user_planning_locked:
                messagebox.showwarning("Planning Locked", "Planning is already saved. User cannot edit saved planning.")
                keep_window_active(popup)
                return
            if not month_allowed_for_row(item, col_name):
                messagebox.showwarning("Planning Not Allowed", "Planning is allowed only between this activity's Schedule Start and Schedule Finish.")
                keep_window_active(popup)
                return

            x, y, w, h = planning_tree.bbox(item, column)
            current_value = planning_tree.item(item)["values"][col_idx]
            entry = Entry(
                planning_tree,
                justify="center",
                font=("Arial", 12, "bold"),
                bg="#fffdf7",
                fg="#1f2933",
                relief="solid",
                bd=1,
                highlightthickness=2,
                highlightbackground="#f0b36a",
                highlightcolor="#f0b36a",
            )
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, "" if str(current_value).strip() in (editable_marker, locked_marker) else current_value)
            entry.focus()
            entry.select_range(0, END)

            def save_cell(e=None):
                try:
                    new_val = float(entry.get() or 0)
                except Exception:
                    new_val = 0
                if new_val < 0:
                    new_val = 0

                values = list(planning_tree.item(item)["values"])
                scope = float(values[4] or 0)
                actual_last_fy = float(values[9] or 0)
                total = 0
                for idx in range(len(fixed_cols), len(values)):
                    total += new_val if idx == col_idx else planning_cell_qty(values[idx])
                if actual_last_fy + total > scope:
                    entry.destroy()
                    messagebox.showerror(
                        "Scope Limit",
                        f"Actual Upto Last FY + monthly planning cannot exceed Scope Qty {scope:.2f}."
                    )
                    keep_window_active(popup)
                    return

                values[col_idx] = f"{new_val:.2f}" if new_val else editable_marker
                values[10] = f"{total:.2f}"
                values[11] = f"{scope - actual_last_fy - total:.2f}"
                planning_tree.item(item, values=values)
                entry.destroy()

            entry.bind("<Return>", save_cell)
            entry.bind("<FocusOut>", save_cell)

        planning_tree.bind("<Double-1>", edit_planning_cell)

        def save_planning():
            nonlocal user_planning_locked
            conn = get_db_connection()
            c = conn.cursor()
            for item in planning_tree.get_children():
                values = list(planning_tree.item(item)["values"])
                activity_type = values[2]
                c.execute("""
                    DELETE FROM monthly_plans
                    WHERE project_id=%s AND plan_name=%s AND activity_type=%s
                """, (self.project_id, self.current_plan, activity_type))
                for idx, month in enumerate(month_cols, start=len(fixed_cols)):
                    planned_qty = planning_cell_qty(values[idx])
                    if planned_qty:
                        c.execute("""
                            INSERT INTO monthly_plans (project_id, plan_name, activity_type, month, planned_qty, row_type)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (self.project_id, self.current_plan, activity_type, month, planned_qty, "plan"))
            conn.commit()
            conn.close()
            if not self.is_admin_user():
                user_planning_locked = True
                for item in planning_tree.get_children():
                    values = list(planning_tree.item(item)["values"])
                    for idx in range(len(fixed_cols), len(values)):
                        if str(values[idx]).strip() == editable_marker:
                            values[idx] = locked_marker
                    planning_tree.item(item, values=values)
            self.load_activities_for_plan()
            messagebox.showinfo("Saved", "Planning saved successfully.")
            keep_window_active(popup)

        btn_frame = Frame(popup, bg="#f0f4f8")
        btn_frame.pack(fill=X, pady=12)
        Button(btn_frame, text="💾 Save Planning", command=save_planning,
               bg="#0066cc", fg="white", font=("Arial", 11, "bold")).pack(side=LEFT, padx=(15, 8))
        Button(btn_frame, text="🔄 Refresh", command=lambda: (popup.destroy(), self.open_planning_popup()),
               bg="#28a745", fg="white", font=("Arial", 11, "bold")).pack(side=LEFT, padx=8)
        Button(btn_frame, text="Close", command=popup.destroy,
               bg="#555555", fg="white", font=("Arial", 11, "bold")).pack(side=RIGHT, padx=15)
        normalize_buttons(popup)

    def pick_date_popup(self, var, parent):
        cal_win = Toplevel(parent)
        cal_win.title("Select Date")
        cal_win.geometry("320x300")
        cal = DateEntry(cal_win, width=25, date_pattern='dd-mm-yy', background='darkblue', foreground='white')
        cal.pack(pady=20)
        def set_d():
            var.set(cal.get_date().strftime("%d-%m-%y"))
            cal_win.destroy()
        Button(cal_win, text="Select This Date", command=set_d, bg="#003087", fg="white").pack(pady=10)
        normalize_buttons(cal_win)

    def save_activities(self):
        if not self.current_plan:
            messagebox.showwarning("Select Plan", "Please select a plan first!")
            return
        if self.get_total_weight() > 100:
            messagebox.showerror("Weight Limit", f"Total Wtg. % cannot exceed 100%.\n\nCurrent total: {self.get_total_weight():.2f}%")
            keep_window_active(self)
            return

        conn = get_db_connection()
        c = conn.cursor()
        existing_ids = []

        for child in self.tree.get_children():
            values = self.recalculate_remaining_qty(self.tree.item(child)["values"])
            self.tree.item(child, values=values)
            tags = self.tree.item(child).get("tags", ())
            activity_id = int(tags[0]) if tags and str(tags[0]).isdigit() else None
            payload = (
                values[1],
                values[2],
                float(values[3]) if values[3] else 0,
                self.parse_weight(values[0]),
                float(values[7]) if values[7] else 0,
                to_storage_date(values[5]),
                to_storage_date(values[6]),
            )

            if activity_id:
                existing_ids.append(activity_id)
                c.execute("""
                    UPDATE activities
                    SET activity_type=%s, uom=%s, scope_qty=%s, weight_percent=%s, actuals_till_last_fy=%s, start_date=%s, finish_date=%s
                    WHERE id=%s AND project_id=%s AND plan_name=%s
                """, payload + (activity_id, self.project_id, self.current_plan))
            else:
                c.execute("""INSERT INTO activities 
                            (project_id, plan_name, activity_type, uom, scope_qty, weight_percent, actuals_till_last_fy, start_date, finish_date) 
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            RETURNING id""",
                          (self.project_id, self.current_plan, values[1], values[2], 
                           float(values[3]) if values[3] else 0, self.parse_weight(values[0]), float(values[7]) if values[7] else 0, to_storage_date(values[5]), to_storage_date(values[6])))
                new_id = c.fetchone()["id"]
                self.tree.item(child, tags=(str(new_id),))
                existing_ids.append(new_id)

            c.execute("""
                DELETE FROM monthly_plans
                WHERE project_id=%s AND plan_name=%s AND activity_type=%s
            """, (self.project_id, self.current_plan, values[1]))
            planned_total = 0.0
            for month_idx, month in enumerate(self.month_columns, start=8):
                planned_qty = self.planning_cell_qty(values[month_idx])
                planned_total += planned_qty
                if planned_qty:
                    allowed, reason = self.is_month_allowed_for_activity(month, values)
                    if not allowed:
                        conn.rollback()
                        conn.close()
                        messagebox.showerror("Planning Not Allowed", f"{values[1]} / {month}: {reason}")
                        keep_window_active(self)
                        return
                    c.execute("""
                        INSERT INTO monthly_plans (project_id, plan_name, activity_type, month, planned_qty, row_type)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (self.project_id, self.current_plan, values[1], month, planned_qty, "plan"))
            if float(values[7] or 0) + planned_total > float(values[3] or 0):
                conn.rollback()
                conn.close()
                messagebox.showerror("Scope Limit", f"{values[1]}: Actuals Till Last FY + monthly planning cannot exceed Scope Qty.")
                keep_window_active(self)
                return

        if existing_ids:
            c.execute("""
                DELETE FROM daily_actuals
                WHERE activity_id IN (
                    SELECT id FROM activities
                    WHERE project_id=%s AND plan_name=%s AND id <> ALL(%s)
                )
            """, (self.project_id, self.current_plan, existing_ids))
            c.execute("""
                DELETE FROM activities
                WHERE project_id=%s AND plan_name=%s AND id <> ALL(%s)
            """, (self.project_id, self.current_plan, existing_ids))
            c.execute("""
                DELETE FROM monthly_plans
                WHERE project_id=%s AND plan_name=%s AND activity_type NOT IN (
                    SELECT activity_type FROM activities WHERE project_id=%s AND plan_name=%s
                )
            """, (self.project_id, self.current_plan, self.project_id, self.current_plan))
        else:
            c.execute("DELETE FROM daily_actuals WHERE activity_id IN (SELECT id FROM activities WHERE project_id=%s AND plan_name=%s)", (self.project_id, self.current_plan))
            c.execute("DELETE FROM activities WHERE project_id=%s AND plan_name=%s", (self.project_id, self.current_plan))
            c.execute("DELETE FROM monthly_plans WHERE project_id=%s AND plan_name=%s", (self.project_id, self.current_plan))

        conn.commit()
        conn.close()
        self.current_plan_saved = True
        self.load_activities_for_plan()
        self.update_scurve_action_state()
        messagebox.showinfo("Success", f"Activities for '{self.current_plan}' saved successfully!")
        keep_window_active(self)

    def delete_selected_row(self):
        if self.is_user_saved_plan_locked():
            messagebox.showwarning("Plan Locked", "This plan is already saved. User cannot delete activities.")
            keep_window_active(self)
            return
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Select", "Please select a row to delete")
            return

        conn = get_db_connection()
        c = conn.cursor()
        for item_id in selected:
            tags = self.tree.item(item_id).get("tags", ())
            activity_id = int(tags[0]) if tags and str(tags[0]).isdigit() else None
            if activity_id:
                c.execute("DELETE FROM daily_actuals WHERE activity_id=%s", (activity_id,))
                c.execute("DELETE FROM activities WHERE id=%s AND project_id=%s AND plan_name=%s", (activity_id, self.project_id, self.current_plan))
        conn.commit()
        conn.close()

        self.tree.delete(*selected)
        messagebox.showinfo("Deleted", "Selected row removed from this plan only.")
        keep_window_active(self)

    def edit_cell(self, event):
        source_tree = self.tree
        region = source_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = source_tree.identify_column(event.x)
        item = source_tree.identify_row(event.y)
        if not item:
            return

        col_idx = int(column[1:]) - 1
        col_name = source_tree["columns"][col_idx]
        full_col_idx = col_idx

        if col_name not in ["Wtg. %", "Actuals Till Last FY"] + self.month_columns:
            return
        if self.is_user_saved_plan_locked() and col_name == "Wtg. %":
            messagebox.showwarning("Plan Locked", "This plan is already saved. User cannot edit activity weight.")
            keep_window_active(self)
            return
        if col_name in ["Actuals Till Last FY"] + self.month_columns and not self.current_plan_saved:
            messagebox.showwarning("Save Activities First", "Save activities first, then enter planning data.")
            keep_window_active(self)
            return
        current_values = list(self.tree.item(item)["values"])
        if col_name in self.month_columns:
            if self.plan_has_saved_planning() and not self.is_admin_user():
                messagebox.showwarning("Planning Locked", "Planning is already saved. User cannot edit saved planning.")
                keep_window_active(self)
                return
            allowed, reason = self.is_month_allowed_for_activity(col_name, current_values)
            if not allowed:
                messagebox.showwarning("Planning Not Allowed", reason)
                keep_window_active(self)
                return

        x, y, w, h = source_tree.bbox(item, column)
        current_value = self.tree.item(item)["values"][full_col_idx]

        entry = Entry(source_tree, width=w//10, justify="center")
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, "" if str(current_value).strip() in (self.editable_marker(), self.locked_marker()) else current_value)
        entry.focus()
        entry.select_range(0, END)

        def save_edit(e=None):
            try:
                new_val = float(entry.get() or 0)
            except:
                new_val = 0
            values = list(self.tree.item(item)["values"])
            if col_name == "Wtg. %" and not self.validate_total_weight(entry.get(), exclude_item=item):
                entry.destroy()
                return
            if col_name in self.month_columns:
                scope = float(values[3] or 0)
                actual_last_fy = float(values[7] or 0)
                planned_total = 0.0
                for idx in range(8, len(values)):
                    planned_total += new_val if idx == full_col_idx else self.planning_cell_qty(values[idx])
                if actual_last_fy + planned_total > scope:
                    entry.destroy()
                    messagebox.showerror("Scope Limit", "Actuals Till Last FY + monthly planning cannot exceed Scope Qty.")
                    keep_window_active(self)
                    return
                values[full_col_idx] = f"{new_val:.2f}" if new_val else self.editable_marker()
            else:
                values[full_col_idx] = f"{new_val:.2f}"
            values = self.recalculate_remaining_qty(values)
            self.tree.item(item, values=values)
            entry.destroy()

        entry.bind("<Return>", save_edit)
        entry.bind("<FocusOut>", save_edit)

    def refresh_current_plan(self):
        if self.current_plan:
            self.load_activities_for_plan()
            messagebox.showinfo("Refreshed", f"Plan '{self.current_plan}' refreshed successfully!")
            keep_window_active(self)
        else:
            messagebox.showwarning("Select Plan", "Please select a plan from the left list first.")
            keep_window_active(self)


# ====================== DAILY PROGRESS REPORT (UPDATED - Matches Screenshot Table) ======================
class DailyProgressWindow(Toplevel):
    def __init__(self, parent, project_id=None, uid=None, main_app=None):
        super().__init__(parent)
        self.project_id = project_id
        self.uid = uid
        self.main_app = main_app

        if self.main_app and project_id and not self.main_app.can_access_project(project_id):
            messagebox.showwarning("Access Denied", "You do not have access to this project.")
            self.destroy()
            return

        self.title(f"Daily Progress Report - {uid if uid else 'All Ongoing Projects'}")
        self.geometry("1850x950")
        self.configure(bg="#f0f4f8")

        self.current_project_id = project_id
        self.can_add_daily_progress = False
        self.current_project_name = ""

        self.build_ui()
        if not project_id:
            self.load_ongoing_projects()
        else:
            self.current_project_name = self.get_project_name(project_id)
            self.load_daily_progress_table()
        apply_page_watermark(self)
        normalize_buttons(self)

    def build_ui(self):
        # Header
        top = Frame(self, bg="#003087", height=65)
        top.pack(fill=X)
        top.pack_propagate(False)
        Label(top, text="📊 DAILY PROGRESS REPORT", bg="#003087", fg="white",
              font=("Arial", 20, "bold")).pack(expand=True, pady=15)

        # Main content
        self.page_scroll = ScrollableFrame(self, bg="#f0f4f8")
        self.page_scroll.pack(fill=BOTH, expand=True, padx=15, pady=10)
        main = Frame(self.page_scroll.scrollable_frame, bg="#f0f4f8")
        main.pack(anchor="nw", fill=BOTH, expand=True)

        # Project selector (left side when no project passed)
        if not self.project_id:
            self.left_frame = Frame(main, bg="#f0f4f8", width=210, relief="groove", bd=3)
            self.left_frame.pack(side=LEFT, fill=Y, padx=(0, 10))

            Label(self.left_frame, text="Ongoing Projects", font=("Arial", 14, "bold"),
                  bg="#f0f4f8", fg="#003087").pack(pady=10)

            list_container = Frame(self.left_frame, bg="#f0f4f8")
            list_container.pack(fill=BOTH, expand=True, padx=10, pady=5)

            self.project_listbox = Text(
                list_container,
                font=("Arial", 12),
                wrap=WORD,
                padx=6,
                pady=6,
                cursor="hand2",
                relief="sunken",
                bd=1,
                width=18,
                height=22,
            )
            self.project_listbox.pack(side=LEFT, fill=BOTH, expand=True)
            project_scroll = ttk.Scrollbar(list_container, orient="vertical", command=self.project_listbox.yview)
            project_scroll.pack(side=RIGHT, fill=Y)
            self.project_listbox.configure(yscrollcommand=project_scroll.set)
            self.project_listbox.configure(state="disabled")

        # Right side - Table
        self.right_frame = Frame(main, bg="#f0f4f8")
        self.right_frame.pack(side=LEFT, fill=BOTH, expand=True)

    def get_project_name(self, project_id):
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT project_name FROM projects WHERE id=%s", (project_id,))
        row = c.fetchone()
        conn.close()
        return row["project_name"] if row else ""

    def wrap_header_text(self, text, max_chars=14):
        words = str(text or "").split()
        if not words:
            return ""
        lines = []
        current = words[0]
        for word in words[1:]:
            if len(current) + 1 + len(word) <= max_chars:
                current += " " + word
            else:
                lines.append(current)
                current = word
        lines.append(current)
        return "\n".join(lines[:3])

    def get_short_activity_name(self, activity_type):
        text = str(activity_type or "").strip()
        if "->" in text:
            return text.split("->")[-1].strip()
        return text

    def build_group_header(self, parent, groups):
        header_height = 38
        canvas = Canvas(parent, bg="white", height=header_height, highlightthickness=1, highlightbackground="#c8d2dc")
        total_width = sum(width for _, width in groups)
        x = 0
        for label_text, width in groups:
            canvas.create_rectangle(x, 0, x + width, header_height, outline="#9aa7b4", fill="white")
            canvas.create_text(
                x + (width / 2),
                header_height / 2,
                text=label_text,
                fill="#111111",
                font=("Arial", 9, "normal"),
                width=max(20, width - 8),
            )
            x += width
        canvas.configure(scrollregion=(0, 0, total_width, header_height))
        return canvas

    def sync_manpower_scrollbar(self, first, last):
        if hasattr(self, "manpower_x_scroll"):
            self.manpower_x_scroll.set(first, last)
        if hasattr(self, "manpower_header_canvas"):
            try:
                self.manpower_header_canvas.xview_moveto(first)
            except Exception:
                pass

    def sync_manpower_xview(self, *args):
        if hasattr(self, "manpower_tree"):
            self.manpower_tree.xview(*args)
        if hasattr(self, "manpower_header_canvas"):
            self.manpower_header_canvas.xview(*args)

    def adjust_numeric_var(self, tk_var, delta):
        try:
            current = int(str(tk_var.get() or "0").strip())
        except Exception:
            current = 0
        tk_var.set(str(max(0, current + delta)))

    def get_allowed_daily_entry_window(self):
        today = datetime.now().date()
        return today - timedelta(days=2), today

    def is_date_allowed_for_daily_entry(self, date_value):
        earliest_allowed, latest_allowed = self.get_allowed_daily_entry_window()
        return earliest_allowed <= date_value <= latest_allowed

    def format_daily_entry_window(self):
        earliest_allowed, latest_allowed = self.get_allowed_daily_entry_window()
        return earliest_allowed.strftime("%d-%m-%y"), latest_allowed.strftime("%d-%m-%y")

    def select_ongoing_project(self, project_index):
        project = self.ongoing_projects[project_index]
        self.current_project_id = project['id']
        self.uid = project['unique_id']
        self.current_project_name = project['project_name']
        self.highlight_project_tag(project_index)

        for widget in self.right_frame.winfo_children():
            widget.destroy()
        self.load_daily_progress_table()

    def highlight_project_tag(self, selected_index):
        if not hasattr(self, "project_listbox"):
            return
        self.project_listbox.configure(state="normal")
        for idx in range(len(getattr(self, "ongoing_projects", []))):
            tag_name = f"project_{idx}"
            self.project_listbox.tag_configure(tag_name, background="white", foreground="#111111")
        self.project_listbox.tag_configure(f"project_{selected_index}", background="#0b79d0", foreground="white")
        self.project_listbox.configure(state="disabled")

    def on_project_text_click(self, event):
        index = self.project_listbox.index(f"@{event.x},{event.y}")
        for tag_name in self.project_listbox.tag_names(index):
            if tag_name.startswith("project_"):
                try:
                    project_index = int(tag_name.split("_")[1])
                except Exception:
                    return
                self.select_ongoing_project(project_index)
                return

    def load_ongoing_projects(self):
        self.project_listbox.configure(state="normal")
        self.project_listbox.delete("1.0", END)
        conn = get_db_connection()
        c = conn.cursor()
        allowed_project_ids = self.main_app.get_allowed_project_ids() if self.main_app else None
        if allowed_project_ids is not None and not allowed_project_ids:
            self.ongoing_projects = []
            conn.close()
            self.project_listbox.configure(state="disabled")
            return
        if allowed_project_ids is None:
            c.execute("""
                SELECT id, unique_id, project_name
                FROM projects
                WHERE stage2_cleared = 'Y'
                ORDER BY id DESC
            """)
        else:
            c.execute("""
                SELECT id, unique_id, project_name
                FROM projects
                WHERE stage2_cleared = 'Y' AND id = ANY(%s)
                ORDER BY id DESC
            """, (list(allowed_project_ids),))
        self.ongoing_projects = c.fetchall()
        conn.close()

        for idx, p in enumerate(self.ongoing_projects):
            tag_name = f"project_{idx}"
            display_text = p["project_name"] + ("\n\n" if idx != len(self.ongoing_projects) - 1 else "")
            self.project_listbox.insert(END, display_text, (tag_name,))
            self.project_listbox.tag_bind(tag_name, "<Button-1>", self.on_project_text_click)
            self.project_listbox.tag_bind(tag_name, "<Enter>", lambda e: self.project_listbox.config(cursor="hand2"))
        self.project_listbox.configure(state="disabled")
        if self.ongoing_projects:
            self.select_ongoing_project(0)
        else:
            self.load_daily_progress_table()

    def reload_daily_progress_view(self):
        for widget in self.right_frame.winfo_children():
            widget.destroy()
        self.load_daily_progress_table()

    def get_manpower_contractor_names(self):
        if not self.current_project_id:
            return []
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""
            SELECT DISTINCT contractor_name
            FROM daily_progress_manpower
            WHERE project_id = %s
              AND contractor_name IS NOT NULL
              AND contractor_name <> ''
              AND category_name <> 'Staff / Supervisory'
            ORDER BY contractor_name
        """, (self.current_project_id,))
        rows = c.fetchall()
        conn.close()
        contractor_names = ["Unistar", "Sharda Co"]
        for row in rows:
            name = row["contractor_name"]
            if name and name not in contractor_names:
                contractor_names.append(name)
        return contractor_names

    def load_daily_progress_table(self):
        self.can_add_daily_progress = project_has_completed_planning(self.current_project_id)
        self.current_daily_plan = get_latest_planned_plan(self.current_project_id) if self.current_project_id else None

        self.selected_daily_record = None
        self.daily_row_widgets = {}
        self.manpower_tree = None
        self.progress_tree = None

        project_title = self.current_project_name or self.uid or ""
        Label(self.right_frame, text=f"Project: {project_title}", font=("Arial", 16, "bold"),
              bg="#f0f4f8", fg="#003087").pack(pady=(5, 10))
        if self.current_daily_plan:
            Label(self.right_frame, text=f"S-Curve Plan: {self.current_daily_plan}", font=("Arial", 11, "bold"),
                  bg="#f0f4f8", fg="#555555").pack(pady=(0, 8))

        if not self.can_add_daily_progress:
            Label(self.right_frame,
                  text="Plan S-curve data first. Daily progress entry is enabled only after S-curve planning quantity is completed and saved.",
                  font=("Arial", 11, "bold"), bg="#f0f4f8", fg="#c8102e").pack(pady=(0, 8))

        self.table_card = Frame(self.right_frame, bg="white", bd=0, relief="flat", padx=16, pady=14)
        self.table_card.pack(fill=X, padx=4, pady=(4, 18))
        Label(
            self.table_card,
            text="Data Entry Table: Editable Window",
            bg="white",
            fg="#1f2937",
            font=("Arial", 10, "bold"),
        ).pack(anchor=W)
        Label(
            self.table_card,
            text="Note: Date field is read-only. Click 'Edit' to modify values.",
            bg="white",
            fg="#4b5563",
            font=("Arial", 8),
        ).pack(anchor=W, pady=(3, 10))

        self.table_canvas = Canvas(self.table_card, bg="white", height=210, highlightthickness=1, highlightbackground="#d1d9e0")
        self.table_canvas.pack(fill=X, expand=False)
        self.table_grid = Frame(self.table_canvas, bg="white")
        self.table_window = self.table_canvas.create_window((0, 0), window=self.table_grid, anchor="nw")
        self.table_x_scroll = ttk.Scrollbar(self.table_card, orient="horizontal", command=self.table_canvas.xview)
        self.table_x_scroll.pack(fill=X)
        self.table_canvas.configure(xscrollcommand=self.table_x_scroll.set)
        self.table_grid.bind("<Configure>", lambda e: self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all")))

        self.summary_frame = Frame(self.right_frame, bg="#f0f4f8")
        self.summary_frame.pack(fill=BOTH, expand=True, padx=4, pady=(2, 8))

        # Buttons
        btn_frame = Frame(self.right_frame, bg="#f0f4f8")
        btn_frame.pack(side=BOTTOM, fill=X, pady=(6, 8))
        btn_inner = Frame(btn_frame, bg="#f0f4f8")
        btn_inner.pack(anchor="center")
        self.add_day_btn = Button(btn_inner, text="Save / Add Day", command=self.add_new_day,
                                  bg="#008000", fg="white", font=("Arial", 11, "bold"), width=18)
        self.add_day_btn.pack(side=LEFT, padx=6)
        self.delete_day_btn = Button(btn_inner, text="Delete", command=self.delete_selected_daily_progress,
                                     bg="#c8102e", fg="white", font=("Arial", 11, "bold"), width=14)
        self.delete_day_btn.pack(side=LEFT, padx=6)
        Button(btn_inner, text="Refresh", command=self.reload_daily_progress_view,
               bg="#0066cc", fg="white", font=("Arial", 11, "bold"), width=14).pack(side=LEFT, padx=6)
        Button(btn_inner, text="Home", command=self.go_home,
               bg="#003087", fg="white", font=("Arial", 11, "bold"), width=14).pack(side=LEFT, padx=6)

        if not self.can_add_daily_progress or (self.main_app and not self.main_app.can_edit("daily_progress")):
            self.add_day_btn.config(state=DISABLED)
            self.delete_day_btn.config(state=DISABLED)

        normalize_buttons(btn_inner)
        self.refresh_table()

    def sync_daily_progress_selection(self, source):
        return

    def get_selected_daily_progress_item(self):
        if getattr(self, "selected_daily_record", None):
            return None, self.selected_daily_record
        for tree in (self.manpower_tree, self.progress_tree):
            if tree and tree.selection():
                item_id = tree.selection()[0]
                return tree, item_id
        return None, None

    def get_item_report_date(self, tree, item):
        if isinstance(item, dict):
            return item.get("report_date_display", "")
        values = tree.item(item).get("values", [])
        columns = list(tree["columns"])
        if "Date" in columns:
            date_index = columns.index("Date")
            if len(values) > date_index:
                return values[date_index]
        return values[0] if values else ""

    def select_daily_row(self, row_info):
        self.selected_daily_record = row_info
        selected_id = str(row_info.get("id"))
        for record_id, widgets in getattr(self, "daily_row_widgets", {}).items():
            bg = "#dbeafe" if str(record_id) == selected_id else "#f8fafc"
            for widget in widgets:
                try:
                    widget.configure(bg=bg)
                except Exception:
                    pass

    def edit_daily_record(self, row_info):
        self.select_daily_row(row_info)
        self.edit_selected_row()

    def make_table_cell(self, parent, text, row, col, width, bg="#f8fafc", fg="#111827", font=None, rowspan=1, columnspan=1, command=None):
        cell = Label(
            parent,
            text=text,
            width=width,
            bg=bg,
            fg=fg,
            font=font or ("Arial", 10),
            anchor=CENTER,
            justify=CENTER,
            relief="solid",
            bd=1,
            padx=5,
            pady=9,
        )
        cell.grid(row=row, column=col, rowspan=rowspan, columnspan=columnspan, sticky="nsew")
        if command:
            cell.bind("<Button-1>", lambda e: command())
            cell.configure(cursor="hand2")
        return cell

    def build_daily_data_grid(self, records):
        for widget in self.table_grid.winfo_children():
            widget.destroy()
        self.daily_row_widgets = {}

        columns = [
            ("date", "Date", 13, "#1f4e79", "white"),
            ("rsp_executive", "Rsp Executive", 15, "#164e7a", "white"),
            ("rsp_non_executive", "Rsp Non-Executive", 19, "#0e93c8", "white"),
            ("executing_agency", "Executing Agency", 18, "#568f98", "white"),
            ("labour_deployed", "Labour Deployed", 17, "#77b487", "white"),
            ("supervisor", "Supervisor", 12, "#82949d", "white"),
            ("design_engineering", "Design Engineering", 19, "#e2b24d", "#111827"),
            ("civil", "Civil", 12, "#ef6849", "white"),
            ("structural_supply", "Structural Supply", 18, "#d62141", "white"),
            ("structural_erection", "Structural Erection", 19, "#ae3b30", "white"),
            ("equipment_supply", "Equipment Supply", 18, "#b6d600", "white"),
            ("equipment_erection", "Equipment Erection", 19, "#167055", "white"),
            ("actions", "Actions", 11, "#137158", "white"),
        ]
        group_specs = [
            ("Date", 0, 1, "#1f4e79"),
            ("MANPOWER RESOURCES", 1, 5, "#2548b8"),
            ("CONSTRUCTION PROGRESS", 6, 6, "#c92918"),
            ("Actions", 12, 1, "#137158"),
        ]
        for label, start_col, span, color in group_specs:
            self.make_table_cell(
                self.table_grid,
                label,
                0,
                start_col,
                sum(columns[idx][2] for idx in range(start_col, start_col + span)),
                bg=color,
                fg="white",
                font=("Arial", 12, "bold"),
                rowspan=2 if span == 1 else 1,
                columnspan=span,
            )
        for col_idx, (_, label, width, bg, fg) in enumerate(columns):
            if col_idx in (0, 12):
                continue
            self.make_table_cell(self.table_grid, label, 1, col_idx, width, bg=bg, fg=fg, font=("Arial", 9, "bold"))

        if not records:
            self.make_table_cell(
                self.table_grid,
                "No daily progress entries available.",
                2,
                0,
                50,
                bg="#f8fafc",
                fg="#64748b",
                font=("Arial", 11, "bold"),
                columnspan=len(columns),
            )
            return

        for row_idx, record in enumerate(records, start=2):
            row_info = {
                "id": record["id"],
                "report_date_display": to_display_date(record["report_date"]),
            }
            row_widgets = []
            for col_idx, (key, _, width, _, _) in enumerate(columns):
                if key == "date":
                    value = row_info["report_date_display"]
                elif key == "actions":
                    btn = Button(
                        self.table_grid,
                        text="Edit",
                        command=lambda info=row_info: self.edit_daily_record(info),
                        bg="#2563eb",
                        fg="white",
                        font=("Arial", 9, "bold"),
                        relief="flat",
                        padx=8,
                        pady=2,
                        cursor="hand2",
                    )
                    btn.grid(row=row_idx, column=col_idx, sticky="nsew", padx=1, pady=1)
                    row_widgets.append(btn)
                    continue
                else:
                    value = record.get(key, 0) or 0
                    if isinstance(value, float) and not value.is_integer():
                        value = f"{value:.2f}"
                    else:
                        value = str(int(value)) if isinstance(value, (int, float)) else str(value)
                cell_bg = "#e9f1f7" if col_idx <= 5 else "#fbf3e7"
                if col_idx in (2, 3):
                    cell_bg = "#e4f4fb"
                elif col_idx in (4, 5):
                    cell_bg = "#e8f5ee"
                elif col_idx in (8, 9):
                    cell_bg = "#fae8ec"
                elif col_idx in (10, 11):
                    cell_bg = "#eff7e4"
                cell = self.make_table_cell(
                    self.table_grid,
                    value,
                    row_idx,
                    col_idx,
                    width,
                    bg=cell_bg,
                    font=("Arial", 10, "bold") if col_idx == 0 else ("Arial", 10),
                    command=lambda info=row_info: self.select_daily_row(info),
                )
                row_widgets.append(cell)
            self.daily_row_widgets[str(record["id"])] = row_widgets

    def refresh_table(self):
        if not hasattr(self, "table_grid"):
            return

        records = get_daily_progress_display_rows(self.current_project_id, self.current_daily_plan)
        self.build_daily_data_grid(records)
        self.render_progress_summary()

    def progress_category_for_activity(self, activity_type):
        text = str(activity_type or "").strip().lower()
        if "design" in text and "engineering" in text:
            return "Design & Engineering"
        if "civil" in text:
            return "Civil-RCC"
        if "supply" in text and ("steel" in text or "structur" in text):
            return "Structural Supply"
        if "erection" in text and ("steel" in text or "structur" in text):
            return "Structural Erection"
        if "supply" in text and ("electrical" in text or "equipment" in text or "mechanical" in text):
            return "Equipment Supply"
        if "erection" in text and ("electrical" in text or "equipment" in text or "mechanical" in text):
            return "Equipment Erection"
        return "Other Progress"

    def planned_qty_for_period(self, activity, period_start=None, period_end=None):
        scope_qty = float(activity.get("scope_qty") or 0)
        start_date = parse_app_date(activity.get("start_date"))
        finish_date = parse_app_date(activity.get("finish_date"))
        if not start_date or not finish_date or scope_qty <= 0:
            return 0.0
        if finish_date < start_date:
            finish_date = start_date
        if period_start is None:
            period_start = start_date
        if period_end is None:
            period_end = datetime.now().date()
        actual_start = max(start_date, period_start)
        actual_end = min(finish_date, period_end)
        if actual_end < actual_start:
            return 0.0
        total_days = (finish_date - start_date).days + 1
        period_days = (actual_end - actual_start).days + 1
        return scope_qty * (period_days / max(1, total_days))

    def summary_periods(self, selected_date=None):
        selected_date = selected_date or datetime.now().date()
        fy_start_year = selected_date.year if selected_date.month >= 4 else selected_date.year - 1
        fy_start = datetime(fy_start_year, 4, 1).date()
        last_fy_end = fy_start - timedelta(days=1)
        month_start = selected_date.replace(day=1)
        if selected_date.month == 12:
            month_end = datetime(selected_date.year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(selected_date.year, selected_date.month + 1, 1).date() - timedelta(days=1)
        fiscal_months = []
        for month in range(4, 13):
            fiscal_months.append(datetime(fy_start_year, month, 1).strftime("%b-%y"))
        for month in range(1, 4):
            fiscal_months.append(datetime(fy_start_year + 1, month, 1).strftime("%b-%y"))
        selected_month_label = selected_date.strftime("%b-%y")
        selected_index = fiscal_months.index(selected_month_label) if selected_month_label in fiscal_months else -1
        fy_months = fiscal_months[:selected_index + 1] if selected_index >= 0 else [selected_month_label]
        return {
            "selected_date": selected_date,
            "fy_label": f"FY {fy_start_year}-{fy_start_year + 1}",
            "fy_start": fy_start,
            "last_fy_end": last_fy_end,
            "month_start": month_start,
            "month_end": month_end,
            "month_label": selected_month_label,
            "fy_months": fy_months,
        }

    def percent_of_scope(self, value, scope):
        scope = float(scope or 0)
        return (float(value or 0) / scope * 100) if scope else 0.0

    def weighted_summary_percent(self, rows, value_key):
        weighted_total = 0.0
        has_weight = False
        for row in rows or []:
            raw_weight = float(row.get("weight_percent") or 0)
            weight_fraction = raw_weight / 100 if raw_weight > 1 else raw_weight
            if weight_fraction <= 0:
                continue
            has_weight = True
            scope = float(row.get("scope") or 0)
            qty = float(row.get(value_key) or 0)
            if scope:
                weighted_total += weight_fraction * (qty / scope)
        if has_weight:
            return weighted_total * 100
        total_scope = sum(float(row.get("scope") or 0) for row in (rows or []))
        total_value = sum(float(row.get(value_key) or 0) for row in (rows or []))
        return self.percent_of_scope(total_value, total_scope)

    def get_summary_actual_totals_by_activity(self, activity_ids, selected_date=None):
        if not activity_ids:
            return {}, {}, {}
        periods = self.summary_periods(selected_date)
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""
            SELECT activity_id,
                   COALESCE(SUM(CASE WHEN actual_date::date <= %s THEN actual_qty ELSE 0 END), 0) AS actual_till_last_fy,
                   COALESCE(SUM(CASE WHEN actual_date::date >= %s AND actual_date::date <= %s THEN actual_qty ELSE 0 END), 0) AS fy_actual,
                   COALESCE(SUM(CASE WHEN actual_date::date >= %s AND actual_date::date <= %s THEN actual_qty ELSE 0 END), 0) AS month_actual
            FROM daily_actuals
            WHERE activity_id = ANY(%s)
            GROUP BY activity_id
        """, (
            periods["last_fy_end"],
            periods["fy_start"],
            periods["month_end"],
            periods["month_start"],
            periods["month_end"],
            list(activity_ids),
        ))
        actual_till_last_fy = {}
        fy_actual = {}
        month_actual = {}
        for row in c.fetchall():
            activity_id = int(row["activity_id"])
            actual_till_last_fy[activity_id] = float(row["actual_till_last_fy"] or 0)
            fy_actual[activity_id] = float(row["fy_actual"] or 0)
            month_actual[activity_id] = float(row["month_actual"] or 0)
        conn.close()
        return actual_till_last_fy, fy_actual, month_actual

    def get_summary_plan_totals_by_activity_type(self, activity_types, selected_date=None):
        periods = self.summary_periods(selected_date)
        cleaned_types = [str(activity_type or "").strip() for activity_type in activity_types if str(activity_type or "").strip()]
        if not cleaned_types:
            return {}
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""
            SELECT activity_type,
                   COALESCE(SUM(CASE WHEN month = %s THEN planned_qty ELSE 0 END), 0) AS month_plan,
                   COALESCE(SUM(CASE WHEN month = ANY(%s) THEN planned_qty ELSE 0 END), 0) AS fy_plan
            FROM monthly_plans
            WHERE project_id=%s
              AND plan_name=%s
              AND activity_type = ANY(%s)
            GROUP BY activity_type
        """, (
            periods["month_label"],
            periods["fy_months"],
            self.current_project_id,
            self.current_daily_plan,
            cleaned_types,
        ))
        plans = {}
        for row in c.fetchall():
            plans[str(row["activity_type"] or "")] = {
                "month_plan": float(row["month_plan"] or 0),
                "fy_plan": float(row["fy_plan"] or 0),
            }
        conn.close()
        return plans

    def get_scope_summary_rows(self):
        if not self.current_project_id or not self.current_daily_plan:
            return [], 0.0, 0.0, {}
        activities = get_activities_for_plan(self.current_project_id, self.current_daily_plan)
        activity_ids = [int(row["id"]) for row in activities]
        actual_till_last_fy_map, fy_actual_map, month_actual_map = self.get_summary_actual_totals_by_activity(activity_ids)
        plan_totals_by_activity = self.get_summary_plan_totals_by_activity_type([row.get("activity_type") for row in activities])
        periods = self.summary_periods()
        rows = []
        for activity in activities:
            activity_id = int(activity["id"])
            activity_type = str(activity.get("activity_type") or "")
            scope_qty = float(activity.get("scope_qty") or 0)
            actual_till_last_fy = actual_till_last_fy_map.get(activity_id, 0.0)
            month_plan = plan_totals_by_activity.get(activity_type, {}).get("month_plan", 0.0)
            month_actual = month_actual_map.get(activity_id, 0.0)
            fy_plan = plan_totals_by_activity.get(activity_type, {}).get("fy_plan", 0.0)
            fy_actual = fy_actual_map.get(activity_id, 0.0)
            cumulative_plan = actual_till_last_fy + fy_plan
            cumulative_actual = actual_till_last_fy + fy_actual
            rows.append({
                "activity": activity_type,
                "scope": scope_qty,
                "uom": activity.get("uom") or "",
                "weight_percent": float(activity.get("weight_percent") or 0),
                "actual_till_last_fy": actual_till_last_fy,
                "month_plan": month_plan,
                "month_actual": month_actual,
                "fy_plan": fy_plan,
                "fy_actual": fy_actual,
                "cumulative_plan": cumulative_plan,
                "cumulative_actual": cumulative_actual,
            })
        planned_percent = self.weighted_summary_percent(rows, "cumulative_plan")
        actual_percent = self.weighted_summary_percent(rows, "cumulative_actual")
        return rows, planned_percent, actual_percent, periods

    def format_summary_number(self, value):
        value = float(value or 0)
        return f"{int(value):,}" if value.is_integer() else f"{value:,.2f}"

    def render_progress_summary(self):
        for widget in self.summary_frame.winfo_children():
            widget.destroy()
        scope_rows, planned_percent, actual_percent, periods = self.get_scope_summary_rows()

        left_card = Frame(self.summary_frame, bg="white", padx=16, pady=14)
        left_card.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 12))
        right_card = Frame(self.summary_frame, bg="white", padx=16, pady=14)
        right_card.pack(side=LEFT, fill=BOTH, expand=True, padx=(12, 0))

        month_label = periods.get("month_label", datetime.now().strftime("%b-%y")) if periods else datetime.now().strftime("%b-%y")
        fy_label = periods.get("fy_label", "Financial Year") if periods else "Financial Year"
        Label(left_card, text="PROJECT PROGRESS SUMMARY (QUANTITY & %)", bg="white", fg="#1e40af", font=("Arial", 13, "bold")).pack(anchor=W, pady=(0, 12))
        scope_grid = Frame(left_card, bg="white")
        scope_grid.pack(fill=X)
        headers = [
            "Activity / Work Package",
            "Scope",
            "UOM",
            "Actual Till\nLast FY",
            f"Month Plan\n{month_label}",
            f"Month Actual\n{month_label}",
            f"FY Plan\n{fy_label}",
            f"FY Actual\n{fy_label}",
            "Cumulative\nPlan",
            "Cumulative\nActual",
        ]
        widths = [28, 10, 8, 13, 11, 12, 13, 13, 13, 13]
        for col, header in enumerate(headers):
            Label(scope_grid, text=header, width=widths[col], bg="#334155", fg="#fef08a" if col in (0, 3, 5) else "white",
                  font=("Arial", 9, "bold"), relief="solid", bd=1, pady=7).grid(row=0, column=col, sticky="nsew")
        if not scope_rows:
            Label(scope_grid, text="No S-Curve activity scope available.", bg="#f8fafc", fg="#64748b", font=("Arial", 10, "bold"),
                  relief="solid", bd=1, pady=10).grid(row=1, column=0, columnspan=len(headers), sticky="nsew")
        if scope_rows:
            overall_values = [
                "Overall Progress",
                "100%",
                "%",
                f"{self.weighted_summary_percent(scope_rows, 'actual_till_last_fy'):.2f}%",
                f"{self.weighted_summary_percent(scope_rows, 'month_plan'):.2f}%",
                f"{self.weighted_summary_percent(scope_rows, 'month_actual'):.2f}%",
                f"{self.weighted_summary_percent(scope_rows, 'fy_plan'):.2f}%",
                f"{self.weighted_summary_percent(scope_rows, 'fy_actual'):.2f}%",
                f"{self.weighted_summary_percent(scope_rows, 'cumulative_plan'):.2f}%",
                f"{self.weighted_summary_percent(scope_rows, 'cumulative_actual'):.2f}%",
            ]
            for col, value in enumerate(overall_values):
                Label(scope_grid, text=value, width=widths[col], bg="#fde6b7", fg="#111827",
                      font=("Arial", 9, "bold"), relief="solid", bd=1, pady=6).grid(row=1, column=col, sticky="nsew")
        for index, row in enumerate(scope_rows):
            row_idx = 2 + (index * 2)
            values = [
                row["activity"],
                self.format_summary_number(row["scope"]),
                row["uom"],
                self.format_summary_number(row["actual_till_last_fy"]),
                self.format_summary_number(row["month_plan"]),
                self.format_summary_number(row["month_actual"]),
                self.format_summary_number(row["fy_plan"]),
                self.format_summary_number(row["fy_actual"]),
                self.format_summary_number(row["cumulative_plan"]),
                self.format_summary_number(row["cumulative_actual"]),
            ]
            for col, value in enumerate(values):
                bg = "#f8fafc" if col < 3 else "#edf2ff"
                Label(scope_grid, text=value, width=widths[col], bg=bg, fg="#1e3a8a" if col in (0, 1, 2) else "#111827",
                      font=("Arial", 9, "bold" if col in (0, 1, 2) else "normal"), relief="solid", bd=1, pady=6).grid(row=row_idx, column=col, sticky="nsew")
            percent_values = [
                "",
                "",
                "%",
                f"{self.percent_of_scope(row['actual_till_last_fy'], row['scope']):.2f}%",
                f"{self.percent_of_scope(row['month_plan'], row['scope']):.2f}%",
                f"{self.percent_of_scope(row['month_actual'], row['scope']):.2f}%",
                f"{self.percent_of_scope(row['fy_plan'], row['scope']):.2f}%",
                f"{self.percent_of_scope(row['fy_actual'], row['scope']):.2f}%",
                f"{self.percent_of_scope(row['cumulative_plan'], row['scope']):.2f}%",
                f"{self.percent_of_scope(row['cumulative_actual'], row['scope']):.2f}%",
            ]
            for col, value in enumerate(percent_values):
                bg = "#fff7e6" if col < 3 else "#eef6ff"
                Label(scope_grid, text=value, width=widths[col], bg=bg, fg="#111827",
                      font=("Arial", 8), relief="solid", bd=1, pady=5).grid(row=row_idx + 1, column=col, sticky="nsew")

        Label(right_card, text="CUMULATIVE PERFORMANCE SNAPSHOT", bg="white", fg="#1e40af", font=("Arial", 13, "bold")).pack(anchor=W, pady=(0, 14))
        metric_row = Frame(right_card, bg="white")
        metric_row.pack(fill=X)
        self.render_metric_box(metric_row, "PLANNED PROGRESS", planned_percent, "#60a5fa", "#1e40af").pack(side=LEFT, fill=X, expand=True, padx=(0, 8))
        self.render_metric_box(metric_row, "ACTUAL PROGRESS", actual_percent, "#4ade80", "#166534").pack(side=LEFT, fill=X, expand=True, padx=(8, 0))

        rate_box = Frame(right_card, bg="#f1f5f9", padx=10, pady=10)
        rate_box.pack(fill=X, pady=(16, 0))
        Label(rate_box, text="EXECUTION PROGRESS RATE", bg="#f1f5f9", fg="#111827", font=("Arial", 9, "bold")).pack(anchor=W)
        bar = Canvas(rate_box, bg="#f1f5f9", height=28, highlightthickness=0)
        bar.pack(fill=X, pady=(4, 0))
        bar.bind("<Configure>", lambda e, pct=actual_percent: self.draw_progress_bar(bar, pct))

        self.render_daily_report_matrix(self.summary_frame)

    def daily_report_cell(self, parent, text, row, col, width=9, bg="white", fg="#0f172a", font=None, columnspan=1, rowspan=1, anchor=CENTER):
        label = Label(
            parent,
            text=text,
            width=width,
            bg=bg,
            fg=fg,
            font=font or ("Arial", 8),
            anchor=anchor,
            justify=CENTER,
            relief="solid",
            bd=1,
            padx=3,
            pady=6,
        )
        label.grid(row=row, column=col, columnspan=columnspan, rowspan=rowspan, sticky="nsew")
        return label

    def format_daily_report_qty(self, value):
        try:
            value = float(value or 0)
        except Exception:
            value = 0
        return str(int(value)) if float(value).is_integer() else f"{value:.2f}"

    def render_daily_report_matrix(self, parent):
        if not self.current_project_id or not self.current_daily_plan:
            return
        today = datetime.now().date()
        try:
            matrix = get_daily_report_month_matrix(self.current_project_id, self.current_daily_plan, today.year, today.month)
        except Exception as exc:
            Label(
                parent,
                text=f"Daily Report could not be loaded: {exc}",
                bg="#fff7ed",
                fg="#c2410c",
                font=("Arial", 10, "bold"),
                padx=10,
                pady=8,
            ).pack(fill=X, pady=(14, 0))
            return

        card = Frame(parent, bg="white", padx=14, pady=12)
        card.pack(fill=BOTH, expand=True, pady=(14, 0))

        report_head = Frame(card, bg="white")
        report_head.pack(fill=X, pady=(0, 10))
        Label(
            report_head,
            text="Daily Report",
            bg="white",
            fg="#0f4aa2",
            font=("Arial", 14, "bold"),
        ).pack(side=LEFT)
        Label(
            report_head,
            text=f"Month: {matrix['month']}",
            bg="#eef6ff",
            fg="#0f376f",
            font=("Arial", 10, "bold"),
            padx=14,
            pady=6,
        ).pack(side=LEFT, padx=14)
        project = matrix.get("project") or {}
        Label(
            report_head,
            text=f"Project: {project.get('project_name') or self.current_project_name or '-'}",
            bg="#f8fafc",
            fg="#0f172a",
            font=("Arial", 9, "bold"),
            padx=12,
            pady=6,
        ).pack(side=LEFT, padx=4)
        Label(
            report_head,
            text=f"Project Code: {project.get('unique_id') or self.uid or '-'}",
            bg="#f8fafc",
            fg="#0f172a",
            font=("Arial", 9, "bold"),
            padx=12,
            pady=6,
        ).pack(side=LEFT, padx=4)

        table_area = Frame(card, bg="white")
        table_area.pack(fill=BOTH, expand=True)
        canvas = Canvas(table_area, bg="white", height=330, highlightthickness=1, highlightbackground="#dbeafe")
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        y_scroll = ttk.Scrollbar(table_area, orient="vertical", command=canvas.yview)
        y_scroll.pack(side=RIGHT, fill=Y)
        x_scroll = ttk.Scrollbar(card, orient="horizontal", command=canvas.xview)
        x_scroll.pack(fill=X)
        grid = Frame(canvas, bg="white")
        window_id = canvas.create_window((0, 0), window=grid, anchor="nw")
        canvas.configure(xscrollcommand=x_scroll.set, yscrollcommand=y_scroll.set)
        grid.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(window_id, height=max(e.height, grid.winfo_reqheight())))

        days = matrix.get("days") or []
        self.daily_report_cell(grid, "Activity / Work Description", 0, 0, width=30, bg="#eaf4ff", fg="#102a56", font=("Arial", 9, "bold"), rowspan=2)
        self.daily_report_cell(grid, "Category", 0, 1, width=14, bg="#eaf4ff", fg="#102a56", font=("Arial", 9, "bold"), rowspan=2)
        self.daily_report_cell(grid, matrix["month"], 0, 2, width=9 * max(1, len(days)), bg="#dbeafe", fg="#0f4aa2", font=("Arial", 10, "bold"), columnspan=max(1, len(days)))
        self.daily_report_cell(grid, "Total\n(Month)", 0, 2 + len(days), width=11, bg="#eaf4ff", fg="#102a56", font=("Arial", 9, "bold"), rowspan=2)

        for index, day_key in enumerate(days, start=2):
            day_date = datetime.strptime(day_key, "%Y-%m-%d").date()
            day_text = f"{day_date.day:02d}\n{day_date.strftime('%a')}"
            fg = "#dc2626" if day_date.weekday() == 6 else "#0f172a"
            bg = "#f8fbff" if day_date.weekday() != 6 else "#fff1f2"
            self.daily_report_cell(grid, day_text, 1, index, width=7, bg=bg, fg=fg, font=("Arial", 8, "bold"))

        row_no = 2
        activity_rows = matrix.get("activity_rows") or []
        if not activity_rows:
            self.daily_report_cell(grid, "No activity progress data available.", row_no, 0, width=44, bg="#f8fafc", fg="#64748b", font=("Arial", 9, "bold"), columnspan=3 + len(days))
            row_no += 1
        for activity in activity_rows:
            self.daily_report_cell(grid, f"{activity.get('serial')}. {activity.get('activity') or '-'}", row_no, 0, width=30, bg="white", fg="#0f172a", font=("Arial", 8, "bold"), anchor=W)
            self.daily_report_cell(grid, activity.get("category") or "-", row_no, 1, width=14, bg="#f8fafc", fg="#334155", font=("Arial", 8))
            values = activity.get("values") or {}
            for index, day_key in enumerate(days, start=2):
                day_date = datetime.strptime(day_key, "%Y-%m-%d").date()
                bg = "#f8fbff" if day_date.weekday() != 6 else "#fff7f7"
                self.daily_report_cell(grid, self.format_daily_report_qty(values.get(day_key, 0)), row_no, index, width=7, bg=bg)
            self.daily_report_cell(grid, self.format_daily_report_qty(activity.get("total", 0)), row_no, 2 + len(days), width=11, bg="#f8fafc", fg="#0f172a", font=("Arial", 8, "bold"))
            row_no += 1

        self.daily_report_cell(grid, "Manpower Category (No. of Persons)", row_no, 0, width=44, bg="#dbeafe", fg="#0f4aa2", font=("Arial", 9, "bold"), columnspan=2, anchor=W)
        for index in range(2, 3 + len(days)):
            self.daily_report_cell(grid, "", row_no, index, width=7 if index < 2 + len(days) else 11, bg="#dbeafe")
        row_no += 1

        for manpower in matrix.get("manpower_rows") or []:
            self.daily_report_cell(grid, manpower.get("label") or "-", row_no, 0, width=30, bg="white", fg="#0f172a", font=("Arial", 8, "bold"), columnspan=2, anchor=W)
            values = manpower.get("values") or {}
            for index, day_key in enumerate(days, start=2):
                day_date = datetime.strptime(day_key, "%Y-%m-%d").date()
                bg = "#f8fbff" if day_date.weekday() != 6 else "#fff7f7"
                self.daily_report_cell(grid, self.format_daily_report_qty(values.get(day_key, 0)), row_no, index, width=7, bg=bg)
            self.daily_report_cell(grid, self.format_daily_report_qty(manpower.get("total", 0)), row_no, 2 + len(days), width=11, bg="#f8fafc", fg="#0f172a", font=("Arial", 8, "bold"))
            row_no += 1

        self.daily_report_cell(grid, "TOTAL (Persons / Day)", row_no, 0, width=44, bg="#dbeafe", fg="#0f4aa2", font=("Arial", 9, "bold"), columnspan=2, anchor=W)
        day_totals = matrix.get("day_totals") or {}
        for index, day_key in enumerate(days, start=2):
            self.daily_report_cell(grid, self.format_daily_report_qty(day_totals.get(day_key, 0)), row_no, index, width=7, bg="#eaf4ff", fg="#0f376f", font=("Arial", 8, "bold"))
        self.daily_report_cell(grid, self.format_daily_report_qty(matrix.get("month_total", 0)), row_no, 2 + len(days), width=11, bg="#dbeafe", fg="#0f4aa2", font=("Arial", 8, "bold"))

    def render_metric_box(self, parent, label, percent, border_color, text_color):
        box = Frame(parent, bg="#eff6ff" if text_color == "#1e40af" else "#ecfdf5", highlightbackground=border_color, highlightthickness=2, padx=18, pady=14)
        Label(box, text=label, bg=box["bg"], fg=text_color, font=("Arial", 9, "bold")).pack()
        Label(box, text=f"{percent:.2f}%", bg=box["bg"], fg=text_color, font=("Arial", 24, "bold")).pack(pady=(6, 0))
        Label(box, text="Scheduled % to Date" if "40af" in text_color else "Actual Execution %", bg=box["bg"], fg="#475569", font=("Arial", 8)).pack()
        return box

    def draw_progress_bar(self, canvas, percent):
        canvas.delete("all")
        width = max(1, canvas.winfo_width())
        height = 18
        y = 5
        fill_width = int(width * min(max(percent, 0), 100) / 100)
        canvas.create_rectangle(0, y, width, y + height, fill="#fb7185", outline="")
        canvas.create_rectangle(0, y, fill_width, y + height, fill="#22c55e", outline="")
        canvas.create_text(max(45, fill_width / 2), y + height / 2, text=f"{percent:.1f}% ACTUAL", fill="white", font=("Arial", 8, "bold"))

    def edit_selected_row(self, event=None):
        if not self.can_add_daily_progress:
            messagebox.showwarning("Planning Incomplete", "Plan S-curve data first. Daily progress is enabled only after quantity planning is completed and saved.")
            return
        if self.main_app and not self.main_app.can_edit("daily_progress"):
            messagebox.showwarning("Edit Denied", "You have view access only for Daily Progress Report.")
            return

        tree, selected_item = self.get_selected_daily_progress_item()
        if not selected_item:
            return
        if isinstance(selected_item, dict):
            record_id = selected_item.get("id")
        else:
            item = tree.item(selected_item)
            record_id = item['tags'][0]
        report_date = self.get_item_report_date(tree, selected_item)
        try:
            selected_date = datetime.strptime(report_date, "%d-%m-%y").date()
        except Exception:
            selected_date = None
        if selected_date and not self.is_date_allowed_for_daily_entry(selected_date):
            start_text, end_text = self.format_daily_entry_window()
            messagebox.showwarning(
                "Date Restricted",
                f"Editing is allowed only for entries between {start_text} and {end_text}.",
            )
            keep_window_active(self)
            return
        self.open_edit_popup(record_id, [report_date])

    def delete_selected_daily_progress(self):
        if not self.can_add_daily_progress:
            messagebox.showwarning("Planning Incomplete", "Plan S-curve data first. Daily progress is enabled only after quantity planning is completed and saved.")
            return
        if self.main_app and not self.main_app.can_edit("daily_progress"):
            messagebox.showwarning("Edit Denied", "You have view access only for Daily Progress Report.")
            return

        tree, selected_item = self.get_selected_daily_progress_item()
        if not selected_item:
            messagebox.showwarning("Select", "Please select a daily progress row to delete.")
            return

        if isinstance(selected_item, dict):
            record_id = selected_item.get("id")
        else:
            item = tree.item(selected_item)
            record_id = item["tags"][0]
        report_date = self.get_item_report_date(tree, selected_item)
        if not messagebox.askyesno("Confirm Delete", f"Delete daily progress entry for {report_date}?"):
            return

        from database import delete_daily_progress
        delete_daily_progress(record_id)
        messagebox.showinfo("Deleted", "Daily progress entry deleted successfully.")
        self.reload_daily_progress_view()
        keep_window_active(self)

    def open_edit_popup(self, record_id, current_values):
        report_date = ""
        for value in current_values:
            if to_storage_date(value):
                report_date = value
                break
        self.open_daily_entry_popup(report_date or current_values[0], current_values)

    def open_daily_entry_popup(self, report_date_display, existing_values=None):
        self.current_daily_plan = get_latest_planned_plan(self.current_project_id) if self.current_project_id else None
        if not self.current_daily_plan:
            messagebox.showwarning("Planning Required", "No saved S-Curve planning found for this project.")
            keep_window_active(self)
            return

        popup = Toplevel(self)
        popup.title("Daily Progress Entry")
        popup.geometry("1280x860")
        popup.grab_set()
        popup.configure(bg="#f0f4f8")

        project_title = self.current_project_name or self.uid or ""
        Label(
            popup,
            text=f"Daily Progress Entry - {report_date_display}",
            font=("Arial", 18, "bold"),
            bg="#f0f4f8",
            fg="#003087",
        ).pack(pady=(15, 6))
        Label(
            popup,
            text=f"Project: {project_title}",
            font=("Arial", 12, "bold"),
            bg="#f0f4f8",
            fg="#003087",
        ).pack(pady=(0, 4))
        Label(
            popup,
            text=f"S-Curve Plan: {self.current_daily_plan}",
            font=("Arial", 11, "bold"),
            bg="#f0f4f8",
            fg="#666666",
        ).pack(pady=(0, 10))
        window_start, window_end = self.format_daily_entry_window()
        Label(
            popup,
            text=f"Allowed entry/edit date window: {window_start} to {window_end}",
            font=("Arial", 10, "bold"),
            bg="#f0f4f8",
            fg="#92400e",
        ).pack(pady=(0, 8))

        scroll_area = ScrollableFrame(popup, bg="#f0f4f8")
        scroll_area.pack(fill=BOTH, expand=True, padx=20, pady=10)
        content = Frame(scroll_area.scrollable_frame, bg="#f0f4f8")
        content.pack(anchor="nw", fill=X, expand=True)

        manpower_box = LabelFrame(
            content,
            text="Manpower Monitoring",
            bg="#f0f4f8",
            font=("Arial", 12, "bold"),
            fg="#003087",
            padx=12,
            pady=10,
        )
        manpower_box.pack(fill=X, pady=(0, 12))

        existing_values = existing_values or [report_date_display] + [0] * 12
        report_date_storage = to_storage_date(report_date_display)
        existing_record = get_daily_progress_by_date(self.current_project_id, report_date_storage)
        existing_manpower_rows = get_daily_progress_manpower(self.current_project_id, report_date_storage)

        rsp_defaults = {
            "rsp_executive": str((existing_record or {}).get("rsp_executive", existing_values[1] if len(existing_values) > 1 else 0) or 0),
            "rsp_non_executive": str((existing_record or {}).get("rsp_non_executive", existing_values[2] if len(existing_values) > 2 else 0) or 0),
            "staff_civil": "0",
            "staff_electrical": "0",
            "staff_mechanical": "0",
            "staff_refractory": "0",
        }
        agency_staff_defaults = []
        contractor_defaults = []
        contractor_map = {}
        for row in existing_manpower_rows:
            section_name = str(row.get("section_name") or "")
            category_name = str(row.get("category_name") or "")
            contractor_name = str(row.get("contractor_name") or "")
            role_name = str(row.get("role_name") or "")
            qty = int(row.get("qty") or 0)
            if section_name == "Rourkela Steel Plant Manpower":
                if category_name == "Executives":
                    rsp_defaults["rsp_executive"] = str(qty)
                elif category_name == "Non-Executives":
                    rsp_defaults["rsp_non_executive"] = str(qty)
            elif section_name == "Executing Agency" and category_name == "Staff / Supervisory":
                key = f"staff_{role_name.strip().lower()}"
                if key in rsp_defaults:
                    rsp_defaults[key] = str(qty)
                agency_staff_defaults.append({"role": role_name, "qty": str(qty)})
            elif section_name == "Executing Agency" and contractor_name:
                if contractor_name not in contractor_map:
                    contractor_map[contractor_name] = {"name": contractor_name, "supervisor": "0", "labour": "0"}
                if role_name.strip().lower() == "supervisor":
                    contractor_map[contractor_name]["supervisor"] = str(qty)
                elif role_name.strip().lower() == "labour":
                    contractor_map[contractor_name]["labour"] = str(qty)

        if contractor_map:
            contractor_defaults = list(contractor_map.values())
        else:
            contractor_defaults = [
                {"name": "Unistar", "supervisor": "0", "labour": "0"},
                {"name": "Sharda Co", "supervisor": "0", "labour": "0"},
            ]
        initial_rsp_defaults = dict(rsp_defaults)
        initial_contractor_defaults = [dict(row) for row in contractor_defaults]

        manpower_box.grid_columnconfigure(1, weight=1)
        manpower_box.grid_columnconfigure(3, weight=1)
        manpower_box.grid_columnconfigure(5, weight=1)

        Label(
            manpower_box,
            text="Rourkela Steel Plant Manpower",
            bg="#f0f4f8",
            fg="#003087",
            font=("Arial", 12, "bold"),
        ).grid(row=0, column=0, columnspan=6, sticky=W, pady=(0, 8))

        manpower_vars = {}
        rsp_fields = [
            ("Executives", "rsp_executive"),
            ("Non-Executives", "rsp_non_executive"),
        ]
        for idx, (label, key) in enumerate(rsp_fields):
            Label(manpower_box, text=label + ":", bg="#f0f4f8", font=("Arial", 11, "bold")).grid(
                row=1, column=idx * 2, padx=10, pady=6, sticky=W
            )
            var = StringVar(value=rsp_defaults[key])
            Entry(manpower_box, textvariable=var, width=16, font=("Arial", 12), justify="center").grid(
                row=1, column=idx * 2 + 1, padx=8, pady=6, sticky=W
            )
            manpower_vars[key] = var

        Label(
            manpower_box,
            text="Executing Agency",
            bg="#f0f4f8",
            fg="#003087",
            font=("Arial", 12, "bold"),
        ).grid(row=2, column=0, columnspan=6, sticky=W, pady=(12, 4))

        agency_header_row = 3
        Label(
            manpower_box,
            text="Staff / Supervisory",
            bg="#f0f4f8",
            fg="#333333",
            font=("Arial", 11, "bold"),
        ).grid(row=agency_header_row, column=0, columnspan=4, sticky=W, pady=(0, 4))

        agency_frame = Frame(manpower_box, bg="#f0f4f8")
        agency_frame.grid(row=agency_header_row + 1, column=0, columnspan=6, sticky="ew")
        agency_frame.grid_columnconfigure(0, weight=3)
        agency_frame.grid_columnconfigure(1, weight=1)

        agency_staff_rows = []

        def render_agency_staff_rows():
            for widget in agency_frame.winfo_children():
                widget.destroy()
            Label(agency_frame, text="Discipline", bg="#f0f4f8", font=("Arial", 11, "bold")).grid(row=0, column=0, padx=6, pady=4, sticky=W)
            Label(agency_frame, text="Qty", bg="#f0f4f8", font=("Arial", 11, "bold")).grid(row=0, column=1, padx=6, pady=4, sticky=W)
            Label(agency_frame, text="", bg="#f0f4f8").grid(row=0, column=2, padx=6, pady=4)

            for idx, row in enumerate(agency_staff_rows, start=1):
                Entry(agency_frame, textvariable=row["role_var"], width=26, font=("Arial", 11)).grid(row=idx, column=0, padx=6, pady=4, sticky="ew")
                Entry(agency_frame, textvariable=row["qty_var"], width=12, font=("Arial", 11), justify="center").grid(
                    row=idx, column=1, padx=6, pady=4, sticky=W
                )
                Button(
                    agency_frame,
                    text="✕",
                    command=lambda current=row: remove_agency_staff_row(current),
                    bg="#c8102e",
                    fg="white",
                    font=("Arial", 8, "bold"),
                    width=2,
                ).grid(row=idx, column=2, padx=6, pady=4)

        def add_agency_staff_row(role_name="", qty="0"):
            agency_staff_rows.append({"role_var": StringVar(value=role_name), "qty_var": StringVar(value=str(qty))})
            render_agency_staff_rows()

        def remove_agency_staff_row(current_row):
            if len(agency_staff_rows) <= 1:
                current_row["role_var"].set("")
                current_row["qty_var"].set("0")
            else:
                agency_staff_rows.remove(current_row)
            render_agency_staff_rows()

        agency_defaults = agency_staff_defaults
        if not agency_defaults:
            agency_defaults = []
            for role_key, role_label in (
                ("staff_civil", "Civil"),
                ("staff_electrical", "Electrical"),
                ("staff_mechanical", "Mechanical"),
                ("staff_refractory", "Refractory"),
            ):
                agency_defaults.append({"role": role_label, "qty": rsp_defaults.get(role_key, "0")})
        for agency_row in agency_defaults:
            add_agency_staff_row(agency_row["role"], agency_row["qty"])
        initial_agency_defaults = [dict(row) for row in agency_defaults]

        Button(
            manpower_box,
            text="+ Add Agency Type",
            command=lambda: add_agency_staff_row("", "0"),
            bg="#0066cc",
            fg="white",
            font=("Arial", 10, "bold"),
            width=18,
        ).grid(row=agency_header_row, column=4, columnspan=2, padx=8, pady=(0, 4), sticky=E)

        contractor_header_row = 5
        Label(
            manpower_box,
            text="Contractors Under Executing Agency",
            bg="#f0f4f8",
            fg="#333333",
            font=("Arial", 11, "bold"),
        ).grid(row=contractor_header_row, column=0, columnspan=4, sticky=W, pady=(12, 4))

        contractor_frame = Frame(manpower_box, bg="#f0f4f8")
        contractor_frame.grid(row=contractor_header_row + 1, column=0, columnspan=6, sticky="ew")
        contractor_frame.grid_columnconfigure(0, weight=3)
        contractor_frame.grid_columnconfigure(1, weight=1)
        contractor_frame.grid_columnconfigure(2, weight=1)

        contractor_rows = []

        def render_contractor_rows():
            for widget in contractor_frame.winfo_children():
                widget.destroy()

            Label(contractor_frame, text="Contractor", bg="#f0f4f8", font=("Arial", 11, "bold")).grid(row=0, column=0, padx=6, pady=4, sticky=W)
            Label(contractor_frame, text="Supervisor", bg="#f0f4f8", font=("Arial", 11, "bold")).grid(row=0, column=1, padx=6, pady=4, sticky=W)
            Label(contractor_frame, text="Labour", bg="#f0f4f8", font=("Arial", 11, "bold")).grid(row=0, column=2, padx=6, pady=4, sticky=W)
            Label(contractor_frame, text="", bg="#f0f4f8").grid(row=0, column=3, padx=6, pady=4)

            for idx, row in enumerate(contractor_rows, start=1):
                Entry(contractor_frame, textvariable=row["name_var"], width=28, font=("Arial", 11)).grid(row=idx, column=0, padx=6, pady=4, sticky="ew")
                Entry(contractor_frame, textvariable=row["supervisor_var"], width=12, font=("Arial", 11), justify="center").grid(
                    row=idx, column=1, padx=6, pady=4, sticky=W
                )
                Entry(contractor_frame, textvariable=row["labour_var"], width=12, font=("Arial", 11), justify="center").grid(
                    row=idx, column=2, padx=6, pady=4, sticky=W
                )
                Button(
                    contractor_frame,
                    text="✕",
                    command=lambda current=row: remove_contractor_row(current),
                    bg="#c8102e",
                    fg="white",
                    font=("Arial", 8, "bold"),
                    width=2,
                ).grid(row=idx, column=3, padx=6, pady=4)

        def add_contractor_row(name="", supervisor="0", labour="0"):
            contractor_rows.append({
                "name_var": StringVar(value=name),
                "supervisor_var": StringVar(value=str(supervisor)),
                "labour_var": StringVar(value=str(labour)),
            })
            render_contractor_rows()

        def remove_contractor_row(current_row):
            if len(contractor_rows) <= 1:
                current_row["name_var"].set("")
                current_row["supervisor_var"].set("0")
                current_row["labour_var"].set("0")
            else:
                contractor_rows.remove(current_row)
            render_contractor_rows()

        for contractor in contractor_defaults:
            add_contractor_row(contractor.get("name", ""), contractor.get("supervisor", "0"), contractor.get("labour", "0"))

        Button(
            manpower_box,
            text="➕ Add Contractor",
            command=lambda: add_contractor_row("", "0", "0"),
            bg="#0066cc",
            fg="white",
            font=("Arial", 10, "bold"),
            width=16,
        ).grid(row=contractor_header_row, column=4, columnspan=2, padx=8, pady=(10, 4), sticky=E)

        activity_box = LabelFrame(
            content,
            text="Activity Wise Daily Progress",
            bg="#f0f4f8",
            font=("Arial", 12, "bold"),
            fg="#003087",
            padx=10,
            pady=10,
        )
        activity_box.pack(fill=BOTH, expand=True)

        activity_table = ScrollableFrame(activity_box, bg="#f0f4f8")
        activity_table.pack(fill=BOTH, expand=True)
        activity_grid = activity_table.scrollable_frame
        activity_grid.configure(bg="#f0f4f8")
        headers = [
            ("Activity", 54, W),
            ("UOM", 10, CENTER),
            ("Scope Qty", 12, CENTER),
            ("Cumulative Actual", 16, CENTER),
            ("Remaining Qty", 14, CENTER),
            ("Actual Today", 12, CENTER),
        ]
        for col_idx, (header_text, width, anchor) in enumerate(headers):
            Label(
                activity_grid,
                text=header_text,
                bg="#e5e7eb",
                fg="#111827",
                font=("Arial", 10, "bold"),
                width=width,
                anchor=anchor,
                relief="solid",
                bd=1,
                padx=4,
                pady=4,
            ).grid(row=0, column=col_idx, sticky="nsew")

        activity_entry_rows = []
        activity_rows = get_activity_progress_rows(self.current_project_id, self.current_daily_plan, report_date_storage)
        if not activity_rows:
            Label(
                activity_box,
                text="No S-curve activities found for this project. Please save project activities in the S-curve window first.",
                bg="#f0f4f8",
                fg="#c8102e",
                font=("Arial", 11, "bold"),
            ).pack(pady=8)
        for row_idx, row in enumerate(activity_rows, start=1):
            scope_qty = float(row["scope_qty"] or 0)
            cumulative_actual = float(row["cumulative_actual"] or 0)
            remaining = scope_qty - cumulative_actual
            actual_today = float(row["actual_today"] or 0)
            actual_var = StringVar(value=f"{actual_today:.2f}" if actual_today else "")

            Label(activity_grid, text=str(row["activity_type"] or ""), bg="white", anchor=W, width=54, relief="solid", bd=1, padx=4, pady=4).grid(row=row_idx, column=0, sticky="nsew")
            Label(activity_grid, text=str(row["uom"] or ""), bg="white", anchor=CENTER, width=10, relief="solid", bd=1, padx=4, pady=4).grid(row=row_idx, column=1, sticky="nsew")
            Label(activity_grid, text=f"{scope_qty:.2f}", bg="white", anchor=CENTER, width=12, relief="solid", bd=1, padx=4, pady=4).grid(row=row_idx, column=2, sticky="nsew")
            Label(activity_grid, text=f"{cumulative_actual:.2f}", bg="white", anchor=CENTER, width=16, relief="solid", bd=1, padx=4, pady=4).grid(row=row_idx, column=3, sticky="nsew")
            Label(activity_grid, text=f"{remaining:.2f}", bg="white", anchor=CENTER, width=14, relief="solid", bd=1, padx=4, pady=4).grid(row=row_idx, column=4, sticky="nsew")
            Entry(activity_grid, textvariable=actual_var, justify="center", font=("Arial", 10), width=12).grid(
                row=row_idx, column=5, sticky="nsew", padx=1, pady=1
            )

            activity_entry_rows.append(
                {
                    "activity_id": int(row["id"]),
                    "activity_type": row["activity_type"],
                    "scope_qty": scope_qty,
                    "cumulative_actual": cumulative_actual,
                    "actual_today_initial": actual_today,
                    "actual_var": actual_var,
                }
            )

        def reset_daily_entry():
            for key, var in manpower_vars.items():
                var.set(initial_rsp_defaults.get(key, "0"))

            agency_staff_rows.clear()
            for agency_row in initial_agency_defaults:
                add_agency_staff_row(agency_row.get("role", ""), agency_row.get("qty", "0"))

            contractor_rows.clear()
            for contractor in initial_contractor_defaults:
                add_contractor_row(contractor.get("name", ""), contractor.get("supervisor", "0"), contractor.get("labour", "0"))

            for row in activity_entry_rows:
                initial_today = float(row.get("actual_today_initial") or 0)
                row["actual_var"].set(f"{initial_today:.2f}" if initial_today else "")
            keep_window_active(popup)

        def save_daily_entry():
            try:
                selected_date = datetime.strptime(report_date_display, "%d-%m-%y").date()
            except Exception:
                messagebox.showerror("Invalid Date", "Please use a valid Daily Progress date.")
                keep_window_active(popup)
                return

            earliest_allowed, latest_allowed = self.get_allowed_daily_entry_window()
            if selected_date < earliest_allowed or selected_date > latest_allowed:
                messagebox.showwarning(
                    "Date Restricted",
                    f"Daily Progress entry is allowed only from {earliest_allowed.strftime('%d-%m-%y')} to {latest_allowed.strftime('%d-%m-%y')}.",
                )
                keep_window_active(popup)
                return

            if not activity_entry_rows:
                messagebox.showwarning("No Activities", "No S-curve activities are available for the selected project.")
                keep_window_active(popup)
                return
            manpower_data = {}
            for key, var in manpower_vars.items():
                try:
                    manpower_data[key] = int(var.get() or 0)
                except Exception:
                    manpower_data[key] = 0
            agency_staff_payload = []
            for row in agency_staff_rows:
                role_name = str(row["role_var"].get() or "").strip()
                if not role_name:
                    continue
                try:
                    qty = int(row["qty_var"].get() or 0)
                except Exception:
                    qty = 0
                agency_staff_payload.append({"role": role_name, "qty": max(0, qty)})
            manpower_data["agency_staff"] = agency_staff_payload

            contractor_payload = []
            for row in contractor_rows:
                contractor_name = str(row["name_var"].get() or "").strip()
                if not contractor_name:
                    continue
                try:
                    supervisor_qty = int(row["supervisor_var"].get() or 0)
                except Exception:
                    supervisor_qty = 0
                try:
                    labour_qty = int(row["labour_var"].get() or 0)
                except Exception:
                    labour_qty = 0
                contractor_payload.append(
                    {
                        "name": contractor_name,
                        "supervisor": supervisor_qty,
                        "labour": labour_qty,
                    }
                )
            manpower_data["contractors"] = contractor_payload

            total_manpower = int(manpower_data.get("rsp_executive", 0) or 0) + int(manpower_data.get("rsp_non_executive", 0) or 0)
            total_manpower += sum(int(row.get("qty", 0) or 0) for row in agency_staff_payload)
            total_manpower += sum(int(row.get("supervisor", 0) or 0) + int(row.get("labour", 0) or 0) for row in contractor_payload)

            activity_payload = []
            total_actual_qty = 0.0
            for row in activity_entry_rows:
                try:
                    actual_qty = float(str(row["actual_var"].get() or "0").strip())
                except Exception:
                    actual_qty = 0.0
                cumulative_before = max(0.0, float(row["cumulative_actual"] or 0) - float(row["actual_today_initial"] or 0))
                if cumulative_before + actual_qty > float(row["scope_qty"] or 0):
                    messagebox.showerror(
                        "Scope Limit",
                        f"{row['activity_type']}: cumulative actual cannot exceed scope quantity.",
                    )
                    keep_window_active(popup)
                    return
                activity_payload.append(
                    {
                        "activity_id": row["activity_id"],
                        "activity_type": row["activity_type"],
                        "actual_qty": actual_qty,
                    }
                )
                total_actual_qty += actual_qty

            if total_manpower <= 0:
                messagebox.showwarning("Manpower Required", "Enter manpower details for the selected date before saving.")
                keep_window_active(popup)
                return

            if total_actual_qty <= 0:
                messagebox.showwarning("Physical Progress Required", "Enter physical progress quantities for the selected date before saving.")
                keep_window_active(popup)
                return

            save_daily_progress_with_activities(
                self.current_project_id,
                report_date_storage,
                manpower_data,
                activity_payload,
            )
            messagebox.showinfo("Success", "Daily progress saved successfully.")
            popup.destroy()
            self.reload_daily_progress_view()
            keep_window_active(self)

        btn_frame = Frame(popup, bg="#f0f4f8")
        btn_frame.pack(fill=X, pady=12)
        Button(
            btn_frame,
            text="Save Daily Progress",
            command=save_daily_entry,
            bg="#008000",
            fg="white",
            font=("Arial", 12, "bold"),
        ).pack(side=LEFT, padx=20)
        Button(
            btn_frame,
            text="Reset",
            command=reset_daily_entry,
            bg="#f59e0b",
            fg="white",
            font=("Arial", 12, "bold"),
        ).pack(side=LEFT, padx=10)
        Button(
            btn_frame,
            text="Close",
            command=popup.destroy,
            bg="#555555",
            fg="white",
            font=("Arial", 11, "bold"),
        ).pack(side=RIGHT, padx=20)
        normalize_buttons(popup)

    def add_new_day(self):
        if not self.can_add_daily_progress:
            messagebox.showwarning("Planning Incomplete", "Plan S-curve data first. Daily progress is enabled only after quantity planning is completed and saved.")
            return
        if self.main_app and not self.main_app.can_edit("daily_progress"):
            messagebox.showwarning("Edit Denied", "You have view access only for Daily Progress Report.")
            return

        popup = Toplevel(self)
        popup.title("Add New Daily Progress")
        popup.geometry("420x280")
        popup.grab_set()
        popup.configure(bg="#f0f4f8")

        Label(popup, text="Add New Daily Progress Entry", font=("Arial", 13, "bold"),
              bg="#f0f4f8", fg="#003087").pack(pady=15)

        f = Frame(popup, bg="#f0f4f8")
        f.pack(pady=10, padx=30)

        Label(f, text="Select Date (DD-MM-YY):", bg="#f0f4f8", font=("Arial", 11)).pack(anchor=W)
        date_var = StringVar(value=datetime.now().strftime("%d-%m-%y"))
        Entry(f, textvariable=date_var, width=22, font=("Arial", 11)).pack(pady=8, anchor=W)
        start_text, end_text = self.format_daily_entry_window()
        Label(
            f,
            text=f"Allowed: {start_text} to {end_text}",
            bg="#f0f4f8",
            fg="#92400e",
            font=("Arial", 10, "bold"),
        ).pack(anchor=W, pady=(0, 5))
        Button(f, text="Pick Date", command=lambda: self.pick_date(date_var, popup),
               bg="#0066cc", fg="white", font=("Arial", 10)).pack(pady=5, anchor=W)

        def create_new_day():
            try:
                selected_date = datetime.strptime(date_var.get(), "%d-%m-%y").date()
            except Exception:
                messagebox.showwarning("Invalid Date", "Please select a valid date in DD-MM-YY format.")
                keep_window_active(popup)
                return
            earliest_allowed, latest_allowed = self.get_allowed_daily_entry_window()
            if selected_date < earliest_allowed or selected_date > latest_allowed:
                messagebox.showwarning(
                    "Date Restricted",
                    f"Daily Progress entry is allowed only from {earliest_allowed.strftime('%d-%m-%y')} to {latest_allowed.strftime('%d-%m-%y')}.",
                )
                keep_window_active(popup)
                return
            popup.destroy()
            self.open_daily_entry_popup(date_var.get())
            keep_window_active(self)

        Button(popup, text="ADD THIS DAY", command=create_new_day,
               bg="#008000", fg="white", font=("Arial", 11, "bold"), height=2, width=22).pack(pady=15)
        normalize_buttons(popup)

    def pick_date(self, var, parent):
        cal_win = Toplevel(parent)
        cal_win.title("Select Date")
        cal_win.geometry("320x300")
        earliest_allowed, latest_allowed = self.get_allowed_daily_entry_window()
        cal = DateEntry(cal_win, width=25, date_pattern='dd-mm-yy', mindate=earliest_allowed, maxdate=latest_allowed)
        cal.pack(pady=20)
        def set_d():
            var.set(cal.get_date().strftime("%d-%m-%y"))
            cal_win.destroy()
        Button(cal_win, text="Select", command=set_d, bg="#003087", fg="white").pack(pady=10)
        normalize_buttons(cal_win)

    def go_home(self):
        self.destroy()
        if self.main_app:
            self.main_app.show_frame("registration")
# =================================================================================
