"""
reports.py
==========
Export layer for the Scheduling & Project Control Module (spec item 11).

This module turns the in-memory results produced by the core engines
(CPMResult, DelayReport, DCMAReport) into deliverable artefacts:

    * CSV   -- universal, opens anywhere, good for re-import
    * Excel -- formatted workbook (openpyxl), one sheet per report section
    * PDF   -- printable scorecards / summaries (reportlab)

The functions are deliberately pure: they take engine objects + a file
path (or return bytes) and never touch the database.  The API/service
layer is responsible for loading data, running the engines and handing
the results here.

Report types implemented (mapping to spec item 11):
    - project_schedule_summary
    - critical_path
    - baseline_variance      (via DelayReport)
    - delay_analysis         (via DelayReport)
    - dcma_compliance        (via DCMAReport)
    - look_ahead             (filtered CPM activities)
    - milestone_tracking     (filtered CPM activities)

Hindrance / risk impact reports are thin wrappers over tabular data and
are produced with the generic `rows_to_*` helpers.

openpyxl / reportlab are optional at import time: if a backend is not
installed the corresponding `*_excel` / `*_pdf` function raises a clear
ReportError, while CSV always works (stdlib only).
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date
from typing import Any, Iterable, Optional, Sequence

from .cpm import CPMResult, CPMActivity
from .delay_analysis import DelayReport
from .dcma import DCMAReport


class ReportError(Exception):
    pass


# ---------------------------------------------------------------------------
# generic tabular helpers
# ---------------------------------------------------------------------------
@dataclass
class Table:
    """A simple titled grid the exporters know how to render."""
    title: str
    columns: list[str]
    rows: list[list[Any]]
    note: str = ""


def _fmt(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, float):
        return f"{v:g}"
    return str(v)


# ---------------------------------------------------------------------------
# builders: engine result -> Table(s)
# ---------------------------------------------------------------------------
def schedule_summary_table(result: CPMResult) -> Table:
    cols = ["ID", "Name", "Dur(wd)", "ES", "EF", "LS", "LF",
            "TF", "FF", "%", "Critical", "Milestone"]
    rows: list[list[Any]] = []
    for a in _sorted_acts(result):
        rows.append([
            a.id, a.name, a.duration, a.es, a.ef, a.ls, a.lf,
            a.total_float, a.free_float, a.percent_complete,
            a.is_critical, a.is_milestone,
        ])
    note = (f"Project {result.project_start} -> {result.project_finish}"
            f"   |   data date: {result.data_date or '-'}"
            f"   |   {len(rows)} activities")
    return Table("Project Schedule Summary", cols, rows, note)


def critical_path_table(result: CPMResult) -> Table:
    cols = ["Seq", "ID", "Name", "Dur(wd)", "ES", "EF", "TF"]
    rows: list[list[Any]] = []
    for i, aid in enumerate(result.critical_path, start=1):
        a = result.activities.get(aid)
        if a is None:
            continue
        rows.append([i, a.id, a.name, a.duration, a.es, a.ef, a.total_float])
    note = f"{len(rows)} activities on the critical path"
    return Table("Critical Path", cols, rows, note)


def look_ahead_table(result: CPMResult, horizon_start: date,
                     horizon_end: date) -> Table:
    cols = ["ID", "Name", "ES", "EF", "TF", "Critical", "%"]
    rows: list[list[Any]] = []
    for a in _sorted_acts(result):
        if a.es is None:
            continue
        # any overlap of [es,ef] with the window
        ef = a.ef or a.es
        if ef < horizon_start or a.es > horizon_end:
            continue
        rows.append([a.id, a.name, a.es, a.ef, a.total_float,
                     a.is_critical, a.percent_complete])
    note = f"Look-ahead window {horizon_start} -> {horizon_end} ({len(rows)} activities)"
    return Table("Look-Ahead Schedule", cols, rows, note)


def milestone_table(result: CPMResult) -> Table:
    cols = ["ID", "Name", "Target (EF)", "LF", "TF", "Critical", "%"]
    rows: list[list[Any]] = []
    for a in _sorted_acts(result):
        if not a.is_milestone:
            continue
        rows.append([a.id, a.name, a.ef, a.lf, a.total_float,
                     a.is_critical, a.percent_complete])
    note = f"{len(rows)} milestones"
    return Table("Milestone Tracking", cols, rows, note)


def delay_table(report: DelayReport) -> Table:
    cols = ["ID", "Name", "BL Start", "BL Finish", "Cur Start", "Cur Finish",
            "Start Var(wd)", "Finish Var(wd)", "TF", "Critical",
            "Class", "Reason"]
    rows: list[list[Any]] = []
    for r in report.rows:
        rows.append([
            r.activity_id, r.name, r.bl_start, r.bl_finish,
            r.cur_start, r.cur_finish, r.start_var_wd, r.finish_var_wd,
            r.total_float, r.is_critical, r.classification.value, r.reason,
        ])
    note = (f"Project finish variance: {report.project_finish_variance_wd} wd"
            f"   |   delayed: {report.delayed_count}"
            f"   |   critical delays: {report.critical_delay_count}")
    return Table("Delay Analysis", cols, rows, note)


def baseline_variance_table(report: DelayReport) -> Table:
    """Baseline-variance view = subset of delay columns focused on dates/float."""
    cols = ["ID", "Name", "BL Start", "Cur Start", "Start Var(wd)",
            "BL Finish", "Cur Finish", "Finish Var(wd)", "TF"]
    rows: list[list[Any]] = []
    for r in report.rows:
        rows.append([
            r.activity_id, r.name, r.bl_start, r.cur_start, r.start_var_wd,
            r.bl_finish, r.cur_finish, r.finish_var_wd, r.total_float,
        ])
    return Table("Baseline Variance", cols, rows,
                 f"{len(rows)} activities compared to baseline")


def dcma_table(report: DCMAReport) -> Table:
    cols = ["#", "Check", "Metric", "Threshold", "Result",
            "Affected", "Total", "Observation", "Suggestion"]
    rows: list[list[Any]] = []
    for c in report.checks:
        rows.append([
            c.number, c.name, c.metric, c.threshold,
            "PASS" if c.passed else "FAIL",
            c.affected, c.total, c.observation, c.suggestion,
        ])
    note = (f"Score: {report.score:.0f}%  "
            f"({report.passed_count}/{report.applicable_count} applicable checks passed)")
    return Table("DCMA 14-Point Compliance", cols, rows, note)


def _sorted_acts(result: CPMResult) -> list[CPMActivity]:
    def key(a: CPMActivity):
        return (a.es or date.max, a.id)
    return sorted(result.activities.values(), key=key)


# ---------------------------------------------------------------------------
# CSV  (stdlib, always available)
# ---------------------------------------------------------------------------
def table_to_csv(table: Table) -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    if table.title:
        w.writerow([table.title])
    if table.note:
        w.writerow([table.note])
    w.writerow(table.columns)
    for row in table.rows:
        w.writerow([_fmt(v) for v in row])
    return buf.getvalue()


def tables_to_csv(tables: Sequence[Table]) -> str:
    return "\n".join(table_to_csv(t) for t in tables)


# ---------------------------------------------------------------------------
# Excel  (openpyxl)
# ---------------------------------------------------------------------------
def tables_to_excel(tables: Sequence[Table], path: str) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:  # pragma: no cover
        raise ReportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        ) from e

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="0E7C7B")     # brand teal
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=13, color="0A5C5B")
    note_font = Font(italic=True, size=9, color="555555")
    fail_fill = PatternFill("solid", fgColor="FBE3E3")
    crit_fill = PatternFill("solid", fgColor="FFF1E0")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    used_names: set[str] = set()
    for table in tables:
        name = _sheet_name(table.title, used_names)
        ws = wb.create_sheet(title=name)
        r = 1
        ws.cell(r, 1, table.title).font = title_font
        r += 1
        if table.note:
            ws.cell(r, 1, table.note).font = note_font
            r += 1
        r += 1  # blank spacer
        header_row = r
        for c, col in enumerate(table.columns, start=1):
            cell = ws.cell(header_row, c, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = border
        r += 1
        # detect special columns for conditional highlight
        cols_lower = [c.lower() for c in table.columns]
        result_idx = cols_lower.index("result") if "result" in cols_lower else None
        crit_idx = cols_lower.index("critical") if "critical" in cols_lower else None
        for row in table.rows:
            highlight = None
            if result_idx is not None and _fmt(row[result_idx]) == "FAIL":
                highlight = fail_fill
            elif crit_idx is not None and _fmt(row[crit_idx]) in ("Yes", "True"):
                highlight = crit_fill
            for c, val in enumerate(row, start=1):
                cell = ws.cell(r, c, _excel_val(val))
                cell.border = border
                if highlight:
                    cell.fill = highlight
            r += 1
        # column widths
        for c, col in enumerate(table.columns, start=1):
            width = max(len(col), *(len(_fmt(rw[c - 1])) for rw in table.rows)) \
                if table.rows else len(col)
            ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 8), 48)
        ws.freeze_panes = ws.cell(header_row + 1, 1)

    wb.save(path)
    return path


def _excel_val(v: Any):
    if isinstance(v, (int, float, date)):
        return v
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return _fmt(v)


def _sheet_name(title: str, used: set[str]) -> str:
    base = title[:28].replace(":", "").replace("/", "-").strip() or "Sheet"
    name = base
    i = 2
    while name in used:
        name = f"{base[:25]} {i}"
        i += 1
    used.add(name)
    return name


# ---------------------------------------------------------------------------
# PDF  (reportlab)
# ---------------------------------------------------------------------------
def tables_to_pdf(tables: Sequence[Table], path: str,
                  document_title: str = "Schedule Report") -> str:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table as RLTable,
                                        TableStyle, Paragraph, Spacer)
    except ImportError as e:  # pragma: no cover
        raise ReportError(
            "reportlab is required for PDF export. "
            "Install with: pip install reportlab"
        ) from e

    teal = colors.HexColor("#0E7C7B")
    teal_dark = colors.HexColor("#0A5C5B")
    fail_bg = colors.HexColor("#FBE3E3")
    crit_bg = colors.HexColor("#FFF1E0")

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], textColor=teal_dark,
                        fontSize=18, spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=teal,
                        fontSize=12, spaceBefore=10, spaceAfter=2)
    note_st = ParagraphStyle("note", parent=styles["Normal"], fontSize=8,
                             textColor=colors.HexColor("#555555"), spaceAfter=4)
    cell_st = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7,
                             leading=8)

    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    flow: list[Any] = [Paragraph(document_title, h1),
                       Paragraph(f"Generated {date.today().isoformat()}", note_st),
                       Spacer(1, 4)]

    for table in tables:
        flow.append(Paragraph(table.title, h2))
        if table.note:
            flow.append(Paragraph(table.note, note_st))
        data = [[Paragraph(str(c), cell_st) for c in table.columns]]
        for row in table.rows:
            data.append([Paragraph(_fmt(v), cell_st) for v in row])
        rl = RLTable(data, repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), teal),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F6FAFA")]),
        ]
        # conditional row shading
        cols_lower = [c.lower() for c in table.columns]
        result_idx = cols_lower.index("result") if "result" in cols_lower else None
        crit_idx = cols_lower.index("critical") if "critical" in cols_lower else None
        for ri, row in enumerate(table.rows, start=1):
            if result_idx is not None and _fmt(row[result_idx]) == "FAIL":
                style.append(("BACKGROUND", (0, ri), (-1, ri), fail_bg))
            elif crit_idx is not None and _fmt(row[crit_idx]) in ("Yes", "True"):
                style.append(("BACKGROUND", (0, ri), (-1, ri), crit_bg))
        rl.setStyle(TableStyle(style))
        flow.append(rl)
        flow.append(Spacer(1, 8))

    doc.build(flow)
    return path


# ---------------------------------------------------------------------------
# high-level convenience: assemble a full report pack
# ---------------------------------------------------------------------------
def build_report_pack(
    result: Optional[CPMResult] = None,
    delay: Optional[DelayReport] = None,
    dcma: Optional[DCMAReport] = None,
    look_ahead_window: Optional[tuple[date, date]] = None,
) -> list[Table]:
    """Collect whichever reports the caller has data for, in a sensible order."""
    tables: list[Table] = []
    if result is not None:
        tables.append(schedule_summary_table(result))
        tables.append(critical_path_table(result))
        if any(a.is_milestone for a in result.activities.values()):
            tables.append(milestone_table(result))
        if look_ahead_window:
            tables.append(look_ahead_table(result, *look_ahead_window))
    if delay is not None:
        tables.append(baseline_variance_table(delay))
        tables.append(delay_table(delay))
    if dcma is not None:
        tables.append(dcma_table(dcma))
    return tables
